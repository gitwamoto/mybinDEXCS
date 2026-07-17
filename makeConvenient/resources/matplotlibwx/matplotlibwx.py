#!/usr/bin/env python
# -*- coding: utf-8 -*-
# matplotlibwx.py
# by Yukiharu Iwamoto
# 2025/9/30 2:11:13 PM

# Macの場合，文字入力後に引用符が勝手に変わったりしてうまく動かない．
# 「システム環境設定」→「キーボード」→「ユーザー辞書」→「スマート引用符とスマートダッシュを使用」のチェックを外す．

version = '2025/9/30 2:11:13 PM'

import os
languages = os.environ.get('LANG')
#languages = ['en']
languages = ['ja']

# https://code-examples.net/ja/docs/matplotlib~2.1/index
# https://code-examples.net/ja/docs/matplotlib~3.0/index
# https://qiita.com/HajimeKawahara/items/03f29367744d5209be42
# https://qiita.com/qsnsr123/items/325d21621cfe9e553c17

import sys
import matplotlib
if sys.platform == 'darwin':
    import unicodedata
    if len(sys.argv) == 1:
        # reference https://qiita.com/ground0state/items/556e768c5fb96044ebcb
        matplotlib.use('WXAgg')
import matplotlib.pyplot as plt
import matplotlib.tri as tri
import matplotlib.ticker as ticker
import matplotlib.patches as patches
import matplotlib.lines as lines
import numpy as np
import math
import subprocess
import re
import ast
import xlrd
import openpyxl
import csv
from PIL import Image, ImageChops
import wx # pip install wxPython==4.1.0
import wx.grid
import requests
import zipfile
import gettext
import codecs
import platform
import pyperclip # pip install pyperclip
import webbrowser
import scipy

def decode_if_necessary(s):
    return s.decode('CP932' if sys.platform == 'win32' else 'UTF-8') if sys.version_info.major <= 2 and type(s) is str else s

def encode_if_necessary(s):
    return s.encode('CP932' if sys.platform == 'win32' else 'UTF-8') if sys.version_info.major <= 2 and type(s) is unicode else s

def eval_exc(expression, globals = None, locals = None):
    try:
         return eval(expression, globals, locals)
    except:
        print(sys.exc_info())
        return None

# How to make translation
# (1) mkdir -p locale/en/LC_MESSAGES
# (2) If you already have old locale/en/LC_MESSAGES/messages.po, rename it by following command:
#     mv locale/en/LC_MESSAGES/messages.po locale/en/LC_MESSAGES/messages.po.old
# (3) xgettext -o locale/messages.pot matplotlibwx.py && msginit --locale=en --input=locale/messages.pot --output-file=locale/en/LC_MESSAGES/messages.po
# (4) If you did not skip the step (2), merge files by following command:
#     msgmerge locale/en/LC_MESSAGES/messages.po.old locale/en/LC_MESSAGES/messages.po --output-file=locale/en/LC_MESSAGES/messages.po
#     rm locale/en/LC_MESSAGES/messages.po.old
# (5) Edit locale/en/LC_MESSAGES/messages.po
# (6) msgfmt -o locale/en/LC_MESSAGES/messages.mo locale/en/LC_MESSAGES/messages.po
#     - Reverse opetration can be done by 'msgunfmt'
gettext.translation(
    domain = 'messages',
    localedir = os.path.join(os.path.dirname(os.path.realpath(decode_if_necessary(__file__))), 'locale'),
    languages = languages,
    fallback = True
).install()

plt.rcParams['font.size'] = 14.0
plt.rcParams['axes.titlesize'] = 14.0
plt.rcParams['axes.labelsize'] = 14.0
plt.rcParams['axes.formatter.limits'] = [-3, 4] # [m, n] -> Scientific notation is used for data < 10^m or 10^n <= data
plt.rcParams['axes.formatter.use_mathtext'] = True # True -> 10^n, False 1e+n
plt.rcParams['axes.formatter.useoffset'] = False
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['xtick.major.size'] = 8.0
plt.rcParams['xtick.minor.size'] = 4.0
plt.rcParams['xtick.labelsize'] = 12.0
plt.rcParams['ytick.direction'] = 'in'
plt.rcParams['ytick.major.size'] = 8.0
plt.rcParams['ytick.minor.size'] = 4.0
plt.rcParams['ytick.labelsize'] = 12.0
plt.rcParams['legend.edgecolor'] = 'k'
plt.rcParams['legend.fontsize'] = 14.0
plt.rcParams['lines.markersize'] = 8.0
plt.rcParams['errorbar.capsize'] = 3.0
#print(plt.rcParams)

plt.cm.jet.set_over('black')
plt.cm.jet.set_under('darkgrey')
plt.cm.gray.set_over('black')
plt.cm.gray.set_under('white')
plt.cm.binary.set_over('white')
plt.cm.binary.set_under('black')
plt.cm.Reds.set_over('black')
plt.cm.Reds.set_under('darkgrey')
plt.cm.Blues.set_over('black')
plt.cm.Blues.set_under('darkgrey')

n_xyticks = 7
n_zticks = 10
zorder_max = 10000
zorder_min = 0
zorder_others = 100
zorder_scatter = 10
zorder_vector = 2
zorder_contour = 1
n_scatter = 5
markers = ('o', '^', 's', 'D', 'v', '>', '<', None) # str
markers_wx = (u'○', u'△', u'□', u'◇', u'▽', u'▷', u'◁', _(u'なし')) # unicode
line_styles = ('-', '--', '-.', ':', None) # str
line_styles_wx = (_(u'実線'), _(u'破線'), _(u'一点鎖線'), _(u'点線'), _(u'線なし')) # unicode
by_what_show_z = ('rainbow', 'wb', 'bw', 'wR', 'wB', 'size',
    'rainbow_and_size', 'wb_and_size', 'bw_and_size', 'wR_and_size', 'wB_and_size') # str
by_what_show_z_wx = (_(u'虹色'), _(u'白→黒'), _(u'黒→白'), _(u'白→赤'), _(u'白→青'), _(u'記号の大きさ'),
    _(u'虹色+記号の大きさ'), _(u'白→黒+記号の大きさ'), _(u'黒→白+記号の大きさ'), _(u'白→赤+記号の大きさ'),
    _(u'白→青+記号の大きさ')) # unicode
paint_styles = ('rainbow', 'wb', 'bw', 'wR', 'wB', None) # str
paint_styles_wx = (_(u'虹色'), _(u'白→黒'), _(u'黒→白'), _(u'白→赤'), _(u'白→青'), _(u'塗りなし')) # unicode

pat_math = re.compile(r'(?<![a-zA-Z0-9_.\s])\s*((?!(?:abs|max|min)\s*\()[a-z0-9]+\s*\(|(?:pi|e)(?![a-zA-Z0-9_.(]))')
pat_eq_plot = re.compile(r'(?:\s+|)(.+?)\s*,\s*([^,=]+?)\s*=\s*\[\s*(.+?)\s*,\s*(.+?)\s*\]\s*(L?/)\s*(.+?)\s*' +
                                      r'(?:,\s*([^,=]+?)\s*=\s*\[\s*(.+?)\s*,\s*(.+?)\s*\]\s*(L?/)\s*(.+?)\s*)?$')
pat_cell = re.compile(r'([0-9]+)\s*!\s*([a-zA-Z]+)\s*([0-9]+)')
pat_col = re.compile(r'\$([0-9]+)')

def get_file_from_google_drive(file_id):
    r = requests.get('https://drive.google.com/uc', params = {'export': 'download', 'id': file_id})
    if r.ok:
        if b'Google Drive - Virus scan warning' in r.content:
            cookies = r.cookies.get_dict()
            if cookies:
                for k in cookies.keys():
                    if k.startswith('download_warning_'):
                        code = cookies[k]
                        break
            else: # https://github.com/wkentaro/gdown/blob/1bf9e20442a0df57eec3e75a15ef4115dbec9b2f/gdown/download.py#L32
                m = re.search(b'id="downloadForm" action=".+?&amp;confirm=(.+?)"', r.content)
                if m:
                    code = m[1]
                else:
                    m = re.search(b'&amp;confirm=t&amp;uuid=(.+?)"', r.content)
                    if m:
                        code = m[1]
            r = requests.get('https://drive.google.com/uc',
                params = {'export': 'download', 'confirm': code, 'id': file_id}, cookies = cookies)
            if not r.ok:
                return None
        return r.content, r.apparent_encoding # type(r.content) = str (Python 2); bytes (Python 3)
    else:
        return None

def get_file_from_github_public(user, repository, branch, file_path):
    r = requests.get('https://raw.githubusercontent.com/' + user + '/' + repository + '/' + branch + '/' + file_path)
    if r.ok:
        return r.content, r.apparent_encoding # type(r.content) = str (Python 2); bytes (Python 3)
    else:
        return None

def naca_4digits_airfoil(digits, points = 51):
    # Example:
    # NACA 2 4 1 2'
    #      | |  |'
    #      | | thickness/chord = 0.12'
    #      | max. camber position = 0.4'
    #      max. camber = 0.02'
    x = np.linspace(0.0, 1.0, points)
    m = 0.01*float(digits[0])
    p = 0.1*float(digits[1])
    t = 0.01*float(digits[2:])
    yt = 5.0*t*(0.2969*np.sqrt(x) + (-0.126 + (-0.3516 + (0.2843 - 0.1015*x)*x)*x)*x)
    yc = np.where(x < p, m/p**2*(2.0*p - x)*x, m/(1.0 - p)**2*(1.0 - 2.0*p + (2.0*p - x)*x))
    theta = np.arctan(np.where(x < p, 2.0*m/p**2*(p - x), 2.0*m/(1.0 - p)**2*(p - x)))
    yts = yt*np.sin(theta)
    ytc = yt*np.cos(theta)
    yts[-1] = ytc[-1] = 0.0
    ps = np.hstack(((x + yts).reshape(-1, 1), (yc - ytc).reshape(-1, 1)))
    ss = np.hstack(((x - yts).reshape(-1, 1), (yc + ytc).reshape(-1, 1)))
    return np.vstack((ps, np.flip(ss[:-1], axis = 0)))

def time_str_a_is_newer_than_b(a, b):
    reg = re.compile(r'([0-9]+)\s*[/\-]\s*([0-9]+)\s*[/\-]\s*([0-9]+)\s+([0-9]+)\s*:\s*([0-9]+)\s*(?::\s*([0-9]+)\s*)?([AaPp][Mm])*')
    ra = reg.search(a)
    rb = reg.search(b)
    if ra is None or rb is None:
        return False
    time_a = [int(i) if i is not None else 0 for i in ra.groups()[:-1]]
    time_b = [int(i) if i is not None else 0 for i in rb.groups()[:-1]]
    # year, month, date, hour, minute, second, AM/PM
    if ra.groups()[-1] is not None and ra.groups()[-1].upper() == 'PM':
        time_a[3] += 12
    if rb.groups()[-1] is not None and rb.groups()[-1].upper() == 'PM':
        time_b[3] += 12
    return True if time_a > time_b else False

def alphabet_to_number(alphabet):
    # A = 1, B = 2, ..., Z = 26, AA = 27, AB = 28, ...
    alphabet = alphabet.upper()
    n = 0
    for i in alphabet:
        n = 26*n + ord(i) - ord('A') + 1
    return n

def number_to_alphabet(number):
    # 1 = A, 2 = B, ..., 26 = Z, 27 = AA, 28 = AB, ...
    number, mod = divmod(number - 1, 26)
    a = chr(ord('A') + mod)
    while number != 0:
        number, mod = divmod(number - 1, 26)
        a = chr(ord('A') + mod) + a
    return a

def appropriate_tick(xmin, xmax, n):
    tmp = abs(xmax - xmin)/(n + 0.01)
    tick = 10.0**math.floor(math.log10(tmp))
    tmp /= tick # 1 <= tmp < 10
    for i in (1.0, 2.0, 2.5, 5.0):
        if tmp < i:
            return i*tick
    return 10.0*tick

def correct_file_name_in_unicode(file_name):
    if sys.version_info.major <= 2 and type(file_name) is str:
        file_name = file_name.decode('UTF-8')
    if file_name == u'':
        return u''
    file_name = os.path.normpath(file_name.strip())
    if sys.platform == 'darwin':
        # 濁点なし文字と濁点に分離されている文字->濁点付きの文字
        file_name = unicodedata.normalize('NFC', file_name)
    elif sys.platform == 'win32':
        # os.path.normpath should be done prior to replace(os.sep, os.altsep)
        file_name = file_name.replace(os.sep, os.altsep)
    elif file_name.startswith(u'file:'):
        file_name = file_name[5:]
    return file_name # unicode

def clip_png(png_file_name):
    # http://pumpkinism113.hatenablog.com/entry/2018/05/16/221813
    png_file_name = correct_file_name_in_unicode(png_file_name) # unicode
    im = Image.open(png_file_name)
    im2 = im.convert('RGB')
    bg = Image.new('RGB', im2.size, im2.getpixel((0, 0)))
    im.crop(ImageChops.difference(im2, bg).getbbox()).save(png_file_name)

def save_and_clip_png(png_file_name, show = True):
    png_file_name = correct_file_name_in_unicode(png_file_name) # unicode
    plt.savefig(png_file_name)
    clip_png(png_file_name)
    if show:
        if sys.platform == 'win32':
            os.startfile(png_file_name)
        elif sys.platform == 'darwin':
            subprocess.call(u'open "{}"'.format(png_file_name), shell = True)
        else:
            subprocess.call(u'xdg-open "{}"'.format(png_file_name), shell = True)

def make_gif_from_pngs(png_file_lists, gif_file_name, duration_ms = 400, loop = 0, remove = False, show = True):
    gif_file_name = correct_file_name_in_unicode(gif_file_name) # unicode
    if not gif_file_name.endswith(u'.gif'):
        gif_file_name += u'.gif'
    images = [Image.open(correct_file_name_in_unicode(png_file)) for png_file in png_file_lists]
    images[0].save(gif_file_name, save_all = True, append_images = images[1:],
        duration = duration_ms, loop = loop)
    if remove:
        for i in png_file_lists:
            os.remove(i)
    if show:
        if sys.platform == 'win32':
            os.startfile(gif_file_name)
        elif sys.platform == 'darwin':
            subprocess.call(u'open "{}"'.format(gif_file_name), shell = True)
        else:
            subprocess.call(u'xdg-open "{}"'.format(gif_file_name), shell = True)

def data_from_file(file_name, columns = (1, 2), every = 1, skip = '#', delimiter = None, terminator = None, param_dict = None):
    # columns = (int | '$1 + 2', ...) for text file
    # columns = ('1!B3', ...) for excel sheet
    if delimiter is None:
        delimiter = r'\t| +\t?'
    if param_dict is not None:
        param_dict['math'] = globals()['math']
    file_name = correct_file_name_in_unicode(file_name) # unicode
    zip_file = False
    if file_name.endswith(u'.zip'):
        zip_file = True
        with zipfile.ZipFile(file_name) as zf:
            zf.extractall(os.path.dirname(file_name))
            file_name = os.path.join(os.path.dirname(file_name), zf.namelist()[0])
    if file_name.endswith(u'.csv'):
        columns1 = []
        for i in columns:
            i = pat_math.sub(r'math.\1', i)
            s = ['']
            while True:
                r = pat_cell.search(i)
                if r is not None:
                    s[-1] += i[:r.start()] + '('
                    s.append([alphabet_to_number(r[2]) - 1, int(r[3]) - 1])
                    s.append(')')
                    if r.end() == len(i):
                        break
                    i = i[r.end():]
                else:
                    s[-1] += i
                    break
            columns1.append(s)
        data = [[] for i in range(len(columns1))]
        try:
            with codecs.open(file_name, 'r', encoding = 'UTF-8') as f:
                lines = list(csv.reader(f))
        except:
            with codecs.open(file_name, 'r', encoding = 'CP932') as f:
                lines = list(csv.reader(f))
        n_lines = len(lines)
        for i in columns1:
            for j in i:
                if type(j) is list:
                    n_lines = min(n_lines, len(lines) - j[1])
        for n in range(n_lines):
            stop = has_None = False
            for i in range(len(columns1)):
                try:
                    s = u''
                    for j in range(len(columns1[i])):
                        if type(columns1[i][j]) is list:
                            v = lines[columns1[i][j][1]][columns1[i][j][0]].strip()
                            if sys.version_info.major <= 2:
                                try:
                                    v = v.decode('UTF-8') # unicode
                                except:
                                    v = v.decode('CP932') # unicode
                            if terminator is not None and terminator in v:
                                stop = True
                            else:
                                s += v
                                columns1[i][j][1] += 1
                        else:
                            s += columns1[i][j]
                    data[i].append(float(eval_exc(s, param_dict)))
                    if data[i][-1] is None:
                        has_None = True
                except:
                    data[i].append(np.nan)
            if stop or has_None:
                for i in range(len(data)):
                    del data[i][-1]
                if stop:
                    break
    elif file_name.endswith(u'.xls') or file_name.endswith(u'.xlsx') or file_name.endswith(u'.xlsm'):
        columns1 = []
        for i in columns:
            i = pat_math.sub(r'math.\1', i)
            s = ['']
            while True:
                r = pat_cell.search(i)
                if r is not None:
                    s[-1] += i[:r.start()] + '('
                    s.append([int(r[1]) - 1, alphabet_to_number(r[2]) - 1, int(r[3]) - 1])
                    s.append(')')
                    if r.end() == len(i):
                        break
                    i = i[r.end():]
                else:
                    s[-1] += i
                    break
            columns1.append(s)
        data = [[] for i in range(len(columns1))]
        if file_name.endswith(u'.xls'):
            wb = xlrd.open_workbook(file_name)
            lines = [[ws.row_values(i) for i in range(ws.nrows)] for ws in wb.sheets()]
            # https://stackoverflow.com/questions/33241837/python-xlrd-book-how-to-close-the-files
            wb.release_resources()
            del wb
        else: # .xlsx
            wb = openpyxl.load_workbook(file_name, data_only = True)
            lines = [[[v.value for v in row] for row in ws.rows] for ws in wb.worksheets]
            wb.close()
        n_lines = None
        for i in columns1:
            for j in i:
                if type(j) is list:
                    n_lines = len(lines[j[0]]) - j[2] if n_lines is None else min(n_lines, len(lines[j[0]]) - j[2])
        for n in range(n_lines):
            stop = has_None = False
            for i in range(len(columns1)):
                try:
                    s = u''
                    for j in range(len(columns1[i])):
                        if type(columns1[i][j]) is list:
                            v = lines[columns1[i][j][0]][columns1[i][j][2]][columns1[i][j][1]]
                            if sys.version_info.major <= 2:
                                if type(v) is not unicode:
                                    v = unicode(v)
                            else:
                                if type(v) is not str:
                                    v = str(v)
                            if terminator is not None and terminator in v:
                                stop = True
                            else:
                                s += v
                                columns1[i][j][2] += 1
                        else:
                            s += columns1[i][j]
                    data[i].append(float(eval_exc(s, param_dict)))
                    if data[i][-1] is None:
                        has_None = True
                except:
#                    print(sys.exc_info())
                    data[i].append(np.nan)
            if stop or has_None:
                for i in range(len(data)):
                    del data[i][-1]
                if stop:
                    break
    else:
        columns1 = []
        for i in columns:
            if type(i) is int:
                columns1.append(i - 1)
            else:
                i = pat_math.sub(r'math.\1', i)
                s = ['']
                while True:
                    r = pat_col.search(i)
                    if r is not None:
                        s[-1] += i[:r.start()] + '('
                        s.append(int(r[1]) - 1)
                        s.append(')')
                        if r.end() == len(i):
                            break
                        i = i[r.end():]
                    else:
                        s[-1] += i
                        break
                columns1.append(s)
        data = [[] for i in range(len(columns1))]
        for line in open(file_name, 'rb'):
            line = line.rstrip()
            try:
                line = line.decode('UTF-8') # unicode
            except:
                line = line.decode('CP932') # unicode
            if line == u'' or skip is not None and line.startswith(skip):
                continue
            line = re.split(delimiter, line)
            if len(line) < 2:
                raise ValueError('Inappropriate delimiter.')
            stop = has_None = False
            for i, j in enumerate(columns1):
                try:
                    if type(j) is int:
                        if terminator is not None and terminator in line[j]:
                            data[i].append(None)
                            stop = True
                        else:
                            data[i].append(float(line[j]))
                    else:
                        s = u''
                        for k in j:
                            if type(k) is int:
                                if terminator is not None and terminator in line[k]:
                                    stop = True
                                else:
                                    s += line[k]
                            else:
                                s += k
                        data[i].append(float(eval_exc(s, param_dict)))
                    if data[i][-1] is None:
                        has_None = True
                except:
                    data[i].append(np.nan)
            if stop or has_None:
                for i in range(len(data)):
                    del data[i][-1]
                if stop:
                    break
    if zip_file:
        os.remove(file_name)
    if every != 1:
        data = [data[i][::every] for i in range(len(data))]
    return np.array(data)

def data_from_equation(equation, param_dict = None):
    # examples of equation
    #  'sqrt(2.2*x), x = [0.0, 1.0]/100'
    #       1              2    3      4      5    6
    #    -> 'sqrt(2.2*x)', 'x', '0.0', '1.0', '/', '100'
    #  '[cos(t), sin(t)], t = [0.0, pi]/100'
    #       1                    2   3       4    5    6
    #    -> '[cos(t), sin(t)]', 't', '0.0', 'pi', '/', '100'
    #  'sqrt(x)*y**2, x = [-1.0, 1.0]/100, y = [-1.0, 1.0]/100'
    #       1               2    3       4      5    6      7    8       9      10   11
    #    -> 'sqrt(x)*y**2', 'x', '-1.0', '1.0', '/', '100', 'y', '-1.0', '1.0', '/', '100'
    #  '[cos(2.0*x), sin(y)], x = [-1.0, 1.0]/100, y = [-1.00, 1.0]/100'
    #       1                       2    3       4      5    6      7    8       9      10   11
    #    -> '[cos(2.0*x), sin(y)]', 'x', '-1.0', '1.0', '/', '100', 'y', '-1.0', '1.0', '/', '100']
    if param_dict is None:
        param_dict = {}
    param_dict['math'] = globals()['math']
    param_dict['scipy'] = globals()['scipy']
    equation = equation.replace('^', '**')
    r = pat_eq_plot.match(equation)
    eq = pat_math.sub(r'math.\1', r[1])

    x = r[2]
    x_val = param_dict[x] if x in param_dict else None
    x0 = float(eval_exc(pat_math.sub(r'math.\1', r[3]), param_dict))
    dx = float(eval_exc(pat_math.sub(r'math.\1', r[4]), param_dict))
    div_x = int(eval_exc(pat_math.sub(r'math.\1', r[6]), param_dict))
    if r[5] == '/':
        log_x = False
        dx = (dx - x0)/div_x
    else: # Log scale
        log_x = True
        dx = (dx/x0)**(1.0/div_x)
    data = [[], [], [], []]
    y = r[7]
    if y is None:
        for i in range(div_x + 1):
            vx = x0*dx**i if log_x else x0 + dx*i
            param_dict[x] = vx
            v = eval_exc(eq, param_dict)
            if isinstance(v, list):
                data[0].append(float(v[0]))
                data[1].append(float(v[1]))
            else:
                data[0].append(vx)
                data[1].append(float(v))
    else: # y is not None
        y_val = param_dict[y] if y in param_dict else None
        y0 = float(eval_exc(pat_math.sub(r'math.\1', r[8]), param_dict))
        dy = float(eval_exc(pat_math.sub(r'math.\1', r[9]), param_dict))
        div_y = int(eval_exc(pat_math.sub(r'math.\1', r[11]), param_dict))
        if r[10] == '/':
            log_y = False
            dy = (dy - y0)/div_y
        else: # Log scale
            log_y = True
            dy = (dy/y0)**(1.0/div_y)
        for j in range(div_y + 1):
            vy = y0*dy**j if log_y else y0 + dy*j
            param_dict[y] = vy
            for i in range(div_x + 1):
                vx = x0*dx**i if log_x else x0 + dx*i
                param_dict[x] = vx
                data[0].append(vx)
                data[1].append(vy)
                v = eval_exc(eq, param_dict)
                if isinstance(v, list):
                    data[2].append(float(v[0]))
                    data[3].append(float(v[1]))
                else:
                    data[2].append(float(v))
        if y_val is None:
            del param_dict[y]
        else:
            param_dict[y] = y_val
    if x_val is None:
        del param_dict[x]
    else:
        param_dict[x] = x_val
    while len(data[-1]) == 0:
        del data[-1]
    return np.array(data)

x_max = None
x_min = None
y_max = None
y_min = None

def figure_setting(fig_size = None, aspect = None, graph_edges = None, first = False, **kwargs):
    # fig_size = None | tuple | 'reflesh'
    # aspect = None | 'auto' | 'equal' | number
    # graph_edges = None | (left, bottom, right, top)
    if type(fig_size) is tuple:
        try:
            # new figure will appear if called, so do not call when you want to overwrite graphs
            plt.figure(figsize = fig_size)
        except:
            print(sys.exc_info())
    elif first:
        plt.clf()
    elif fig_size == 'reflesh':
        plt.cla()
    if aspect is not None:
        plt.gca().set_aspect(aspect, adjustable = 'box')
    if graph_edges is not None:
        plt.subplots_adjust(left = graph_edges[0], bottom = graph_edges[1], right = graph_edges[2], top = graph_edges[3])

def common_setting(data, ranges, ticks, log_scale, fig_size, aspect, graph_edges, title, labels, grids,
    title_size = plt.rcParams['axes.titlesize']):
    # fig_size = None | tuple | 'reflesh'
    # aspect = None | 'auto' | 'equal' | number
    # graph_edges = None | (left, bottom, right, top)
    figure_setting(fig_size = fig_size, aspect = aspect, graph_edges = graph_edges)
    plt.xscale('log' if log_scale[0] else 'linear')
    plt.yscale('log' if log_scale[1] else 'linear')
    plt.minorticks_on()
    try:
        x = data[0]
        if log_scale[0]:
            x = x[x > 0.0]
        if ranges[0] is None or ranges[0][0] is None:
            l, l_auto = np.nanmin(x), True
            global x_min
            if x_min is not None and x_min < l:
                l = x_min
            else:
                x_min = l
        else:
            l, l_auto = ranges[0][0], False
        if ranges[0] is None or ranges[0][1] is None:
            r, r_auto = np.nanmax(x), True
            global x_max
            if x_max is not None and x_max > r:
                r = x_max
            else:
                x_max = r
        else:
            r, r_auto = ranges[0][1], False
        if log_scale[0]:
            if l_auto:
                l = 10.0**math.floor(math.log10(l))
            if r_auto:
                r = 10.0**math.ceil(math.log10(r))
            plt.xlim(l, r)
            plt.gca().xaxis.minor.formatter.labelOnlyBase = True
            plt.gca().xaxis.set_major_locator(
                ticker.LogLocator(base = 10.0 if ticks[0] is None else ticks[0]))
        else:
            t = appropriate_tick(l, r, n_xyticks) if ticks[0] is None else ticks[0]
            if l_auto:
                l = math.floor(l/t)*t
            if r_auto:
                r = math.ceil(r/t)*t
            plt.xlim(l, r)
#            plt.gca().xaxis.set_major_formatter(ticker.FormatStrFormatter('%g'))
            plt.gca().xaxis.set_major_formatter(ticker.ScalarFormatter(useMathText = True))
            t = np.arange(l, r + 0.1*t, t)
            t[np.abs(t) < 1.0e-10*abs(r - l)] = 0.0
            plt.xticks(t)
    except:
        print(sys.exc_info())
    try:
        y = data[1]
        if log_scale[1]:
            y = y[y > 0.0]
        if ranges[1] is None or ranges[1][0] is None:
            l, l_auto = np.nanmin(y), True
            global y_min
            if y_min is not None and y_min < l:
                l = y_min
            else:
                y_min = l
        else:
            l, l_auto = ranges[1][0], False
        if ranges[1] is None or ranges[1][1] is None:
            r, r_auto = np.nanmax(y), True
            global y_max
            if y_max is not None and y_max > r:
                r = y_max
            else:
                y_max = r
        else:
            r, r_auto = ranges[1][1], False
        if log_scale[1]:
            if l_auto:
                l = 10.0**math.floor(math.log10(l))
            if r_auto:
                r = 10.0**math.ceil(math.log10(r))
            plt.ylim(l, r)
            plt.gca().yaxis.minor.formatter.labelOnlyBase = True
            plt.gca().yaxis.set_major_locator(
                ticker.LogLocator(base = 10.0 if ticks[1] is None else ticks[1]))
        else:
            t = appropriate_tick(l, r, n_xyticks) if ticks[1] is None else ticks[1]
            if l_auto:
                l = math.floor(l/t)*t
            if r_auto:
                r = math.ceil(r/t)*t
            plt.ylim(l, r)
#            plt.gca().yaxis.set_major_formatter(ticker.FormatStrFormatter('%g'))
            plt.gca().yaxis.set_major_formatter(ticker.ScalarFormatter(useMathText = True))
            t = np.arange(l, r + 0.1*t, t)
            t[np.abs(t) < 1.0e-10*abs(r - l)] = 0.0
            plt.yticks(t)
    except:
        print(sys.exc_info())
    if title is not None:
        plt.title(title, fontsize = title_size)
    if labels is not None:
        if labels[0] is not None:
            try:
                plt.xlabel(labels[0])
            except:
                print(sys.exc_info())
        if labels[1] is not None:
            try:
                plt.ylabel(labels[1])
            except:
                print(sys.exc_info())
    if grids is not None:
        if grids[0]:
            plt.gca().grid(which = 'major', axis = 'x', color = 'black', linestyle = '--', linewidth = 1.0)
            plt.gca().grid(which = 'minor', axis = 'x', color = 'black', linestyle = '--', linewidth = 0.5)
        if grids[1]:
            plt.gca().grid(which = 'major', axis = 'y', color = 'black', linestyle = '--', linewidth = 1.0)
            plt.gca().grid(which = 'minor', axis = 'y', color = 'black', linestyle = '--', linewidth = 0.5)

def insert_text(x, y, text, fontsize = plt.rcParams['font.size'], horizontal_alignment = 'left', coordinate = 'axes', zorder = 0):
    # coordinate = 'axes' | 'data'
    plt.text(x, y, text, fontsize = fontsize,
        horizontalalignment = horizontal_alignment,
        transform = plt.gca().transAxes if coordinate == 'axes' else plt.gca().transData, zorder = zorder)

def insert_arrow(x_from, y_from, x_to, y_to, line_width = 1.0, line_style = '-', head_width = 12.0, color = 'black',
    coordinate = 'axes', zorder = 0):
    # line_style = '-' | '--' | '-.' | ':'
    # coordinate = 'axes' | 'data'
    plt.arrow(x_from, y_from, x_to - x_from, y_to - y_from,
        width = 0, head_width = 0.002*head_width, length_includes_head = True,
        linewidth = line_width, linestyle = line_style, color = color,
        transform = plt.gca().transAxes if coordinate == 'axes' else plt.gca().transData, zorder = zorder)

def insert_line(x_from, y_from, x_to, y_to, line_width = 1.0, line_style = '-', color = 'black', coordinate = 'axes', zorder = 0):
    # line_style = '-' | '--' | '-.' | ':'
    # coordinate = 'axes' | 'data'
    plt.gca().add_artist(lines.Line2D([x_from, x_to], [y_from, y_to], linewidth = line_width, linestyle = line_style,
        color = color, transform = plt.gca().transAxes if coordinate == 'axes' else plt.gca().transData, zorder = zorder))

def insert_circle(x, y, width, radius, line_width = 1.0, fill = False, face_color = 'white', edge_color = 'black',
    coordinate = 'axes', zorder = 0):
    # coordinate = 'axes' | 'data'
    plt.gca().add_patch(patches.Circle((x, y), radius,
        lw = line_width, fill = fill, fc = face_color, ec = edge_color,
        transform = plt.gca().transAxes if coordinate == 'axes' else plt.gca().transData, zorder = zorder))

def insert_ellipse(x, y, width, height, angle = 0.0, line_width = 1.0, fill = False, face_color = 'white',
    edge_color = 'black', coordinate = 'axes', zorder = 0):
    # coordinate = 'axes' | 'data'
    plt.gca().add_patch(patches.Ellipse((x, y), width, height, angle = angle,
        lw = line_width, fill = fill, fc = face_color, ec = edge_color,
        transform = plt.gca().transAxes if coordinate == 'axes' else plt.gca().transData, zorder = zorder))

def insert_rectangle(x, y, width, height, angle = 0.0, line_width = 1.0, fill = False, face_color = 'white',
    edge_color = 'black', coordinate = 'axes', zorder = 0):
    # coordinate = 'axes' | 'data'
    plt.gca().add_patch(patches.Rectangle((x - 0.5*width, y - 0.5*height), width, height, angle = angle,
        lw = line_width, fill = fill, fc = face_color, ec = edge_color,
        transform = plt.gca().transAxes if coordinate == 'axes' else plt.gca().transData, zorder = zorder))

def insert_polygon(xy, line_width = 1.0, fill = False, face_color = 'white', edge_color = 'black',
    coordinate = 'axes', zorder = 0):
    # coordinate = 'axes' | 'data'
    plt.gca().add_patch(patches.Polygon(np.array(xy),
        lw = line_width, fill = fill, fc = face_color, ec = edge_color, closed = True,
        transform = plt.gca().transAxes if coordinate == 'axes' else plt.gca().transData, zorder = zorder))

def insert_naca_4digits_airfoil(digits, xle, yle, xte, yte, line_width = 1.0, fill = False, face_color = 'white',
    edge_color = 'black', coordinate = 'axes', zorder = 0, points = 51):
    # coordinate = 'axes' | 'data'
    chord = math.sqrt((xte - xle)**2 + (yte - yle)**2)
    c = (xte - xle)/chord
    s = (yte - yle)/chord
    xy = chord*naca_4digits_airfoil(digits, points)
    xy = np.hstack(((xle + xy[:, 0]*c - xy[:, 1]*s).reshape(-1, 1), (yle + xy[:, 0]*s + xy[:, 1]*c).reshape(-1, 1)))
    plt.gca().add_patch(patches.Polygon(np.array(xy),
        lw = line_width, fill = fill, fc = face_color, ec = edge_color, closed = True,
        transform = plt.gca().transAxes if coordinate == 'axes' else plt.gca().transData, zorder = zorder))

def plot_scatter(data, err = (None, None), ranges = (None, None), ticks = (None, None), log_scale = (False, False),
    fig_size = None, aspect = 'auto', graph_edges = None, title = None, labels = None, grids = (False, False),
    marker = 'o', marker_size = None, marker_color = 'white', marker_edge_color = 'red',
    line_style = None, line_color = 'red', label_in_legend = None, legend_location = None,
    show_z_by = 'rainbow', marker_size_ratio = None, zorder = 0):
    # fig_size = None | tuple | 'reflesh'
    # aspect = None | 'auto' | 'equal' | number
    # graph_edges = None | (left, bottom, right, top)
    # marker =  'o' | '^' | 's' | 'D' | 'v' |'<' | '>' and so on
    # marker_edge_color can take 'none'.
    # line_style = None | '-' | '--' | '-.' | ':'
    # legend_location = None | (left, top)
    # show_z_by = {'rainbow' | 'wb' | 'bw' | 'wR' | 'wB'} + {'_and_size' | ''} | 'size'
    common_setting(data, ranges, ticks, log_scale, fig_size, aspect, graph_edges, title, labels, grids)
    if err[0] is not None or err[1] is not None:
        plt.errorbar(data[0], data[1], xerr = err[0], yerr = err[1], linestyle = 'None', color = line_color, zorder = zorder)
    if line_style is not None:
        plt.plot(data[0], data[1], linestyle = line_style, color = line_color,
            label = label_in_legend if marker is None else None, zorder = zorder)
    if marker is not None:
        try:
            if isinstance(marker_color, (tuple, list)) and not isinstance(marker_color[0], (tuple, list)):
                marker_color = (marker_color, )
        except:
            pass
        if len(data) == 2 or show_z_by is None:
            plt.scatter(data[0], data[1], marker = marker, s = marker_size,
                c = marker_color, edgecolors = marker_edge_color, label = label_in_legend, zorder = zorder)
        else:
            try:
                try:
                    vmin = ranges[2][0]
                except:
                    vmin = None
                try:
                    vmax = ranges[2][1]
                except:
                    vmax = None
                try:
                    ls = False if log_scale[2] is None else log_scale[2]
                except:
                    ls = False
                z = data[2]
                if ls:
                    z = np.log(z)
                paint = show_z_by.replace('_and_', '').replace('size', '')
                if 'size' in show_z_by:
                    if marker_size is None:
                        marker_size = plt.rcParams['lines.markersize']**2
                    if marker_size_ratio is None:
                        marker_size_ratio = 4.0
                    smin = np.nanmin(z) if vmin is None else vmin
                    smax = np.nanmax(z) if vmax is None else vmax
                    if paint != '':
                        if vmin is None or vmax is None:
                            sc = plt.scatter(data[0], data[1], marker = marker,
                                s = marker_size + marker_size*(marker_size_ratio - 1.0)/(smax - smin)*(z - smin), c = z,
                                edgecolors = marker_edge_color, label = label_in_legend, vmin = vmin, vmax = vmax,
                                cmap = plt.get_cmap('jet' if paint == 'rainbow' else ('binary' if paint == 'wb' else
                                ('gray' if paint == 'bw' else ('Reds' if paint == 'wR' else 'Blues')))),
                                norm = matplotlib.colors.LogNorm() if ls else matplotlib.colors.Normalize(),
                                zorder = zorder)
                        else:
                            sc = plt.scatter(data[0], data[1], marker = marker,
                                s = marker_size + marker_size*(marker_size_ratio - 1.0)/(smax - smin)*(z - smin), c = z,
                                edgecolors = marker_edge_color, label = label_in_legend, vmin = vmin, vmax = vmax,
                                cmap = plt.get_cmap('jet' if paint == 'rainbow' else ('binary' if paint == 'wb' else
                                ('gray' if paint == 'bw' else ('Reds' if paint == 'wR' else 'Blues')))),
                                zorder = zorder)
                        cbar = plt.colorbar(sc)
                    else:
                        sc = plt.scatter(data[0], data[1], marker = marker,
                            s = marker_size + marker_size*(marker_size_ratio - 1.0)/(smax - smin)*(z - smin),
                            c = marker_color, edgecolors = marker_edge_color, label = label_in_legend, zorder = zorder)
                elif paint != '':
                    if vmin is None or vmax is None:
                        sc = plt.scatter(data[0], data[1], marker = marker, s = marker_size, c = z,
                            edgecolors = marker_edge_color, label = label_in_legend, vmin = vmin, vmax = vmax,
                            cmap = plt.get_cmap('jet' if paint == 'rainbow' else ('binary' if paint == 'wb' else
                            ('gray' if paint == 'bw' else ('Reds' if paint == 'wR' else 'Blues')))),
                            norm = matplotlib.colors.LogNorm() if ls else matplotlib.colors.Normalize(),
                            zorder = zorder)
                    else:
                        sc = plt.scatter(data[0], data[1], marker = marker, s = marker_size, c = z,
                            edgecolors = marker_edge_color, label = label_in_legend, vmin = vmin, vmax = vmax,
                            cmap = plt.get_cmap('jet' if paint == 'rainbow' else ('binary' if paint == 'wb' else
                            ('gray' if paint == 'bw' else ('Reds' if paint == 'wR' else 'Blues')))),
                            zorder = zorder)
                    cbar = plt.colorbar(sc)
            except:
                print(sys.exc_info())
            try:
                cbar.ax.set_ylabel(labels[2])
            except:
                pass
    if label_in_legend is not None:
        if legend_location is not None:
            plt.legend(bbox_to_anchor = legend_location, loc = 'upper left', borderaxespad = 0)
        else:
            plt.legend()

def plot_vector(data, ranges = (None, None, None), ticks = (None, None, None), log_scale = (False, False, False),
    fig_size = None, aspect = 'auto', graph_edges = None, title = None, labels = None, grids = (False, False),
    color = 'black', scale = 1.0, legend = (0.8, 1.03, 1.0, None), zorder = 0):
    # fig_size = None | tuple | 'reflesh'
    # aspect = None | 'auto' | 'equal' | number
    # graph_edges = None | (left, bottom, right, top)
    # color = 'len' | 'len1' | color name
    # legend = (x, y, u, 'label') | None
    common_setting(data, ranges, ticks, log_scale, fig_size, aspect, graph_edges, title, labels, grids)
    if color == 'len' or color == 'len1' or len(data) > 4:
        if color == 'len':
            length = np.hypot(data[2], data[3])
            quiv = plt.quiver(data[0], data[1], data[2], data[3], length, angles = 'xy',
                scale_units = 'xy', scale = 1.0/scale, cmap = plt.cm.jet, zorder = zorder,
                norm = matplotlib.colors.LogNorm() if log_scale[2] else matplotlib.colors.Normalize())
        elif color == 'len1':
            length = np.hypot(data[2], data[3])
            quiv = plt.quiver(data[0], data[1], data[2]/length, data[3]/length, length, angles = 'xy',
                scale_units = 'xy', scale = 1.0/scale, cmap = plt.cm.jet, zorder = zorder,
                norm = matplotlib.colors.LogNorm() if log_scale[2] else matplotlib.colors.Normalize())
        else:
            length = data[4]
            quiv = plt.quiver(data[0], data[1], data[2], data[3], length, angles = 'xy',
                scale_units = 'xy', scale = 1.0/scale, cmap = plt.cm.jet, zorder = zorder,
                norm = matplotlib.colors.LogNorm() if log_scale[2] else matplotlib.colors.Normalize())
        try:
            if log_scale[2]:
                length = length[length > 0.0]
            if ranges[2] is None or ranges[2][0] is None:
                l, l_auto = np.nanmin(length), True
            else:
                l, l_auto = ranges[2][0], False
            if ranges[2] is None or ranges[2][1] is None:
                r, r_auto = np.nanmax(length), True
            else:
                r, r_auto = ranges[2][1], False
            if log_scale[2]:
                log10_l = math.log10(l)
                log10_r = math.log10(r)
                if l_auto:
                    l = 10.0**math.floor(log10_l)
                    log10_l = math.log10(l)
                if r_auto:
                    r = 10.0**math.ceil(log10_r)
                    log10_r = math.log10(r)
                plt.clim(l, r) # plt.clim should be called last
                cbar = plt.colorbar(extend = 'both',
                    ticks = ticker.LogLocator(base = 10.0 if ticks[2] is None else ticks[2]),
                    format = ticker.LogFormatterMathtext(labelOnlyBase = True))
                t = np.array([[i + j for j in np.log10(np.arange(1, 9))]
                    for i in range(int(math.floor(log10_l)), int(math.ceil(log10_r)))])
                t.reshape(-1)
                t = (t - log10_l)/(log10_r - log10_l)
                t = t[t > 0.0]
                t = t[t < 1.0]
                cbar.ax.yaxis.set_minor_locator(ticker.FixedLocator(t))
            else:
                t = appropriate_tick(l, r, n_zticks) if ticks[2] is None else ticks[2]
                if l_auto:
                    l = math.floor(l/t)*t
                if r_auto:
                    r = math.ceil(r/t)*t
                plt.clim(l, r) # plt.clim should be called last
                t = np.arange(l, r + 0.1*t, t)
                t[np.abs(t) < 1.0e-10*abs(r - l)] = 0.0
#                cbar = plt.colorbar(ticks = t, extend = 'both', format = ticker.FormatStrFormatter('%g'))
                cbar = plt.colorbar(ticks = t, extend = 'both', format = ticker.ScalarFormatter(useMathText = True))
                t = appropriate_tick(t[0], t[1], 5)
                cbar.ax.yaxis.set_minor_locator(ticker.FixedLocator(
                    np.arange(t, r - l - 0.1*t, t)/(r - l) if sys.version_info.major <= 2 else np.arange(t, r - l - 0.1*t, t)))
        except:
            print(sys.exc_info())
        if labels is not None and labels[2] is not None:
            try:
                cbar.ax.set_ylabel(labels[2])
            except:
                print(sys.exc_info())
    else:
        quiv = plt.quiver(data[0], data[1], data[2], data[3], angles = 'xy', scale_units = 'xy',
            scale = 1.0/scale, color = color, zorder = zorder)
    if legend is not None and sys.version_info.major > 2:
        # In the latest version of matplotlib (2.2.5) for Python 2, Quiverkey has a bug.
        # "The scale_units='xy' case is not being handled correctly by Quiverkey,"
        # stated in https://github.com/matplotlib/matplotlib/issues/13616
        plt.quiverkey(quiv, X = legend[0], Y = legend[1], U = legend[2],
            label = ('' if legend[2] == 1.0 else '{:g}'.format(legend[2])) + (legend[3] if legend[3] is not None else ''),
            coordinates = 'axes', labelpos = 'E')

def plot_contour(data, ranges = (None, None, None), ticks = (None, None, None), log_scale = (False, False, False),
    fig_size = None, aspect = 'auto', graph_edges = None, title = None, labels = None, grids = (False, False),
    paint = 'rainbow', smooth_paint = True, structured_grid = None, show_triangle = False, zorder = 0):
    # fig_size = None | tuple | 'reflesh'
    # aspect = None | 'auto' | 'equal' | number
    # graph_edges = None | (left, bottom, right, top)
    # paint = 'rainbow' | 'wb' | 'bw' | 'wR' | 'wB' | None
    # structured_grid = (row, col) | (-1, col) | (row, -1) | (-1, -1) | None
    common_setting(data, ranges, ticks, log_scale, fig_size, aspect, graph_edges, title, labels, grids)
    try:
        if structured_grid is None:
            triang = tri.Triangulation(data[0], data[1])
            triang.set_mask(np.any(np.isnan(data[2, triang.triangles]), axis = 1))
        elif data.ndim != 3:
            if structured_grid[0] == -1 and structured_grid[1] == -1:
                dx0, dy0 = data[0][1] - data[0][0], data[1][1] - data[1][0]
                for i in range(1, data.size):
                    dx1, dy1 = data[0][i + 1] - data[0][i], data[1][i + 1] - data[1][i]
                    # 0.001 [rad] = 0.05729577951308232 [deg]
                    if math.fabs(math.atan2(dx0*dy1 - dy0*dx1, dx0*dx1 + dy0*dy1)) > 0.001:
                        structured_grid = (-1, i + 1)
                        break
            data = data.reshape(3, structured_grid[0], structured_grid[1])
        z = data[2]
        if log_scale[2]:
            z = z[z > 0.0]
        if ranges[2] is None or ranges[2][0] is None:
            l, l_auto = np.nanmin(z), True
        else:
            l, l_auto = ranges[2][0], False
        if ranges[2] is None or ranges[2][1] is None:
            r, r_auto = np.nanmax(z), True
        else:
            r, r_auto = ranges[2][1], False
        if paint is not None:
            cmap = plt.get_cmap('jet' if paint == 'rainbow' else ('binary' if paint == 'wb' else
                ('gray' if paint == 'bw' else ('Reds' if paint == 'wR' else 'Blues'))))
        if log_scale[2]:
            log10_l = math.log10(l)
            log10_r = math.log10(r)
            if l_auto:
                l = 10.0**math.floor(log10_l)
                log10_l = math.log10(l)
            if r_auto:
                r = 10.0**math.ceil(log10_r)
                log10_r = math.log10(r)
            locator = ticker.LogLocator(base = 10.0 if ticks[2] is None else ticks[2])
            levels = 10.0**np.arange(log10_l, log10_r + 0.1)
            if structured_grid is None:
                if show_triangle:
                    plt.triplot(triang, color = 'dimgray', linewidth = 0.5, zorder = zorder + 0.2)
                plt.tricontour(triang, data[2], levels, colors = 'black', linewidths = 0.7, locator = locator,
                    zorder = zorder + 0.1)
            else:
                plt.contour(data[0], data[1], data[2], levels, colors = 'black', linewidths = 0.7, locator = locator,
                    zorder = zorder + 0.1)
            if paint is not None:
                if smooth_paint:
                    levels = np.logspace(log10_l, log10_r, 256)
                if structured_grid is None:
                    # tricontourf shuld be called after tricontour
                    plt.tricontourf(triang, data[2], levels, cmap = cmap, locator = locator, zorder = zorder,
                        norm = matplotlib.colors.LogNorm()) # extend kwarg does not work yet with log scale
                else:
                    plt.contourf(data[0], data[1], data[2], levels, cmap = cmap, locator = locator, zorder = zorder,
                        norm = matplotlib.colors.LogNorm()) # extend kwarg does not work yet with log scale ???
                plt.clim(l, r) # plt.clim should be called last
                cbar = plt.colorbar(ticks = locator, format = ticker.LogFormatterMathtext(labelOnlyBase = True))
                t = np.array([[i + j for j in np.log10(np.arange(1, 9))]
                    for i in range(int(math.floor(log10_l)), int(math.ceil(log10_r)))])
                t.reshape(-1)
                t = (t - log10_l)/(log10_r - log10_l)
                t = t[t > 0.0]
                t = t[t < 1.0]
                cbar.ax.yaxis.set_minor_locator(ticker.FixedLocator(t))
        else:
            t = appropriate_tick(l, r, n_zticks) if ticks[2] is None else ticks[2]
            if l_auto:
                l = math.floor(l/t)*t
            if r_auto:
                r = math.ceil(r/t)*t
            t = np.arange(l, r + 0.1*t, t)
            t[np.abs(t) < 1.0e-10*abs(r - l)] = 0.0
            if structured_grid is None:
                if show_triangle:
                    plt.triplot(triang, color = 'dimgray', linewidth = 0.5, zorder = zorder + 0.2)
                plt.tricontour(triang, data[2], t, colors = 'black', linewidths = 0.7, zorder = zorder + 0.1)
            else:
                plt.contour(data[0], data[1], data[2], t, colors = 'black', linewidths = 0.7, zorder = zorder + 0.1)
            if paint is not None:
                if structured_grid is None:
                    # tricontourf shuld be called after tricontour
                    plt.tricontourf(triang, data[2], np.linspace(l, r, 256) if smooth_paint else t,
                        cmap = cmap, extend = 'both', zorder = zorder, norm = matplotlib.colors.Normalize())
                else:
                    plt.contourf(data[0], data[1], data[2], np.linspace(l, r, 256) if smooth_paint else t,
                        cmap = cmap, extend = 'both', zorder = zorder, norm = matplotlib.colors.Normalize())
                plt.clim(l, r) # plt.clim should be called last
                cbar = plt.colorbar(ticks = t, extend = 'both', format = ticker.ScalarFormatter(useMathText = True))
                t = appropriate_tick(t[0], t[1], 5)
                cbar.ax.yaxis.set_minor_locator(ticker.FixedLocator(
                    np.arange(t, r - l - 0.1*t, t)/(r - l) if sys.version_info.major <= 2 else np.arange(t, r - l - 0.1*t, t)))
    except:
        print(sys.exc_info())
    try:
        if paint is not None and labels is not None and labels[2] is not None:
            cbar.ax.set_ylabel(labels[2])
    except:
        print(sys.exc_info())

def plot(dict_lists, show = True):
    global x_max, x_min, y_max, y_min
    x_max = x_min = y_max = y_min = None
    figure_setting(first = True, **dict_lists[0])
    dict_lists[0]['fig_size'] = dict_lists[0]['aspect'] = dict_lists[0]['graph_edges'] = None
    if 'texts' in dict_lists[0] and dict_lists[0]['texts'] is not None:
        for i in dict_lists[0]['texts']:
            insert_text(coordinate = i[0], x = i[1], y = i[2], horizontal_alignment = i[3],
                text = i[4], fontsize = i[5], zorder = i[6])
    if 'arrows' in dict_lists[0] and dict_lists[0]['arrows'] is not None:
        for i in dict_lists[0]['arrows']:
            if i[7] > 0.0:
                insert_arrow(coordinate = i[0], x_from = i[1], y_from = i[2], x_to = i[3], y_to = i[4],
                    line_width = i[5], line_style = i[6], head_width = i[7], zorder = i[8])
            else:
                insert_line(coordinate = i[0], x_from = i[1], y_from = i[2], x_to = i[3], y_to = i[4],
                    line_width = i[5], line_style = i[6], zorder = i[8])
    if 'circles' in dict_lists[0] and dict_lists[0]['circles'] is not None:
        for i in dict_lists[0]['circles']:
            insert_circle(coordinate = i[0], x = i[1], y = i[2], radius = i[3], line_width = i[4],
                fill = i[5], zorder = i[6])
    if 'ellipses' in dict_lists[0] and dict_lists[0]['ellipses'] is not None:
        for i in dict_lists[0]['ellipses']:
            insert_ellipse(coordinate = i[0], x = i[1], y = i[2], width = i[3], height = i[4],
                angle = i[5], line_width = i[6], fill = i[7], zorder = i[8])
    if 'rectangles' in dict_lists[0] and dict_lists[0]['rectangles'] is not None:
        for i in dict_lists[0]['rectangles']:
            insert_rectangle(coordinate = i[0], x = i[1], y = i[2], width = i[3], height = i[4],
                angle = i[5], line_width = i[6], fill = i[7], zorder = i[8])
    if 'polygons' in dict_lists[0] and dict_lists[0]['polygons'] is not None:
        for i in dict_lists[0]['polygons']:
            insert_polygon(coordinate = i[0], xy = i[1], line_width = i[2], fill = i[3], zorder = i[4])
    if 'naca4' in dict_lists[0] and dict_lists[0]['naca4'] is not None:
        for i in dict_lists[0]['naca4']:
            insert_naca_4digits_airfoil(digits = i[0], coordinate = i[1], xle = i[2], yle = i[3],
                xte = i[4], yte = i[5], line_width = i[6], fill = i[7], zorder = i[8])
    for i in dict_lists[1:]:
        try:
            if 'file_name' in i:
                d = {'file_name': i['file_name'], 'columns': i['columns'], 'every': i['every']}
                if 'err_columns' in i and i['err_columns'] is not None:
                    i['err_columns'] = list(i['err_columns'])
                    for j, k in enumerate(i['err_columns']):
                        if type(k) is tuple:
                            k = [l for l in sorted(set(k), key = k.index) if l is not None]
                            if len(k) == 0:
                                i['err_columns'][j] = None
                            elif len(k) == 1:
                                i['err_columns'][j] = k[0]
                            elif i['err_columns'][j] != k:
                                i['err_columns'][j] = k
                            d['columns'] += tuple(k)
                        elif k is not None:
                            d['columns'] += (k,)
                has_err = False if i['columns'] == d['columns'] else True
                for j in ('skip', 'delimiter', 'terminator', 'param_dict'):
                    if j in dict_lists[0]:
                        d[j] = dict_lists[0][j]
                d = {'data': data_from_file(**d)}
                if has_err:
                    d['err'] = []
                    e = len(i['columns'])
                    for j in i['err_columns']:
                        if type(j) is list: # i['err_columns'] = list(i['err_columns'])
                            d['err'].append([d['data'][e], d['data'][e + 1]])
                            e += 2
                        elif j is not None:
                            d['err'].append(d['data'][e])
                            e += 1
                        else:
                            d['err'].append(None)
                    d['data'] = d['data'][:len(i['columns'])]
            else:
                d = {'equation': i['equation']}
                if 'param_dict' in dict_lists[0]:
                    d['param_dict'] = dict_lists[0]['param_dict']
                d = {'data': data_from_equation(**d)}
            for j in ('ranges', 'ticks', 'log_scale', 'fig_size', 'aspect', 'graph_edges', 'title', 'labels', 'grids'):
                if j in dict_lists[0]:
                    d[j] = dict_lists[0][j]
        except:
            print(sys.exc_info())
        if i['type'] == 'scatter':
            try:
                for j in plot_scatter.__code__.co_varnames[:plot_scatter.__code__.co_argcount]:
                    if j in i:
                        d[j] = i[j]
            except:
                print(sys.exc_info())
            if 'legend_location' in dict_lists[0]:
                d['legend_location'] = dict_lists[0]['legend_location']
            plot_scatter(**d)
        elif i['type'] == 'vector':
            try:
                for j in plot_vector.__code__.co_varnames[:plot_vector.__code__.co_argcount]:
                    if j in i:
                        d[j] = i[j]
            except:
                print(sys.exc_info())
            plot_vector(**d)
        else: # contour
            try:
                for j in plot_contour.__code__.co_varnames[:plot_contour.__code__.co_argcount]:
                    if j in i:
                        d[j] = i[j]
            except:
                print(sys.exc_info())
            plot_contour(**d)
    if 'png_file_name' in dict_lists[0] and dict_lists[0]['png_file_name'] is not None:
        save_and_clip_png(dict_lists[0]['png_file_name'], show = show)
    elif show:
        if sys.platform == 'win32':
            with wx.MessageDialog(None,
                _(u'この後に現れる（Figure 1というタイトルを持つ）グラフウインドウは，ウインドウが閉じられるまで処理をブロックします！\n' +
                  u'グラフを描きなおしたい場合は，その前に必ずグラフウインドウを閉じて下さい．'),
                _(u'処理をブロックします！'), style = wx.ICON_INFORMATION) as md:
                md.ShowModal()
        plt.show()

def evaluate_plot_settings(s, dir_name = u'.'):
    s = ast.literal_eval('[' + s + '\n]') # '\n]'の'\n'は最終行がコメントの時に必要
    dir_name = correct_file_name_in_unicode(dir_name) # unicode
    for i in range(len(s)):
        for j in s[i].keys():
            if sys.version_info.major <= 2 and type(s[i][j]) is str:
                s[i][j] = s[i].pop(j).decode('UTF-8') # unicode
        if 'png_file_name' in s[i] and s[i]['png_file_name'] is not None:
            s[i]['png_file_name'] = os.path.normpath(os.path.join(dir_name, s[i]['png_file_name']))
        elif 'file_name' in s[i] and s[i]['file_name'] is not None:
            s[i]['file_name'] = os.path.normpath(os.path.join(dir_name, s[i]['file_name']))
        if i == 0:
            continue
        try:
            if s[i]['type'] == 'global':
                s[0], s[i] = s[i], s[0]
            elif s[i]['type'] == 'contour':
                if 'grid_pattern' in s[i]:
                    s[i]['structured_grid'] = (-1, -1) if s[i]['grid_pattern'] else None
                    del s[i]['grid_pattern']
            elif s[i]['type'] not in ('global', 'scatter', 'vector'):
                print("Unknown value '%s' for key 'type'." % s[i]['type'])
        except:
            print("No key 'type' in dictionary.")
            quit()
    if s[0]['type'] != 'global':
        print("No value 'global' for key 'type'.")
        quit()
    if 'texts' in s[0] and s[0]['texts'] is not None:
        if type(s[0]['texts'][0]) is not tuple:
            s[0]['texts'] = (s[0]['texts'],)
        l = []
        for i in s[0]['texts']:
            if len(i) < 6: # サイズ指定がない頃のサポート
                l.append(('axes', i[0], i[1],
                    'right' if i[2].startswith('r') else ('center' if i[2].startswith('c') else 'left'),
                    i[3].decode('UTF-8') if sys.version_info.major <= 2 else i[3],
                    plt.rcParams['font.size'], i[4] if len(i) > 4 else zorder_others))
            elif len(i) < 7: # 座標系指定がない頃のサポート
                l.append(('axes', i[0], i[1],
                    'right' if i[2].startswith('r') else ('center' if i[2].startswith('c') else 'left'),
                    i[3].decode('UTF-8') if sys.version_info.major <= 2 else i[3], i[4], i[5]))
            else:
                l.append((i[0], i[1], i[2],
                    'right' if i[3].startswith('r') else ('center' if i[3].startswith('c') else 'left'),
                    i[4].decode('UTF-8') if sys.version_info.major <= 2 else i[4], i[5], i[6]))
        s[0]['texts'] = tuple(l)
    if 'arrows' in s[0] and s[0]['arrows'] is not None:
        if type(s[0]['arrows'][0]) is not tuple:
            s[0]['arrows'] = (s[0]['arrows'],)
        l = []
        for k in s[0]['arrows']:
            if len(k) < 7: # zorderがない頃のサポート
                l.append(tuple(['axes'] + list(k[:5]) + ['-', k[5], zorder_others]))
            elif len(k) < 8: # 座標系指定がない頃のサポート
                l.append(tuple(['axes'] + list(k[:5]) + ['-'] + list(k[5:])))
            elif len(k) < 9: # 線種指定がない頃のサポート
                l.append(tuple(list(k[:6]) + ['-'] + list(k[6:])))
            else:
                l.append(k)
        s[0]['arrows'] = tuple(l)
    for i, j in (('circles', 7), ('ellipses', 9), ('rectangles', 9)):
        if i in s[0] and s[0][i] is not None:
            if type(s[0][i][0]) is not tuple:
                s[0][i] = (s[0][i],)
            l = []
            for k in s[0][i]:
                if len(k) < j - 1: # zorderがない頃のサポート
                    l.append(tuple(['axes'] + list(k) + [zorder_others]))
                elif len(k) < j: # 座標系指定がない頃のサポート
                    l.append(tuple(['axes'] + list(k)))
                else:
                    l.append(k)
            s[0][i] = tuple(l)
    if 'polygons' in s[0] and s[0]['polygons'] is not None:
        if type(s[0]['polygons'][0]) is not tuple:
            s[0]['polygons'] = (s[0]['polygons'],)
        l = []
        for i in s[0]['polygons']:
            try:
                if len(i) < 4: # zorderがない頃のサポート
                    if type(i[0][0]) is list:
                        l.append(tuple(['axes'] + list(i) + [zorder_others]))
                    else:
                        l.append(('axes', [[i[0][j], i[0][j + 1]] for j in range(0, len(i[0]), 2)], i[1], i[2], zorder_others))
                elif len(i) < 5: # 座標系指定がない頃のサポート
                    if type(i[0][0]) is list:
                        l.append(tuple(['axes'] + list(i)))
                    else:
                        l.append(('axes', [[i[0][j], i[0][j + 1]] for j in range(0, len(i[0]), 2)], i[1], i[2], i[3]))
                else:
                    if type(i[0][0]) is list:
                        l.append(i)
                    else:
                        l.append((i[0], [[i[1][j], i[1][j + 1]] for j in range(0, len(i[1]), 2)], i[2], i[3], i[4]))
            except:
                print(sys.exc_info())
        s[0]['polygons'] = tuple(l)
    if 'naca4' in s[0] and s[0]['naca4'] is not None:
        if type(s[0]['naca4'][0]) is not tuple:
            s[0][i] = (s[0]['naca4'],)
        l = []
        for i in s[0]['naca4']:
            l.append(i)
        s[0]['naca4'] = tuple(l)
    if 'param_dict' in s[0] and s[0]['param_dict'] is not None:
        for i in list(s[0]['param_dict']):
            s[0]['param_dict'][i.decode('UTF-8') if sys.version_info.major <= 2 else i] = s[0]['param_dict'].pop(i)
    zorder_max = zorder_min = s[1]['zorder']
    for i in range(1, len(s)):
        zorder_max = max(zorder_max, s[i]['zorder'])
        zorder_min = min(zorder_min, s[i]['zorder'])
    s[0]['zorder_max'] = zorder_max
    s[0]['zorder_min'] = zorder_min
    return s

def load_plot_settings(file_name):
    file_name = correct_file_name_in_unicode(file_name) # unicode
    with codecs.open(file_name, 'r', encoding = 'UTF-8') as f:
        s = evaluate_plot_settings(f.read(), os.path.dirname(file_name))
    return s, file_name

def make_script(s, file_name):
    file_name = correct_file_name_in_unicode(file_name) # unicode
    with codecs.open(file_name, 'w', encoding = 'UTF-8') as f:
        f.write(u'#!/usr/bin/env python\n# -*- coding: utf-8 -*-\n\n')
        f.write(u'import sys\n')
        f.write(u"if sys.platform == 'darwin':\n")
        f.write(u'    import unicodedata\n')
        f.write(u'import os\n')
        f.write(u'def decode_if_necessary(s):\n')
        f.write(u"    return s.decode('CP932' if sys.platform == 'win32' else 'UTF-8') if sys.version_info.major <= 2 else s\n")
        f.write(u"cwd = os.path.dirname(os.path.realpath(decode_if_necessary(__file__)))\n")
        f.write(u"if sys.platform == 'darwin':\n")
        f.write(u'    # 濁点なし文字と濁点に分離されている文字->濁点付きの文字\n')
        f.write(u"    cwd = unicodedata.normalize('NFC', cwd)\n")
        f.write(u"elif sys.platform == 'win32':\n")
        f.write(u'    cwd = cwd.replace(os.sep, os.altsep)\n')
        f.write(u'os.chdir(cwd) # 作業フォルダをこのファイルがあるフォルダにする\n\n')
        import_file = correct_file_name_in_unicode(os.path.realpath(decode_if_necessary(__file__)))
        f.write(u"##### %sがある場所，必要に応じて書き換える #####\n" % os.path.basename(import_file))
        f.write(u"sys.path.append(u'%s')\n" % os.path.dirname(import_file))
        module_name = os.path.splitext(os.path.basename(import_file))[0] # unicode
        f.write(u'import %s\n' % module_name)
        f.write(u"\nif __name__ == '__main__':\n")
        f.write(u'    s = (\n        r"%s"\n    )\n' % s.strip().replace(u'\n', u'" +\n        r"'))
        f.write(u'    # プロット，画像を表示させたい場合はshow = Trueにする\n')
        f.write(u"    %s.plot(%s.evaluate_plot_settings(s, u'.'), show = False)\n" % (module_name, module_name))
        f.write(u'    # gifアニメーションを作る\n')
        f.write('#    ' + module_name + '.make_gif_from_pngs(\n' +
            u"#        png_file_lists = [u'a1.png', u'a2.png', ...], # 元データであるpngファイルのリスト，要変更\n" +
            u"#        gif_file_name = u'a.gif', # 作られるgifアニメーションファイル名，要変更\n" +
            u'#        duration_ms = 400, # 1コマあたりの時間 [ms]，必要に応じて変更\n' +
            u'#        loop = 0, # 繰り返し回数，必要に応じて変更\n' +
            u'#        remove = False, # 元データであるpngファイルを消したいときはTrueにする\n' +
            u'#        show = False # gifアニメーションを表示させたい場合はTrueにする\n' +
            u'#    )\n')

def make_template(file_name = u'template.txt'):
    file_name = correct_file_name_in_unicode(file_name) # unicode
    if not file_name.endswith(u'.txt'):
        file_name += u'.txt'
    s = os.path.splitext(os.path.basename(file_name))[0]
    with codecs.open(file_name, 'w', encoding = 'UTF-8') as f:
        f.write(u"{\n")
        f.write(u"# 全体の設定\n")
        f.write(u"    'type': 'global',\n")
        f.write(u"# 'png_file_name': 'figure.png'\n" +
            u"#     → figure.png という名前のpng画像ファイルにグラフを保存する，このファイルからの相対パス\n")
        f.write(u"# 'png_file_name': None # → 画像ファイルを保存せず，matplotlibのビューワーにグラフを表示する\n")
        f.write(u"    'png_file_name': '%s_example.png',\n" % s)
        f.write(u"# 'fig_size': None # → 図の大きさを指定しない\n")
        f.write(u"# 'fig_size': (4, 4) # → 図の大きさ縦4インチ，横4インチにする\n")
        f.write(u"    'fig_size': None,\n")
        f.write(u"# データファイル中で読み飛ばす行の先頭文字\n")
        f.write(u"    'skip': '#',\n")
        f.write(u"# 'delimiter': None # → 半角スペースかタブ区切りでデータを読み込む\n")
        f.write(u"# 'delimiter': ',' # → コンマ区切りでデータを読み込む\n")
        f.write(u"    'delimiter': None,\n")
        f.write(u"# 'terminator': None # → 最後の行まで読み込む\n")
        f.write(u"# 'terminator': '!!STOP' # → !!STOPという文字が含まれる行が出てきたら読み込み終了\n")
        f.write(u"    'terminator': '!!STOP',\n")
        f.write(u"# 'ranges': (None, (-1, 2), (-3, 8))\n" +
            u"#     → x軸の範囲を指定せず，y, z軸の範囲をそれぞれ-1 ≤ y ≤ 2, -3 ≤ z ≤ 8にする\n")
        f.write(u"    'ranges': (None, None, None),\n")
        f.write(u"# 'ticks': (None, 0.5, 1) # → x軸の刻みを指定せず，y, z軸の刻みをそれぞれ0.5, 1にする\n")
        f.write(u"    'ticks': (None, None, None),\n")
        f.write(u"# 'log_scale': (False, True, True) # → y, z軸を対数軸にする\n")
        f.write(u"    'log_scale': (False, False, False),\n")
        f.write(u"# x, y, z軸のラベル，$\\mathsf{ と }$ の間に書くとTeX形式になる\n")
        f.write(u"    'labels': (r'x', r'y', r'z'),\n")
        f.write(u"# x, y軸の目盛線，(True, False) → x軸の目盛線を長く描き，y軸は刻みだけ描く\n")
        f.write(u"    'grids': (False, False),\n")
        f.write(u"# 'aspect': 'auto' # → グラフの縦横比を特に指定しない\n")
        f.write(u"# 'aspect': 'equal' # → x, y軸方向に1進む長さを等しくする\n")
        f.write(u"# 'aspect': 3.0 # → y軸方向に1進む長さをx軸のそれの3.0倍にする\n")
        f.write(u"    'aspect': 'equal',\n")
        f.write(u"# 'title': 'graph for $\\mathsf{a_{xy}^2}$'\n" +
            u"#     → グラフの上に書く題名，$\\mathsf{ と }$ の間に書くとTeX形式になる\n")
        f.write(u"# 'title': None # → グラフの上に題名を書かない\n")
        f.write(u"    'title': None,\n")
        f.write(u"# 'legend_location': (1, 1) # → グラフ左/下端を0，右/上端を1とした座標に対して，凡例の左上を(1, 1)に指定する\n")
        f.write(u"# 'legend_location': (1, None) # → グラフ左/下端を0，右/上端を1とした座標に対して，凡例の左をグラフの右端に指定する\n")
        f.write(u"# 'legend_location': None # → 凡例の場所を指定しない\n")
        f.write(u"    'legend_location': None,\n")
        f.write(u"# 'graph_edges': (0.1, 0.2, 0.8, 0.9)\n" +
            u"#     → 画面左/下端を0，右/上端を1とした座標に対して，グラフの左，下，右，上端をそれぞれ0.1, 0.2, 0.8, 0.9に指定する\n")
        f.write(u"# 'graph_edges': (None, None, 0.7, None)\n" +
            u"#     → 画面左/下端を0，右/上端を1とした座標に対して，グラフの右端を0.7に指定する\n")
        f.write(u"# 'graph_edges': None # → グラフの端を指定しない\n")
        f.write(u"    'graph_edges': None,\n")
        f.write(u"# 'texts': (('axes', 0.1, 0.8, 'l', r'$\\mathsf{a/b}$', " +
            u"{:g}, {:d}),\n".format(plt.rcParams['font.size'], zorder_others) +
            u"#           ('data', 0.5, 0.1, 'c', r'exp.', {:g}, {:d}),\n".format(
                plt.rcParams['font.size'], zorder_others) +
            u"#           ('data', 0.7, 0.9, 'r', r'theory', {:g}, {:d}),)\n".format(
                plt.rcParams['font.size'] + 2.0, zorder_others + 1) +
            u"#     → axes座標で(0.1, 0.8), data座標で(0.5, 0.1), data座標で(0.7, 0.9)の場所に\n"
            u"#       それぞれ左揃え，中央揃え，右揃え，zorder = {:d}, {:d}, {:d}，".format(
                zorder_others, zorder_others, zorder_others + 1) +
            u"サイズ{:g}, {:g}, {:g}の文字を書く\n".format(
                plt.rcParams['font.size'], plt.rcParams['font.size'], plt.rcParams['font.size'] + 2.0) +
            u"#     * axes座標はグラフ枠の左下を(0, 0), 右上を(1, 1)にとる座標，data座標はふつうのx, y座標\n")
        f.write(u"# 'texts': None # 書く文字はない\n")
        f.write(u"    'texts': (('axes', 0.05, 0.9, 'l', r'test', {:g}, {:d}),),\n".format(
            plt.rcParams['font.size'], zorder_others))
        f.write(u"# 'arrows': (('axes', 0.1, 0.2, 0.4, 0.3, 2.0, '-', 5.0, {:d}),\n".format(zorder_others) +
            u"#            ('data', 0.1, 0.3, 0.4, 0.4, 1.0, '--', 0.0, {:d}),)\n".format(zorder_others + 1) +
            u"#     → axes座標で(0.1, 0.2)から(0.4, 0.3)に線幅2.0，実線，矢の幅5.0，" +
            u"zorder = {:d}の矢印，\n".format(zorder_others) +
            u"#       data座標で(0.1, 0.3)から(0.4, 0.4)に線幅1.0，破線，矢の幅0.0，" +
            u"zorder = {:d}の矢印(=直線)を描く\n".format(zorder_others + 1) +
            u"#     * axes座標はグラフ枠の左下を(0, 0), 右上を(1, 1)にとる座標，data座標はふつうのx, y座標\n" +
            u"#     * 線種の指定方法は散布図のline_styleと同じ\n")
        f.write(u"# 'arrows': None # 描く矢印はない\n")
        f.write(u"    'arrows': (('axes', 0.1, 0.2, 0.4, 0.3, 2.0, ':', 24.0, {:d}),),\n".format(zorder_others))
        f.write(u"# 'ellipses': (('data', 0.1, 0.2, 0.6, 0.4, 20.0, 1.0, True, {:d}),)\n".format(zorder_others) +
            u"#     → data座標で(0.1, 0.2)に中心があり，data座標で幅0.6，高さ0.4で20.0度反時計回りに回転させた線幅1.0，\n" +
            u"#       塗りつぶしあり，zorder = {:d}の楕円を描く\n".format(zorder_others) +
            u"#     * axes座標はグラフ枠の左下を(0, 0), 右上を(1, 1)にとる座標，data座標はふつうのx, y座標\n")
        f.write(u"# 'ellipses': None # 描く楕円はない\n")
        f.write(u"    'ellipses': (('axes', 0.8, 0.4, 0.2, 0.1, 45.0, 1.0, False, {:d}),),\n".format(zorder_others))
        f.write(u"# 'polygons': (('data', [0.1, 0.1, 0.3, 0.1, 0.2, 0.4], 1.0, False, {:d}),)\n".format(zorder_others) +
            u"#     → data座標で(0.1, 0.1), (0.3, 0.1), (0.2, 0.4)に頂点があり，線幅1.0，塗りつぶしなし，" +
            u"zorder = {:d}の多角形を描く\n".format(zorder_others) +
            u"#     * axes座標はグラフ枠の左下を(0, 0), 右上を(1, 1)にとる座標，data座標はふつうのx, y座標\n")
        f.write(u"# 'polygons': None # 描く多角形はない\n")
        f.write(u"    'polygons': (('axes', [0.2, 0.2, 0.2, 0.4, 0.3, 0.3], 1.5, True, {:d}),),\n".format(zorder_others))
        f.write(u"# 'naca4': (('2412', 'axes', 0.1, 0.1, 0.3, 0.1, 1.0, True, {:d}),)\n".format(zorder_others) +
            u"#     → axes座標で前縁，後縁がそれぞれ(0.1, 0.1), (0.3, 0.1)にあり，線幅1.0，塗りつぶしあり，\n" +
            u"#       zorder = {:d}のNACA2412翼を描く\n".format(zorder_others) +
            u"#     * axes座標はグラフ枠の左下を(0, 0), 右上を(1, 1)にとる座標，data座標はふつうのx, y座標\n")
        f.write(u"# 'naca4': None # 描くNACA4桁系列翼はない\n")
        f.write(u"    'naca4': (('4415', 'data', -0.6, 0.4, -0.1, 0.1, 1.2, True, {:d}),),\n".format(zorder_others))
        f.write(u"# 'param_dict': {'a': 10.0, 'b': 2.2}\n" +
            u"#     → ファイル読み取り列指定を数式で行う際に，パラメータa, bをそれぞれ10.0, 2.2として使う\n")
        f.write(u"# 'param_dict': None # 使うパラメータはない\n")
        f.write(u"    'param_dict': None,\n")
        f.write(u"},\n")
        f.write(u"# '以下の散布図，ベクトル線図，等高線図の辞書（{から}まで）を必要な数だけ作る\n")
        f.write(u"{\n")
        f.write(u"# 散布図\n")
        f.write(u"    'type': 'scatter',\n")
        f.write(u"# {:d} ≤ zorder ≤ {:d}，zorderが大きいほど前面に描く\n".format(zorder_min, zorder_max))
        f.write(u"    'zorder': 10,\n")
        f.write(u"# データファイル名，このファイルからの相対パス\n")
        f.write(u"    'file_name': '%s_example1.txt',\n" % s)
        f.write(u"# <テキストファイルの場合>\n")
        f.write(u"#  'columns': ('$1/$2*sqrt($3)', '4')\n" +
            u"#     → 1列目の値/2列目の値*3列目の値の平方根，4列目の値をそれぞれx, y軸のデータに使う\n")
        f.write(u"#  'columns': ('1', '2', '5')\n" +
            u"#     → 【記号の色や大きさでz軸の値を表したい時】1列目，2列目，5列目の値をそれぞれx, y, z軸のデータに使う\n")
        f.write(u"# <エクセルファイル（拡張子がxlsまたはxlsx）またはcsvファイルの場合>\n")
        f.write(u"#  'columns': ('1!A2', 'sqrt(1!B2)')\n" +
            u"#     → 1枚目のシートのセルA2から下の値，B2から下の値の平方根をそれぞれx, y軸のデータに使う\n")
        f.write(u"    'columns': (1, 2),\n")
        f.write(u"# 'err_columns': ((3, 4), (5, 6))\n" +
            u"#     → 3, 4, 5, 6列目の値をそれぞれx方向-, +側，y方向-, +側誤差棒のデータに使う\n")
        f.write(u"# 'err_columns': (None, 5) # → 5列目の値をy方向-+両側誤差棒のデータに使う\n")
        f.write(u"# 'err_columns': None # → 誤差棒をつけない\n")
        f.write(u"    'err_columns': None,\n")
        f.write(u"# 'every': 2 # → 2行毎にデータを読み込む\n")
        f.write(u"    'every': 1,\n")
        for i, j in zip(markers[:-1], markers_wx[:-1]):
            f.write(u"# 'marker': '" + i + u"' # → データ点に" + j + u"記号をつける\n")
        f.write(u"# 'marker': None # → データ点に記号をつけない\n")
        f.write(u"    'marker': 'o',\n")
        f.write(u"# 'marker_size': None # → 記号の大きさを指定しない\n")
        f.write(u"# 'marker_size': 8.0 # → 記号の大きさを8.0ポイントにする\n")
        f.write(u"    'marker_size': None,\n")
        f.write(u"# 'marker_color': 'white' # → 記号を白で塗りつぶす\n")
        f.write(u"# 'marker_color': (1.0, 0.1, 0.1) # → 記号を(R, G, B)で指定した色で塗りつぶす，0 ≤ R, G, B ≤ 1\n")
        f.write(u"    'marker_color': 'white',\n")
        f.write(u"# 'marker_edge_color': 'red' # → 記号の輪郭の色を赤にする\n")
        f.write(u"# 'marker_edge_color': 'none' # → 記号の輪郭を付けない\n")
        f.write(u"# 'marker_edge_color': (1.0, 0.1, 0.1) " +
            u"# → 記号の輪郭の色を(R, G, B)で指定した色にする，0 ≤ R, G, B ≤ 1\n")
        f.write(u"    'marker_edge_color': 'red',\n")
        f.write(u"# 'line_style': None # → データ点を線でつながない\n")
        for i, j in zip(line_styles[:-1], line_styles_wx[:-1]):
            f.write(u"# 'line_style': '%s' # → データ点を%sでつなぐ\n" % (i, j))
        f.write(u"    'line_style': None,\n")
        f.write(u"# 'line_color': 'red' # → 線の色を赤にする\n")
        f.write(u"# 'line_color': (1.0, 0.1, 0.1) # → 線の色を(R, G, B)で指定した色にする，0 ≤ R, G, B ≤ 1\n")
        f.write(u"    'line_color': 'red',\n")
        for i, j in zip(by_what_show_z, by_what_show_z_wx):
            f.write(u"# 'show_z_by': '%s' # → %sでz軸の値を表す\n" % (i, j))
        f.write(u"# 'show_z_by': None # → z軸の値はない\n")
        f.write(u"    'show_z_by': None,\n")
        f.write(u"# 'marker_size_ratio': 5.0 # → (z軸最大値を表す記号の大きさ)/(z軸最小値を表す記号の大きさ)を5.0倍にする\n")
        f.write(u"# 'marker_size_ratio': None # → z軸最大値を表す記号の大きさを指定しない\n")
        f.write(u"    'marker_size_ratio': None,\n")
        f.write(u"# 'label_in_legend': r'graph for $\\mathsf{a_{xy}^2}$'\n" +
            u"#     → 凡例に書くラベル，$\\mathsf{ と }$ の間に書くとTeX形式になる\n")
        f.write(u"# 'label_in_legend': None # → 凡例に題名を書かない\n")
        f.write(u"    'label_in_legend': None,\n")
        f.write(u"},\n")
        f.write(u"{\n")
        f.write(u"# ベクトル線図\n")
        f.write(u"    'type': 'vector',\n")
        f.write(u"# {:d} ≤ zorder ≤ {:d}，zorderが大きいほど前面に描く\n".format(zorder_min, zorder_max))
        f.write(u"    'zorder': 2,\n")
        f.write(u"# ファイル名，このファイルからの相対パス\n")
        f.write(u"    'file_name': '%s_example2.xlsx',\n" % s)
        f.write(u"# <テキストファイルの場合>\n")
        f.write(u"#  'columns': (1, 2, 3, 4) # → 1, 2, 3, 4列目の値をそれぞれx, y軸，u, vのデータに使う\n")
        f.write(u"#  'columns': ('$1/$2', 'sqrt($3)*10', 3, 4)\n" +
            u"#     → 1列目の値/2列目の値，3列目の値の平方根*10, 3, 4列目の値をそれぞれx, y軸，u, vのデータに使う\n")
        f.write(u"# <エクセルファイル（拡張子がxlsまたはxlsx）またはcsvファイルの場合>\n")
        f.write(u"#  'columns': ('1!A2', '1!B2', 'sqrt(1!C2)', '1!D2')\n" +
            u"#     → 1枚目のシートのセルA2から下の値，B2から下の値，C2から下の値の平方根，D2から下の値をそれぞれx, y軸，\n" +
            u"#       u, vのデータに使う\n")
        f.write(u"    'columns': ('1!A3', '1!B3', '1!D3', '1!E3'),\n")
        f.write(u"# 'every': 2 # → 2行毎にデータを読み込む\n")
        f.write(u"    'every': 1,\n")
        f.write(u"# 'color': 'red' # → 矢印の色を赤にする\n")
        f.write(u"# 'color': (1.0, 0.1, 0.1) # → 矢印の色を(R, G, B)で指定した色にする，0 ≤ R, G, B ≤ 1\n")
        f.write(u"# 'color': 'len' # → ベクトルの大きさに応じて矢印の色を虹色で変える\n")
        f.write(u"# 'color': 'len1' # → ベクトルの大きさに応じて矢印の色を虹色で変え，表示する矢印の長さを全て1にする\n")
        f.write(u"    'color': 'red',\n")
        f.write(u"# ベクトルの長さにかける倍率，大きいほど長く表示される\n")
        f.write(u"    'scale': 1.0,\n")
        f.write(u"# legend: (0.8, 1.03, 1.0, r'$\\mathsf{u/U_m}$')\n" +
            u"#     → 矢印先端の横方向座標を0.8，縦方向座標を1.03として，長さ1.0の矢印で凡例を付ける\n" +
            u"#       座標はグラフ左/下端を0，右/上端を1とした座標で指定する\n")
        f.write(u"# legend: None # → 凡例を付けない\n")
        f.write(u"    'legend': (0.8, 1.03, 1.0, r'V'),\n")
        f.write(u"},\n")
        f.write(u"{\n")
        f.write(u"# 等高線\n")
        f.write(u"    'type': 'contour',\n")
        f.write(u"# {:d} ≤ zorder ≤ {:d}，zorderが大きいほど前面に描く\n".format(zorder_min, zorder_max))
        f.write(u"    'zorder': 1,\n")
        f.write(u"# ファイル名，このファイルからの相対パス\n")
        f.write(u"    'file_name': '%s_example2.xlsx',\n" % s)
        f.write(u"# 'grid_pattern': True # → データが碁盤の目状に並んでいる\n")
        f.write(u"# 'grid_pattern': False # → データが碁盤の目状に並んでいない\n")
        f.write(u"    'grid_pattern': False,\n")
        f.write(u"# 'show_triangle': True # → データが碁盤の目状に並んでいない時に行う，デローニー分割した三角形を表示する\n")
        f.write(u"# 'show_triangle': False # → データが碁盤の目状に並んでいない時に行う，デローニー分割した三角形を表示しない\n")
        f.write(u"    'show_triangle': False,\n")
        f.write(u"# <テキストファイルの場合>\n")
        f.write(u"#  'columns': (1, 2, 3) # → 1, 2, 3列目の値をそれぞれx, y, z軸のデータに使う\n")
        f.write(u"#  'columns': ('$1/$2', 'sqrt($3)*10', 3)\n" +
            u"#     → 1列目の値/2列目の値，3列目の値の平方根*10, 3列目の値をそれぞれx, y, z軸のデータに使う\n")
        f.write(u"# <エクセルファイル（拡張子がxlsまたはxlsx）またはcsvファイルの場合>\n")
        f.write(u"#  'columns': ('1!A2', '1!B2', 'sqrt(1!C2)')\n"
            u"#     → 1枚目のシートのセルA2から下の値，B2から下の値，C2から下の値の平方根をそれぞれx, y, z軸のデータに使う\n")
        f.write(u"    'columns': ('1!A3', '1!B3', '1!C3'),\n")
        f.write(u"# 'every': 2 # → 2行毎にデータを読み込む\n")
        f.write(u"    'every': 1,\n")
        for i, j in zip(paint_styles[:-1], paint_styles_wx[:-1]):
            f.write(u"# 'paint': '%s' # → %sで塗る\n" % (i, j))
        f.write(u"# 'paint': None # → %s\n" % paint_styles_wx[-1])
        f.write(u"    'paint': 'rainbow',\n")
        f.write(u"# 'smooth_paint': True # → 面塗りを滑らかにする\n")
        f.write(u"    'smooth_paint': True,\n")
        f.write(u"},\n")
    s = os.path.splitext(file_name)[0] # unicode
    with codecs.open(u'%s_example1.txt' % s, 'w', encoding = 'UTF-8') as f:
        f.write(u'#y = x\n')
        f.write(u'#x\ty\n')
        for x in np.arange(-1.0, 1.01, 0.1):
            f.write(u'{:g}\t{:g}\n'.format(x, x))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row = 1, column = 1, value = 'x**2 + y**2')
    for i, j in enumerate(('x', 'y', 'z', 'u', 'v'), start = 1):
        ws.cell(row = 2, column = i, value = j)
    i = 2
    for x in np.arange(-1.0, 1.01, 0.1):
        for y in np.arange(-1.0, 1.01, 0.1):
            i += 1
            ws.cell(row = i, column = 1, value = float(x))
            ws.cell(row = i, column = 2, value = float(y))
            ws.cell(row = i, column = 3, value = float(math.sqrt(x**2 + y**2)))
            ws.cell(row = i, column = 4, value = float(-0.1*y))
            ws.cell(row = i, column = 5, value = float(0.1*x))
    wb.save(u'%s_example2.xlsx' % s)

    return file_name

class FileDropTargetForSetting(wx.FileDropTarget):
    def __init__(self, window, parent):
        wx.FileDropTarget.__init__(self)
        self.window = window
        self.parent = parent

    def OnDropFiles(self, x, y, filenames):
        if os.path.isfile(filenames[-1]):
            self.window.SetFocus()
            self.window.SetPath(correct_file_name_in_unicode(filenames[-1]))
            self.parent.load_settings(filenames[-1])
            return True
        else:
            return False

class MyFileDropTarget(wx.FileDropTarget):
    def __init__(self, window):
        wx.FileDropTarget.__init__(self)
        self.window = window

    def OnDropFiles(self, x, y, filenames):
        if os.path.isfile(filenames[-1]):
            self.window.SetFocus()
            self.window.SetPath(correct_file_name_in_unicode(filenames[-1]))
            return True
        else:
            return False

class MyTable(wx.grid.GridTableBase):
    def __init__(self):
        wx.grid.GridTableBase.__init__(self)

    def GetNumberRows(self):
        return len(self.data)

    def GetNumberCols(self):
        return len(self.data[0])

    def GetRowLabelValue(self, row):
        try:
            return self.row_labels[row]
        except:
            return str(row + 1).decode('UTF-8') if sys.version_info.major <= 2 else str(row + 1)

    def GetColLabelValue(self, col):
        return self.col_labels[col]

    def IsEmptyCell(self, row, col):
        try:
            return not self.data[row][col]
        except IndexError:
            return True

    def GetTypeName(self, row, col):
        return self.data_types[col]

    def AppendRows(self, numRows = 1):
        self.data.extend([self.new_data[:] for i in range(numRows)])
        return True

    def InsertRows(self, pos = 0, numRows = 1):
        self.data[pos:pos] = [self.new_data[:] for i in range(numRows)]
        return True

    def DeleteRows(self, pos = 0, numRows = 1):
        try:
            if self.GetNumberRows() == 0:
                return False
            del self.data[pos:pos + numRows]
            return True
        except:
            return False

tooltip_tex = (_(u'$\\mathsf{ と }$ の間に書くとTeX形式になり，上付き，下付き文字やギリシア文字が使えます．\n例）') +
    u'$\\mathsf{\\rho U_{inf}^2/2}$')

class TableForGraph(MyTable):
    def __init__(self):
        MyTable.__init__(self)
        self.ROW_X = 0
        self.ROW_Y = 1
        self.ROW_Z = 2
        self.row_labels = (_(u'x軸'), _(u'y軸'), _(u'z軸'))
        self.row_label_size = 60
        self.col_label_size = 22
        self.COL_FROM = 0
        self.COL_SIM = 1
        self.COL_TO = 2
        self.COL_TICKS = 3
        self.COL_LOG_SCALE = 4
        self.COL_LABEL = 5
        self.col_labels = (_(u'範囲(から)'), u'〜', _(u'範囲(まで)'), _(u'刻み'), _(u'対数軸'), _(u'ラベル'))
        self.col_sizes = (100, 40, 100, 100, 60, 290)
        s1 = _(u'空白の場合，自動スケールになります．')
        s2 = _(u'等高線またはベクトルの大きさを表す色に対する軸\n')
        self.tooltips = ((s1, None, s1, s1, None, tooltip_tex), (s1, None, s1, s1, None, tooltip_tex),
                         (s2 + s1, None, s2 + s1, s2 + s1, None, s2 + tooltip_tex))
        self.data_types = (wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_STRING,
                           wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_STRING,
                           wx.grid.GRID_VALUE_BOOL, wx.grid.GRID_VALUE_STRING)
        self.data = [[None, u'〜', None, None, False, u'x'],
                     [None, u'〜', None, None, False, u'y'],
                     [None, u'〜', None, None, False, u'z']]

    def GetValue(self, row, col):
        if col in (self.COL_FROM, self.COL_TO, self.COL_TICKS):
            if self.data[row][col] is None:
                return u''
            else:
                v = str(float(self.data[row][col]))
                return v.decode('UTF-8') if sys.version_info.major <= 2 else v
        elif col == self.COL_SIM:
            return self.data[row][col]
        elif col == self.COL_LOG_SCALE:
            return u'1' if self.data[row][col] else u''
        else:
            return u'' if self.data[row][col] is None else self.data[row][col]

    def SetValue(self, row, col, value):
        if col in (self.COL_FROM, self.COL_TO, self.COL_TICKS):
            try:
                self.data[row][col] = float(value)
            except:
                self.data[row][col] = None
        elif col == self.COL_LOG_SCALE:
            self.data[row][col] = bool(value)
        elif col == self.COL_LABEL:
            value = value.decode('UTF-8') if sys.version_info.major <= 2 and type(value) is str else value
            self.data[row][col] = None if value == u'' else value

tooltip_zorder = (u'{:d} ≤ Z-order ≤ {:d}\n'.format(zorder_min, zorder_max)) + _(u'大きいほど前面に描きます．')
tooltip_coordinate = _(u'axes座標はグラフ枠の左下を(0, 0), 右上を(1, 1)にとる座標，data座標はふつうのx, y座標')

class TableForText(MyTable):
    def __init__(self):
        MyTable.__init__(self)
        self.row_label_size = 40
        self.col_label_size = 22
        self.COL_COORDINATE = 0
        self.COL_X = 1
        self.COL_Y = 2
        self.COL_ALIGN = 3
        self.COL_STRING = 4
        self.COL_FONTSIZE = 5
        self.COL_ZORDER = 6
        self.col_labels = (_(u'座標系'), _(u'横座標'), _(u'縦座標'), _(u'文字揃え'), _(u'文字列'), _(u'サイズ'), u'Z-order')
        self.col_sizes = (60, 70, 70, 70, 320, 60, 60)
        self.tooltips = ((tooltip_coordinate, None, None, None, tooltip_tex, None, tooltip_zorder),)
        self.data_types = (wx.grid.GRID_VALUE_CHOICE + ':axes,data', wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_STRING,
                           wx.grid.GRID_VALUE_CHOICE + _(u':左,中央,右'), wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_STRING,
                           wx.grid.GRID_VALUE_NUMBER + ':{:d},{:d}'.format(zorder_min, zorder_max))
        self.new_data = ['axes', None, None, 'l', None, plt.rcParams['font.size'], zorder_others]
        self.data = [self.new_data[:], self.new_data[:], self.new_data[:]]

    def GetValue(self, row, col):
        if col in (self.COL_COORDINATE, self.COL_ZORDER):
            return self.data[row][col]
        elif col == self.COL_ALIGN:
            return _(u'左') if self.data[row][col] == 'l' else (_(u'中央') if self.data[row][col] == 'c' else _(u'右'))
        elif col == self.COL_STRING:
            return u'' if self.data[row][col] is None else self.data[row][col]
        else:
            if self.data[row][col] is None:
                return u''
            else:
                v = str(float(self.data[row][col]))
                return v.decode('UTF-8') if sys.version_info.major <= 2 else v

    def SetValue(self, row, col, value):
        if col == self.COL_COORDINATE:
            self.data[row][col] = value
        elif col == self.COL_ALIGN:
            if sys.version_info.major <= 2 and type(value) is str:
                value = value.decode('UTF-8')
            self.data[row][col] = 'l' if value == _(u'左') else ('c' if value == _(u'中央') else 'r')
        elif col == self.COL_STRING:
            value = value.strip().decode('UTF-8') if sys.version_info.major <= 2 and type(value) is str else value.strip()
            self.data[row][col] = None if value == u'' else value
        elif col == self.COL_ZORDER:
            try:
                self.data[row][col] = int(value)
            except:
                self.data[row][col] = self.new_data[self.COL_ZORDER]
        else:
            try:
                self.data[row][col] = float(value)
            except:
                self.data[row][col] = None

    def Clear(self):
        for i in range(self.GetNumberRows()):
            self.data[i] = self.new_data[:]

class TableForArrow(MyTable):
    def __init__(self):
        MyTable.__init__(self)
        self.row_label_size = 40
        self.col_label_size = 22
        self.COL_COORDINATE = 0
        self.COL_X_FROM = 1
        self.COL_Y_FROM = 2
        self.COL_X_TO = 3
        self.COL_Y_TO = 4
        self.COL_LINE_WIDTH = 5
        self.COL_LINE_STYLE = 6
        self.COL_HEAD_WIDTH = 7
        self.COL_ZORDER = 8
        self.col_labels = (_(u'座標系'), _(u'始点横座標'), _(u'始点縦座標'), _(u'終点横座標'), _(u'終点縦座標'),
                           _(u'線幅'), _(u'線種'), _(u'矢の幅'), u'Z-order')
        self.col_sizes = (60, 85, 85, 85, 85, 85, 80, 85, 60)
        self.tooltips = ((tooltip_coordinate, None, None, None, None, None, None,
                          _(u'0にすれば直線になります'), tooltip_zorder),)
        self.data_types = (wx.grid.GRID_VALUE_CHOICE + ':axes,data', wx.grid.GRID_VALUE_STRING,
                           wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_STRING,
                           wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_CHOICE + ':' + ','.join(line_styles_wx[:-1]),
                           wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_NUMBER + ':{:d},{:d}'.format(zorder_min, zorder_max))
        self.new_data = ['axes', None, None, None, None, 1.0, line_styles[0], 12.0, zorder_others]
        self.data = [self.new_data[:], self.new_data[:], self.new_data[:]]

    def GetValue(self, row, col):
        if col in (self.COL_COORDINATE, self.COL_ZORDER):
            return self.data[row][col]
        elif col == self.COL_LINE_STYLE:
            return line_styles_wx[line_styles.index(self.data[row][col])]
        else:
            if self.data[row][col] is None:
                return u''
            else:
                v = str(float(self.data[row][col]))
                return v.decode('UTF-8') if sys.version_info.major <= 2 else v

    def SetValue(self, row, col, value):
        if col == self.COL_COORDINATE:
            self.data[row][col] = value
        elif col == self.COL_LINE_STYLE:
            self.data[row][col] = line_styles[line_styles_wx.index(value)]
        elif col == self.COL_ZORDER:
            try:
                self.data[row][col] = int(value)
            except:
                self.data[row][col] = self.new_data[self.COL_ZORDER]
        else:
            try:
                self.data[row][col] = float(value)
            except:
                self.data[row][col] = self.new_data[self.COL_LINE_WIDTH] if col == self.COL_LINE_WIDTH else (
                    self.new_data[self.COL_HEAD_WIDTH] if col == self.COL_HEAD_WIDTH else None)

    def Clear(self):
        for i in range(self.GetNumberRows()):
            self.data[i] = self.new_data[:]

class TableForEllipse(MyTable):
    def __init__(self):
        MyTable.__init__(self)
        self.row_label_size = 40
        self.col_label_size = 22
        self.COL_COORDINATE = 0
        self.COL_X = 1
        self.COL_Y = 2
        self.COL_WIDTH = 3
        self.COL_HEIGHT = 4
        self.COL_ANGLE = 5
        self.COL_LINE_WIDTH = 6
        self.COL_FILL = 7
        self.COL_ZORDER = 8
        self.col_labels = (_(u'座標系'), _(u'中心横座標'), _(u'中心縦座標'),
                           _(u'幅'), _(u'高さ'), _(u'角度[deg]'), _(u'線幅'), _(u'塗りつぶし'), u'Z-order')
        self.col_sizes = (60, 95, 95, 85, 85, 80, 80, 70, 60)
        self.tooltips = ((tooltip_coordinate, None, None, None, None, _(u'反時計回りに正'), None, None, tooltip_zorder),)
        self.data_types = (wx.grid.GRID_VALUE_CHOICE + ':axes,data', wx.grid.GRID_VALUE_STRING,
                           wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_STRING,
                           wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_BOOL,
                           wx.grid.GRID_VALUE_NUMBER + ':{:d},{:d}'.format(zorder_min, zorder_max))
        self.new_data = ['axes', None, None, None, None, 0.0, 1.0, False, zorder_others]
        self.data = [self.new_data[:], self.new_data[:], self.new_data[:]]

    def GetValue(self, row, col):
        if col in (self.COL_COORDINATE, self.COL_ZORDER):
            return self.data[row][col]
        elif col == self.COL_FILL:
            return u'1' if self.data[row][col] else u''
        else:
            if self.data[row][col] is None:
                return u''
            else:
                v = str(float(self.data[row][col]))
                return v.decode('UTF-8') if sys.version_info.major <= 2 else v

    def SetValue(self, row, col, value):
        if col == self.COL_COORDINATE:
            self.data[row][col] = value
        elif col == self.COL_FILL:
            self.data[row][col] = bool(value)
        elif col == self.COL_ZORDER:
            try:
                self.data[row][col] = int(value)
            except:
                self.data[row][col] = self.new_data[self.COL_ZORDER]
        else:
            try:
                self.data[row][col] = float(value)
            except:
                self.data[row][col] = self.new_data[self.COL_ANGLE] if col == self.COL_ANGLE else (
                    self.new_data[self.COL_LINE_WIDTH] if col == self.COL_LINE_WIDTH else None)

    def Clear(self):
        for i in range(self.GetNumberRows()):
            self.data[i] = self.new_data[:]

class TableForPolygon(MyTable):
    def __init__(self):
        MyTable.__init__(self)
        self.row_label_size = 40
        self.col_label_size = 22
        self.COL_COORDINATE = 0
        self.COL_XY = 1
        self.COL_LINE_WIDTH = 2
        self.COL_FILL = 3
        self.COL_ZORDER = 4
        self.col_labels = (_(u'座標系'), _(u'頂点の横座標と縦座標'), _(u'線幅'), _(u'塗りつぶし'), u'Z-order')
        self.col_sizes = (60, 440, 80, 70, 60)
        self.tooltips = ((tooltip_coordinate,
                         _(u'1点目の横座標, 1点目の縦座標, 2点目の横座標, 2点目の縦座標, ...\nのようにコンマ区切りで記入'),
                         None, None, tooltip_zorder),)
        self.data_types = (wx.grid.GRID_VALUE_CHOICE + ':axes,data', wx.grid.GRID_VALUE_STRING,
                           wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_BOOL,
                           wx.grid.GRID_VALUE_NUMBER + ':{:d},{:d}'.format(zorder_min, zorder_max))
        self.new_data = ['axes', None, 1.0, False, zorder_others]
        self.data = [self.new_data[:], self.new_data[:], self.new_data[:]]

    def GetValue(self, row, col):
        if col in (self.COL_COORDINATE, self.COL_ZORDER):
            return self.data[row][col]
        elif col == self.COL_XY:
            return (u''  if self.data[row][col] is None
                else u', '.join([u'{:g}, {:g}'.format(i[0], i[1]) for i in self.data[row][col]]))
        elif col == self.COL_FILL:
            return u'1' if self.data[row][col] else u''
        else:
            if self.data[row][col] is None:
                return u''
            else:
                v = str(float(self.data[row][col]))
                return v.decode('UTF-8') if sys.version_info.major <= 2 else v

    def SetValue(self, row, col, value):
        if col == self.COL_COORDINATE:
            self.data[row][col] = value
        elif col == self.COL_XY:
            try:
                value = [float(i) for i in value.split(u',')]
                self.data[row][col] = [[value[i], value[i + 1]] for i in range(0, len(value), 2)]
            except:
                self.data[row][col] = None
        elif col == self.COL_FILL:
            self.data[row][col] = bool(value)
        elif col == self.COL_ZORDER:
            try:
                self.data[row][col] = int(value)
            except:
                self.data[row][col] = self.new_data[self.COL_ZORDER]
        else: # COL_LINE_WIDTH
            try:
                self.data[row][col] = float(value)
            except:
                self.data[row][col] = self.new_data[self.COL_LINE_WIDTH]

    def Clear(self):
        for i in range(self.GetNumberRows()):
            self.data[i] = self.new_data[:]

class TableForNaca4(MyTable):
    def __init__(self):
        MyTable.__init__(self)
        self.row_label_size = 40
        self.col_label_size = 22
        self.COL_4DIGITS = 0
        self.COL_COORDINATE = 1
        self.COL_XLE = 2
        self.COL_YLE = 3
        self.COL_XTE = 4
        self.COL_YTE = 5
        self.COL_LINE_WIDTH = 6
        self.COL_FILL = 7
        self.COL_ZORDER = 8
        self.col_labels = (_(u'4桁'), _(u'座標系'), _(u'前縁横座標'), _(u'前縁縦座標'),
                           _(u'後縁横座標'), _(u'後縁縦座標'), _(u'線幅'), _(u'塗りつぶし'), u'Z-order')
        self.col_sizes = (60, 60, 95, 95, 95, 95, 80, 70, 60)
        self.tooltips = ((None, tooltip_coordinate, None, None, None, None, None, None, tooltip_zorder),)
        self.data_types = (wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_CHOICE + ':axes,data',
                           wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_STRING,
                           wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_STRING,
                           wx.grid.GRID_VALUE_BOOL, wx.grid.GRID_VALUE_NUMBER + ':{:d},{:d}'.format(zorder_min, zorder_max))
        self.new_data = [None, 'axes', None, None, None, None, 1.0, False, zorder_others]
        self.data = [self.new_data[:], self.new_data[:], self.new_data[:]]

    def GetValue(self, row, col):
        if col in (self.COL_4DIGITS, self.COL_COORDINATE, self.COL_ZORDER):
            return self.data[row][col]
        elif col == self.COL_FILL:
            return u'1' if self.data[row][col] else u''
        else:
            if self.data[row][col] is None:
                return u''
            else:
                v = str(float(self.data[row][col]))
                return v.decode('UTF-8') if sys.version_info.major <= 2 else v

    def SetValue(self, row, col, value):
        if col == self.COL_4DIGITS:
            try:
                self.data[row][col] = (u'%04d' % int(value))[:4]
            except:
                self.data[row][col] = None
        elif col == self.COL_COORDINATE:
            self.data[row][col] = value
        elif col == self.COL_FILL:
            self.data[row][col] = bool(value)
        elif col == self.COL_ZORDER:
            try:
                self.data[row][col] = int(value)
            except:
                self.data[row][col] = self.new_data[self.COL_ZORDER]
        else:
            try:
                self.data[row][col] = float(value)
            except:
                self.data[row][col] = self.new_data[self.COL_LINE_WIDTH] if col == self.COL_LINE_WIDTH else None

    def Clear(self):
        for i in range(self.GetNumberRows()):
            self.data[i] = self.new_data[:]

class TableForParamDict(MyTable):
    def __init__(self):
        MyTable.__init__(self)
        self.row_label_size = 40
        self.col_label_size = 22
        self.COL_SYMBOL = 0
        self.COL_VALUE = 1
        self.col_labels = (_(u'パラメータ'), _(u'数値'))
        self.col_sizes = (200, 300)
        self.tooltips = ((_(u'数字で始まるパラメータは不可'), None),)
        self.data_types = (wx.grid.GRID_VALUE_STRING, wx.grid.GRID_VALUE_STRING)
        self.new_data = [None, None]
        self.data = [self.new_data[:], self.new_data[:], self.new_data[:]]

    def GetValue(self, row, col):
        if col == self.COL_SYMBOL:
            return u'' if self.data[row][col] is None else self.data[row][col]
        else:
            if self.data[row][col] is None:
                return u''
            else:
                v = str(float(self.data[row][col]))
                return v.decode('UTF-8') if sys.version_info.major <= 2 else v

    def SetValue(self, row, col, value):
        if col == self.COL_SYMBOL:
            value = (value.decode('UTF-8') if sys.version_info.major <= 2 and type(value) is str else value).strip()
            self.data[row][col] = None if value == u'' or value[0].isdigit() else value
        else:
            try:
                self.data[row][col] = float(value)
            except:
                self.data[row][col] = None

    def Clear(self):
        for i in range(self.GetNumberRows()):
            self.data[i] = self.new_data[:]

class GridWithCellToolTip(wx.grid.Grid):
    def __init__(self, parent, id = wx.ID_ANY, pos = wx.DefaultPosition, size = wx.DefaultSize,
        style = wx.WANTS_CHARS, name = u'', table = None):
        wx.grid.Grid.__init__(self, parent, id, pos, size, style, name)
        self.table = table
        if self.table is not None:
            self.SetTable(self.table, takeOwnership = True)
            try:
                self.SetRowLabelSize(self.table.row_label_size)
            except:
                pass
            try:
                self.SetColLabelSize(self.table.col_label_size)
            except:
                pass
            try:
                for i, j in enumerate(self.table.col_sizes):
                    self.SetColSize(i, j)
            except:
                pass
        self.EnableDragRowSize(False) # needed to avoid "Segmentation fault: 11" in button_delete_gridOnButtonClick
        self.GetGridWindow().Bind(wx.EVT_MOTION, self.OnMouseOver)

    def OnMouseOver(self, event):
        # Method to calculate where the mouse is pointing and then set the tooltip dynamically.
        # https://stackoverflow.com/questions/20589686/tooltip-message-when-hovering-on-cell-with-mouse-in-wx-grid-wxpython
        try:
            if self.table.tooltips is not None:
                # Use CalcUnscrolledPosition() to get the mouse position within the entire grid including what's offscreen
                c = self.XYToCell(*self.CalcUnscrolledPosition(event.GetX(), event.GetY()))
                row = min(c.Row, len(self.table.tooltips) - 1)
                col = min(c.Col, len(self.table.tooltips[0]) - 1)
                s = self.table.tooltips[row][col]
                if s is None:
                    s = u''
                event.GetEventObject().SetToolTip(s)
        except:
#            print(sys.exc_info())
            pass
        event.Skip()

    def AppendRows(self, numRows = 1, updateLabels = True):
        if wx.grid.Grid.AppendRows(self, numRows, updateLabels):
            self.UpdateView(numRows)
            return True
        else:
            return False

    def InsertRows(self, pos = 0, numRows = 1, updateLabels = True):
        if wx.grid.Grid.InsertRows(self, pos, numRows, updateLabels):
            self.UpdateView(numRows)
            return True
        else:
            return False

    def DeleteRows(self, pos = 0, numRows = 1, updateLabels = True):
        if wx.grid.Grid.DeleteRows(self, pos, numRows, updateLabels):
            self.UpdateView(-numRows)
            return True
        else:
            return False

    def UpdateView(self, numOfRowsIncreased):
        self.BeginBatch()
        self.ProcessTableMessage(
            wx.grid.GridTableMessage(self.table, wx.grid.GRIDTABLE_NOTIFY_ROWS_APPENDED, numOfRowsIncreased))
        self.ProcessTableMessage(
            wx.grid.GridTableMessage(self.table, wx.grid.GRIDTABLE_REQUEST_VIEW_GET_VALUES))
        self.EndBatch()
        self.AdjustScrollbars()
        self.ForceRefresh()

    def Copy(self):
        col = self.GetGridCursorCol()
        if self.table.data_types[col] == wx.grid.GRID_VALUE_STRING:
            pyperclip.copy(self.table.GetValue(self.GetGridCursorRow(), col))

    def Paste(self):
        col = self.GetGridCursorCol()
        if self.table.data_types[col] == wx.grid.GRID_VALUE_STRING:
            self.table.SetValue(self.GetGridCursorRow(), col, pyperclip.paste())
            self.ForceRefresh()

    def Cut(self):
        col = self.GetGridCursorCol()
        if self.table.data_types[col] == wx.grid.GRID_VALUE_STRING:
            row = self.GetGridCursorRow()
            pyperclip.copy(self.table.GetValue(row, col))
            self.table.SetValue(row, col, None)
            self.ForceRefresh()

###########################################################################
## Class FrameMain
###########################################################################

class FrameMain(wx.Frame):
    def __init__(self, parent):
        wx.Frame.__init__(self, parent, id = wx.ID_ANY,
            title = 'matplotlibwx (' + version + ') by Python ' + platform.python_version(),
            pos = wx.DefaultPosition, size = wx.Size(800, 700), style = wx.DEFAULT_FRAME_STYLE | wx.TAB_TRAVERSAL)
        self.SetSizeHints(wx.DefaultSize, wx.DefaultSize)
        self.SetForegroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOWTEXT))
        self.SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOW))

        tooltip_either_empty = _(u'どちらかが空白の場合，指定しません．')
        tooltip_graph_coordinate = _(u'空白の場合，指定しません．\nグラフ左/下端を0，右/上端を1とします．')
        tooltip_screen_coordinate = _(u'空白の場合，指定しません．\n画面左/下端を0，右/上を1とします．')
        tooltip_append_row = _(u'下に行を追加')
        tooltip_delete_row = _(u'選択している行を削除')
        label_reset = _(u'リセット')
        tooltip_column = (_(u"<テキストファイルの場合>\n" +
                            u" 1以上の整数または ' と ' で囲んだ数式\n" +
                            u" 数式中では列番号の前に$をつけます．\n 例）") +
                            u"'sqrt($3)*10'\n" +
                          _(u"<エクセルファイル（拡張子がxlsまたはxlsx）またはcsvファイルの場合>\n" +
                            u" 'シート番号!読み取り開始のセル番号' で指定します．\n" +
                            u" 行方向にデータがあるものとします．\n" +
                            u" テキストファイルの場合と同様に，数式が使えます．\n 例）") +
                            u"'sqrt(1!A2)'")
        tooltip_err_column = _(u'両側が空白の場合，指定しません．\nいずれかだけの場合，-+両側で同じデータを使用します．')
        tooltip_drag_and_drop = _(u'ドラッグ&ドロップで決めることもできます．')
        tooltip_equation_scatter = (_(u"次の書式を用いた数式のプロットも可能です．\n" +
                                      u"分割数の前の'/'を'L/'にすれば，対数軸上で等分割にできます：\n" +
                                      u" 数式, 変数 = [最小値, 最大値]/分割数\n [xの数式, yの数式], 変数 = [最小値, 最大値]/分割数\n例）") +
                                      u"sqrt(2.2*x), x = [0.0, 1.0]/100\n        [cos(t), sin(t)], t = [0.0, pi]/100")
        tooltip_open_file = _(u'ファイルを開く')

        spinCtrlWidth = 118 if sys.platform == 'linux2' else 60

        bSizer1 = wx.BoxSizer(wx.VERTICAL)

        self.scrolledWindow_graph = wx.ScrolledWindow(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize,
            wx.HSCROLL | wx.VSCROLL)
        self.scrolledWindow_graph.SetScrollRate(5, 5)
        bSizer2 = wx.BoxSizer(wx.VERTICAL)

        bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

        staticText1 = wx.StaticText(self.scrolledWindow_graph, wx.ID_ANY, _(u'設定ファイル：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        self.filePicker_setting = wx.FilePickerCtrl(self.scrolledWindow_graph, wx.ID_ANY, wx.EmptyString,
            _(u'設定ファイルを開く'), u'*.*', wx.DefaultPosition, wx.DefaultSize,
            wx.FLP_USE_TEXTCTRL | wx.FLP_OPEN | wx.FLP_FILE_MUST_EXIST | wx.FLP_SMALL, name = 'filePicker_setting')
        self.filePicker_setting.SetForegroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOWTEXT))
        self.filePicker_setting.SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOW))
        self.filePicker_setting.SetToolTip(tooltip_drag_and_drop)
        self.filePicker_setting.SetDropTarget(FileDropTargetForSetting(self.filePicker_setting, self))
        bSizer3.Add(self.filePicker_setting, 1, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        self.button_setting_open_file = wx.Button(self.scrolledWindow_graph, wx.ID_ANY, _(u'開く'),
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'bfilePicker_setting')
        self.button_setting_open_file.SetToolTip(_(u'設定ファイルを開く'))
        bSizer3.Add(self.button_setting_open_file, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        bSizer2.Add(bSizer3, 0, wx.EXPAND, 5)

        sbSizer1 = wx.StaticBoxSizer(wx.StaticBox(self.scrolledWindow_graph, wx.ID_ANY, _(u'画像')), wx.HORIZONTAL)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'保存先：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        sbSizer1.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.filePicker_png = wx.FilePickerCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
            _(u'pngファイルを保存'), u'*.png', wx.DefaultPosition, wx.DefaultSize,
            wx.FLP_USE_TEXTCTRL | wx.FLP_SAVE | wx.FLP_OVERWRITE_PROMPT | wx.FLP_SMALL)
        self.filePicker_png.SetForegroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOWTEXT))
        self.filePicker_png.SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOW))
        self.filePicker_png.SetToolTip(_(u'空白の場合，matplotlibのビューワーにグラフを表示します．\n') + tooltip_drag_and_drop)
        self.filePicker_png.SetDropTarget(MyFileDropTarget(self.filePicker_png))
        sbSizer1.Add(self.filePicker_png, 1, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'大きさ：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        sbSizer1.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_fig_size_w = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_fig_size_w.SetToolTip(tooltip_either_empty)
        sbSizer1.Add(self.textCtrl_fig_size_w, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, u'in ×',
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        sbSizer1.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM , 5)

        self.textCtrl_fig_size_h = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_fig_size_h.SetToolTip(tooltip_either_empty)
        sbSizer1.Add(self.textCtrl_fig_size_h, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, u'in',
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        sbSizer1.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        bSizer2.Add(sbSizer1, 0, wx.EXPAND, 5)

        sbSizer1 = wx.StaticBoxSizer(wx.StaticBox(self.scrolledWindow_graph, wx.ID_ANY, _(u'データファイル')), wx.HORIZONTAL)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'読み飛ばす行の先頭文字：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)

        sbSizer1.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_skip = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'#', wx.DefaultPosition, wx.Size(60, -1), 0)
        sbSizer1.Add(self.textCtrl_skip, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'区切り文字：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        sbSizer1.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_delimiter = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_delimiter.SetToolTip(_(u'空白の場合，半角スペースかタブ区切りになります．'))
        sbSizer1.Add(self.textCtrl_delimiter, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'読み込み終了文字列：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        sbSizer1.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_terminator = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'!!STOP',
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_terminator.SetToolTip(_(u'空白の場合，最後の行まで読み込みます．'))
        sbSizer1.Add(self.textCtrl_terminator, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        bSizer2.Add(sbSizer1, 0, wx.EXPAND, 5)

        sbSizer1 = wx.StaticBoxSizer(wx.StaticBox(self.scrolledWindow_graph, wx.ID_ANY, _(u'グラフ')), wx.VERTICAL)

        self.grid_graph = GridWithCellToolTip(sbSizer1.GetStaticBox(), wx.ID_ANY,
            wx.DefaultPosition, wx.DefaultSize, 0, table = TableForGraph())
        self.grid_graph.SetDefaultCellAlignment(wx.ALIGN_CENTER, wx.ALIGN_CENTER)
        attr = wx.grid.GridCellAttr()
        attr.SetReadOnly(True)
        self.grid_graph.SetColAttr(self.grid_graph.table.COL_SIM, attr)
        attr = wx.grid.GridCellAttr()
        attr.SetAlignment(wx.ALIGN_LEFT, wx.ALIGN_CENTER)
        self.grid_graph.SetColAttr(self.grid_graph.table.COL_LABEL, attr)
        sbSizer1.Add(self.grid_graph, 1, wx.EXPAND, 5)

        bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

        self.checkBox_x_grid = wx.CheckBox(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'x軸目盛線'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        self.checkBox_x_grid.SetValue(False)
        bSizer3.Add(self.checkBox_x_grid, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.checkBox_y_grid = wx.CheckBox(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'y軸目盛線'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        self.checkBox_y_grid.SetValue(False)
        bSizer3.Add(self.checkBox_y_grid, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'y軸を1進む長さ/x軸を1進む長さ：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.comboBox_aspect = wx.ComboBox(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
            wx.DefaultPosition, wx.DefaultSize, [u"auto", u"equal"], wx.CB_DROPDOWN)
        self.comboBox_aspect.SetValue(u"auto")
        self.comboBox_aspect.SetForegroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOWTEXT))
        self.comboBox_aspect.SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOW))
        self.comboBox_aspect.SetToolTip(_(u'数値またはauto, equal\n数値は(y軸の長さ)/(x軸の長さ)*(x軸の最大値-最小値)/(y軸の最大値-最小値)の値を入力します．\nequalは1と同じ意味です．'))
        bSizer3.Add(self.comboBox_aspect, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'グラフの上に書く題名：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_title = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
            wx.DefaultPosition, wx.Size(-1, -1), 0)
        self.textCtrl_title.SetToolTip(tooltip_tex)
        bSizer3.Add(self.textCtrl_title, 1, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

        bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'凡例の左端：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.Left, 5)

        self.textCtrl_legend_left = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_legend_left.SetToolTip(tooltip_graph_coordinate)
        bSizer3.Add(self.textCtrl_legend_left, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'上端：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        self.textCtrl_legend_top = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_legend_top.SetToolTip(tooltip_graph_coordinate)
        bSizer3.Add(self.textCtrl_legend_top, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'グラフの左端：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_graph_left = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_graph_left.SetToolTip(tooltip_screen_coordinate)
        bSizer3.Add(self.textCtrl_graph_left, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'下端：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        self.textCtrl_graph_bottom = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_graph_bottom.SetToolTip(tooltip_screen_coordinate)
        bSizer3.Add(self.textCtrl_graph_bottom, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'右端：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        self.textCtrl_graph_right = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_graph_right.SetToolTip(tooltip_screen_coordinate)
        bSizer3.Add(self.textCtrl_graph_right, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'上端：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        self.textCtrl_graph_top = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_graph_top.SetToolTip(tooltip_screen_coordinate)
        bSizer3.Add(self.textCtrl_graph_top, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

        bSizer2.Add(sbSizer1, 0, wx.EXPAND, 5)

        sbSizer1 = wx.StaticBoxSizer(wx.StaticBox(self.scrolledWindow_graph, wx.ID_ANY, _(u'文字列')), wx.VERTICAL)

        self.grid_text = GridWithCellToolTip(sbSizer1.GetStaticBox(), wx.ID_ANY,
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'grid_text', table = TableForText())
        self.grid_text.SetDefaultCellAlignment(wx.ALIGN_CENTER, wx.ALIGN_CENTER)
        attr = wx.grid.GridCellAttr()
        attr.SetAlignment(wx.ALIGN_LEFT, wx.ALIGN_CENTER)
        self.grid_text.SetColAttr(self.grid_text.table.COL_STRING, attr)
        sbSizer1.Add(self.grid_text, 1, wx.EXPAND, 5)

        bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

        self.button_append_text = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, u'+',
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'agrid_text')
        self.button_append_text.SetToolTip(tooltip_append_row)
        bSizer3.Add(self.button_append_text, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        self.button_delete_text = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, u'-',
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'dgrid_text')
        self.button_delete_text.SetToolTip(tooltip_delete_row)
        bSizer3.Add(self.button_delete_text, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        self.button_clear_text = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, label_reset,
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'cgrid_text')
        bSizer3.Add(self.button_clear_text, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

        bSizer2.Add(sbSizer1, 0, wx.EXPAND, 5)

        sbSizer1 = wx.StaticBoxSizer(wx.StaticBox(self.scrolledWindow_graph, wx.ID_ANY,
            _(u'矢印．矢の幅を0にすれば直線になります．')), wx.VERTICAL)

        self.grid_arrow = GridWithCellToolTip(sbSizer1.GetStaticBox(), wx.ID_ANY,
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'grid_arrow', table = TableForArrow())
        self.grid_arrow.SetDefaultCellAlignment(wx.ALIGN_CENTER, wx.ALIGN_CENTER)
        sbSizer1.Add(self.grid_arrow, 1, wx.EXPAND, 5)

        bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

        self.button_append_arrow = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, u'+',
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'agrid_arrow')
        self.button_append_arrow.SetToolTip(tooltip_append_row)
        bSizer3.Add(self.button_append_arrow, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        self.button_delete_arrow = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, u'-',
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'dgrid_arrow')
        self.button_delete_arrow.SetToolTip(tooltip_delete_row)
        bSizer3.Add(self.button_delete_arrow, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        self.button_clear_arrow = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, label_reset,
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'cgrid_arrow')
        bSizer3.Add(self.button_clear_arrow, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

        bSizer2.Add(sbSizer1, 0, wx.EXPAND, 5)

        sbSizer1 = wx.StaticBoxSizer(wx.StaticBox(self.scrolledWindow_graph, wx.ID_ANY,
            _(u'楕円')), wx.VERTICAL)

        self.grid_ellipse = GridWithCellToolTip(sbSizer1.GetStaticBox(), wx.ID_ANY,
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'grid_ellipse', table = TableForEllipse())
        self.grid_ellipse.SetDefaultCellAlignment(wx.ALIGN_CENTER, wx.ALIGN_CENTER)
        sbSizer1.Add(self.grid_ellipse, 1, wx.EXPAND, 5)

        bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

        self.button_append_ellipse = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, u'+',
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'agrid_ellipse')
        self.button_append_ellipse.SetToolTip(tooltip_append_row)
        bSizer3.Add(self.button_append_ellipse, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        self.button_delete_ellipse = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, u'-',
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'dgrid_ellipse')
        self.button_delete_ellipse.SetToolTip(tooltip_delete_row)
        bSizer3.Add(self.button_delete_ellipse, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        self.button_clear_ellipse = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, label_reset,
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'cgrid_ellipse')
        bSizer3.Add(self.button_clear_ellipse, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

        bSizer2.Add(sbSizer1, 0, wx.EXPAND, 5)

        sbSizer1 = wx.StaticBoxSizer(wx.StaticBox(self.scrolledWindow_graph, wx.ID_ANY,
            _(u'多角形')), wx.VERTICAL)

        self.grid_polygon = GridWithCellToolTip(sbSizer1.GetStaticBox(), wx.ID_ANY,
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'grid_polygon', table = TableForPolygon())
        self.grid_polygon.SetDefaultCellAlignment(wx.ALIGN_CENTER, wx.ALIGN_CENTER)
        sbSizer1.Add(self.grid_polygon, 1, wx.EXPAND, 5)

        bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

        self.button_append_polygon = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, u'+',
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'agrid_polygon')
        self.button_append_polygon.SetToolTip(tooltip_append_row)
        bSizer3.Add(self.button_append_polygon, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        self.button_delete_polygon = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, u'-',
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'dgrid_polygon')
        self.button_delete_polygon.SetToolTip(tooltip_delete_row)
        bSizer3.Add(self.button_delete_polygon, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        self.button_clear_polygon = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, label_reset,
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'cgrid_polygon')
        bSizer3.Add(self.button_clear_polygon, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

        bSizer2.Add(sbSizer1, 0, wx.EXPAND, 5)

        sbSizer1 = wx.StaticBoxSizer(wx.StaticBox(self.scrolledWindow_graph, wx.ID_ANY,
            _(u'NACA 4桁系列翼')), wx.VERTICAL)

        self.grid_naca4 = GridWithCellToolTip(sbSizer1.GetStaticBox(), wx.ID_ANY,
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'grid_naca4', table = TableForNaca4())
        self.grid_naca4.SetDefaultCellAlignment(wx.ALIGN_CENTER, wx.ALIGN_CENTER)
        sbSizer1.Add(self.grid_naca4, 1, wx.EXPAND, 5)

        bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

        self.button_append_naca4 = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, u'+',
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'agrid_naca4')
        self.button_append_naca4.SetToolTip(tooltip_append_row)
        bSizer3.Add(self.button_append_naca4, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        self.button_delete_naca4 = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, u'-',
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'dgrid_naca4')
        self.button_delete_naca4.SetToolTip(tooltip_delete_row)
        bSizer3.Add(self.button_delete_naca4, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        self.button_clear_naca4 = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, label_reset,
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'cgrid_naca4')
        bSizer3.Add(self.button_clear_naca4, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

        bSizer2.Add(sbSizer1, 0, wx.EXPAND, 5)

        sbSizer1 = wx.StaticBoxSizer(wx.StaticBox(self.scrolledWindow_graph, wx.ID_ANY,
            _(u'ファイル読み取り列指定を数式で行う際に使うパラメータ')), wx.VERTICAL)

        self.grid_param_dict = GridWithCellToolTip(sbSizer1.GetStaticBox(), wx.ID_ANY,
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'grid_param_dict', table = TableForParamDict())
        self.grid_param_dict.SetDefaultCellAlignment(wx.ALIGN_CENTER, wx.ALIGN_CENTER)
        sbSizer1.Add(self.grid_param_dict, 1, 0, 5)

        bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

        self.button_append_param_dict = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, u'+',
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'agrid_param_dict')
        self.button_append_param_dict.SetToolTip(tooltip_append_row)
        bSizer3.Add(self.button_append_param_dict, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        self.button_delete_param_dict = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, u'-',
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'dgrid_param_dict')
        self.button_delete_param_dict.SetToolTip(tooltip_delete_row)
        bSizer3.Add(self.button_delete_param_dict, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        self.button_clear_param_dict = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, label_reset,
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'cgrid_param_dict')
        bSizer3.Add(self.button_clear_param_dict, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

        bSizer2.Add(sbSizer1, 0, wx.EXPAND, 5)

        self.spinCtrls_zorder_scatter = []
        self.filePickers_scatter = []
        self.buttons_scatter_open_file = []
        self.textCtrls_scatter_column_x = []
        self.textCtrls_scatter_column_y = []
        self.textCtrls_scatter_err_column_xp = []
        self.textCtrls_scatter_err_column_xn = []
        self.textCtrls_scatter_err_column_yp = []
        self.textCtrls_scatter_err_column_yn = []
        self.textCtrls_scatter_every = []
        self.textCtrls_scatter_label = []
        self.choices_scatter_marker = []
        self.textCtrls_scatter_marker_size = []
        self.colourPickers_scatter_marker = []
        self.checkBoxes_scatter_marker_edge = []
        self.colourPickers_scatter_marker_edge = []
        self.choices_scatter_line_style = []
        self.colourPickers_scatter_line = []
        self.textCtrls_scatter_column_z = []
        self.choices_show_z_by = []
        self.textCtrls_scatter_marker_size_ratio = []

        for i in range(n_scatter):
            sbSizer1 = wx.StaticBoxSizer(
                wx.StaticBox(self.scrolledWindow_graph, wx.ID_ANY, _(u'散布図') + (u'{:d}'.format(i + 1))), wx.VERTICAL)

            bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, u'Z-order：',
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

            self.spinCtrls_zorder_scatter.append(wx.SpinCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
                wx.DefaultPosition, wx.Size(spinCtrlWidth, -1), wx.SP_ARROW_KEYS, zorder_min, zorder_max, zorder_scatter + i))
            self.spinCtrls_zorder_scatter[i].SetToolTip(tooltip_zorder)
            bSizer3.Add(self.spinCtrls_zorder_scatter[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'ファイル：'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

            self.filePickers_scatter.append(wx.FilePickerCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
                _(u'ファイルを開く'), u'*.*', wx.DefaultPosition, wx.DefaultSize,
                wx.FLP_USE_TEXTCTRL | wx.FLP_OPEN | wx.FLP_FILE_MUST_EXIST | wx.FLP_SMALL, name = 'filePicker_scatter{:d}'.format(i)))
            self.filePickers_scatter[i].SetForegroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOWTEXT))
            self.filePickers_scatter[i].SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOW))
            self.filePickers_scatter[i].SetToolTip(tooltip_drag_and_drop + '\n\n' + tooltip_equation_scatter)
            self.filePickers_scatter[i].SetDropTarget(MyFileDropTarget(self.filePickers_scatter[i]))
            bSizer3.Add(self.filePickers_scatter[i], 1, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            self.buttons_scatter_open_file.append(wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'開く'),
                wx.DefaultPosition, wx.DefaultSize, 0, name = 'bfilePicker_scatter{:d}'.format(i)))
            self.buttons_scatter_open_file[i].SetToolTip(tooltip_open_file)
            bSizer3.Add(self.buttons_scatter_open_file[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

            sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

            bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'xの列：'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

            self.textCtrls_scatter_column_x.append(wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'1',
                wx.DefaultPosition, wx.Size(60, -1), 0))
            self.textCtrls_scatter_column_x[i].SetToolTip(tooltip_column)
            bSizer3.Add(self.textCtrls_scatter_column_x[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'yの列：'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

            self.textCtrls_scatter_column_y.append(wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'2',
                wx.DefaultPosition, wx.Size(60, -1), 0))
            self.textCtrls_scatter_column_y[i].SetToolTip(tooltip_column)
            bSizer3.Add(self.textCtrls_scatter_column_y[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'xの-, +側誤差棒の列：'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

            self.textCtrls_scatter_err_column_xn.append(wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'',
                wx.DefaultPosition, wx.Size(58, -1), 0))
            self.textCtrls_scatter_err_column_xn[i].SetToolTip(tooltip_err_column + u"\n\n" + tooltip_column)
            bSizer3.Add(self.textCtrls_scatter_err_column_xn[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u','),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

            self.textCtrls_scatter_err_column_xp.append(wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'',
                wx.DefaultPosition, wx.Size(58, -1), 0))
            self.textCtrls_scatter_err_column_xp[i].SetToolTip(tooltip_err_column + u"\n\n" + tooltip_column)
            bSizer3.Add(self.textCtrls_scatter_err_column_xp[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'yの-, +側誤差棒の列：'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

            self.textCtrls_scatter_err_column_yn.append(wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'',
                wx.DefaultPosition, wx.Size(58, -1), 0))
            self.textCtrls_scatter_err_column_yn[i].SetToolTip(tooltip_err_column + u"\n\n" + tooltip_column)
            bSizer3.Add(self.textCtrls_scatter_err_column_yn[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u','),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

            self.textCtrls_scatter_err_column_yp.append(wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'',
                wx.DefaultPosition, wx.Size(58, -1), 0))
            self.textCtrls_scatter_err_column_yp[i].SetToolTip(tooltip_err_column + u"\n\n" + tooltip_column)
            bSizer3.Add(self.textCtrls_scatter_err_column_yp[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

            bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'読み込み：'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

            self.textCtrls_scatter_every.append(wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'1',
                wx.DefaultPosition, wx.Size(60, -1), 0))
            bSizer3.Add(self.textCtrls_scatter_every[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'行ごと'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'凡例に書くラベル：'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            self.textCtrls_scatter_label.append(wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
                wx.DefaultPosition, wx.Size(-1, -1), 0))
            self.textCtrls_scatter_label[i].SetToolTip(tooltip_tex)
            bSizer3.Add(self.textCtrls_scatter_label[i], 1, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

            sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

            bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'記号：'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

            self.choices_scatter_marker.append(wx.Choice(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.DefaultPosition,
                wx.DefaultSize, markers_wx, 0))
            self.choices_scatter_marker[i].SetSelection(i%(len(markers_wx) - 1))
            bSizer3.Add(self.choices_scatter_marker[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'大きさ：'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

            self.textCtrls_scatter_marker_size.append(wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
                wx.DefaultPosition, wx.DefaultSize, 0))
            self.textCtrls_scatter_marker_size[i].SetToolTip(_(u'空白の場合，指定しません．'))
            self.textCtrls_scatter_marker_size[i].SetMinSize(wx.Size(60, -1))
            bSizer3.Add(self.textCtrls_scatter_marker_size[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, u'pt', wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'記号の塗色：'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

            self.colourPickers_scatter_marker.append(wx.ColourPickerCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY,
                wx.Colour(255, 255, 255), wx.DefaultPosition, wx.Size(30, -1), wx.CLRP_DEFAULT_STYLE))
            bSizer3.Add(self.colourPickers_scatter_marker[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            self.checkBoxes_scatter_marker_edge.append(wx.CheckBox(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'記号の輪郭色：'),
                wx.DefaultPosition, wx.DefaultSize, 0))
            self.checkBoxes_scatter_marker_edge[i].SetToolTip(_(u'チェックを外すと輪郭を付けません．'))
            self.checkBoxes_scatter_marker_edge[i].SetValue(True)
            bSizer3.Add(self.checkBoxes_scatter_marker_edge[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            self.colourPickers_scatter_marker_edge.append(wx.ColourPickerCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY,
                wx.Colour(255, 0, 0), wx.DefaultPosition, wx.Size(30, -1), wx.CLRP_DEFAULT_STYLE))
            bSizer3.Add(self.colourPickers_scatter_marker_edge[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'線種：'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

            self.choices_scatter_line_style.append(wx.Choice(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.DefaultPosition,
                wx.DefaultSize, line_styles_wx, 0))
            self.choices_scatter_line_style[i].SetSelection(len(line_styles_wx) - 1)
            bSizer3.Add(self.choices_scatter_line_style[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'線色：'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

            self.colourPickers_scatter_line.append(wx.ColourPickerCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY,
                wx.Colour(255, 0, 0), wx.DefaultPosition, wx.Size(30, -1), wx.CLRP_DEFAULT_STYLE))
            bSizer3.Add(self.colourPickers_scatter_line[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

            sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

            bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'zの列：'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

            self.textCtrls_scatter_column_z.append(wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'',
                wx.DefaultPosition, wx.Size(60, -1), 0))
            self.textCtrls_scatter_column_z[i].SetToolTip(_(u'空白の場合，普通の散布図になります．\n') +  tooltip_column)
            bSizer3.Add(self.textCtrls_scatter_column_z[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'表し方：'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

            self.choices_show_z_by.append(wx.Choice(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.DefaultPosition,
                wx.DefaultSize, by_what_show_z_wx, 0))
            self.choices_show_z_by[i].SetSelection(0)
            bSizer3.Add(self.choices_show_z_by[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'z最大の記号の大きさ＝'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

            self.textCtrls_scatter_marker_size_ratio.append(wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
                wx.DefaultPosition, wx.DefaultSize, 0))
            self.textCtrls_scatter_marker_size_ratio[i].SetToolTip(
                _(u'空白の場合，指定しません．\nz最小の記号の大きさは「大きさ」で指定します．'))
            self.textCtrls_scatter_marker_size_ratio[i].SetMinSize(wx.Size(60, -1))
            bSizer3.Add(self.textCtrls_scatter_marker_size_ratio[i], 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

            staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'×z最小の記号の大きさ'),
                wx.DefaultPosition, wx.DefaultSize, 0)
            staticText1.Wrap(-1)
            bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

            sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

            bSizer2.Add(sbSizer1, 0, wx.EXPAND, 5)

        sbSizer1 = wx.StaticBoxSizer(wx.StaticBox(self.scrolledWindow_graph, wx.ID_ANY, _(u'ベクトル線図')), wx.VERTICAL)

        bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, u'Z-order：',
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.spinCtrl_zorder_vector = wx.SpinCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString, wx.DefaultPosition,
            wx.Size(spinCtrlWidth, -1), wx.SP_ARROW_KEYS, zorder_min, zorder_max, zorder_vector)
        self.spinCtrl_zorder_vector.SetToolTip(tooltip_zorder)
        bSizer3.Add(self.spinCtrl_zorder_vector, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'ファイル：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.filePicker_vector = wx.FilePickerCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
            _(u'ファイルを開く'), u'*.*', wx.DefaultPosition, wx.DefaultSize,
            wx.FLP_USE_TEXTCTRL | wx.FLP_OPEN | wx.FLP_FILE_MUST_EXIST | wx.FLP_SMALL, name = 'filePicker_vector')
        self.filePicker_vector.SetForegroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOWTEXT))
        self.filePicker_vector.SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOW))
        self.filePicker_vector.SetToolTip(tooltip_drag_and_drop + '\n\n' +
            _(u"次の書式を用いた数式のプロットも可能です．\n" +
              u"分割数の前の'/'を'L/'にすれば，対数軸上で等分割にできます：\n" +
              u" [uの数式, vの数式], xの変数 = [最小値, 最大値]/分割数, yの変数 = [最小値, 最大値]/分割数\n例）") +
              u"[cos(2.0*x), sin(y)], x = [-1.0, 1.0]/100, y = [-1.00, 1.0]/100")
        self.filePicker_vector.SetDropTarget(MyFileDropTarget(self.filePicker_vector))
        bSizer3.Add(self.filePicker_vector, 1, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        self.button_vector_open_file = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'開く'),
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'bfilePicker_vector')
        self.button_vector_open_file.SetToolTip(tooltip_open_file)
        bSizer3.Add(self.button_vector_open_file, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

        bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'xの列：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_vector_column_x = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'1',
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_vector_column_x.SetToolTip(tooltip_column)
        bSizer3.Add(self.textCtrl_vector_column_x, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'yの列：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_vector_column_y = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'2',
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_vector_column_y.SetToolTip(tooltip_column)
        bSizer3.Add(self.textCtrl_vector_column_y, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'uの列：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_vector_column_u = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'3',
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_vector_column_u.SetToolTip(tooltip_column)
        bSizer3.Add(self.textCtrl_vector_column_u, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'vの列：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_vector_column_v = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'4',
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_vector_column_v.SetToolTip(tooltip_column)
        bSizer3.Add(self.textCtrl_vector_column_v, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'読み込み：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_vector_every = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'1',
            wx.DefaultPosition, wx.Size(60, -1), 0)
        bSizer3.Add(self.textCtrl_vector_every, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'行ごと'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

        bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'矢印色：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.colourPicker_vector = wx.ColourPickerCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.Colour(255, 0, 0),
            wx.DefaultPosition, wx.Size(30, -1), wx.CLRP_DEFAULT_STYLE)
        bSizer3.Add(self.colourPicker_vector, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        self.choice_vector_color = wx.Choice(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.DefaultPosition,
            wx.DefaultSize, [_(u'虹色なし'), _(u'虹色あり'), _(u'虹色+長さ1')], 0)
        self.choice_vector_color.SetSelection(0)
        self.choice_vector_color.SetToolTip(_(u'「虹色なし」は単色指定．\n' +
            u'「虹色あり」はベクトルの大きさに応じて矢印の色を虹色で変える．\n' +
            u'「虹色+長さ1」はベクトルの大きさに応じて矢印の色を虹色で変え，表示する矢印の長さを全て1にする．'))
        bSizer3.Add(self.choice_vector_color, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'倍率：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_vector_scale = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'1.0',
            wx.DefaultPosition, wx.DefaultSize, 0)
        self.textCtrl_vector_scale.SetToolTip(_(u'ベクトルの長さにかける倍率．大きいほど長く表示される．'))
        self.textCtrl_vector_scale.SetMinSize(wx.Size(60, -1))
        bSizer3.Add(self.textCtrl_vector_scale, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

        bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY,
            (_(u'!!Python 3でのみ有効!!→') if sys.version_info.major <= 2 else u'') + _(u'凡例の横座標：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_vector_legend_x = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'0.8',
            wx.DefaultPosition, wx.DefaultSize, 0)
        self.textCtrl_vector_legend_x.SetToolTip(_(u'凡例矢印先端の横座標，0/1だとグラフ枠の左/右端'))
        self.textCtrl_vector_legend_x.SetMinSize(wx.Size(60, -1))
        if sys.version_info.major <= 2:
            self.textCtrl_vector_legend_x.Disable()
        bSizer3.Add(self.textCtrl_vector_legend_x, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'縦座標：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_vector_legend_y = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'1.03',
            wx.DefaultPosition, wx.DefaultSize, 0)
        self.textCtrl_vector_legend_y.SetToolTip(_(u'凡例矢印先端の縦座標，0/1だとグラフ枠の下/上端'))
        self.textCtrl_vector_legend_y.SetMinSize(wx.Size(60, -1))
        if sys.version_info.major <= 2:
            self.textCtrl_vector_legend_y.Disable()
        bSizer3.Add(self.textCtrl_vector_legend_y, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'矢印長さ：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_vector_legend_u = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'1.0',
            wx.DefaultPosition, wx.DefaultSize, 0)
        self.textCtrl_vector_legend_u.SetMinSize(wx.Size(60, -1))
        if sys.version_info.major <= 2:
            self.textCtrl_vector_legend_u.Disable()
        bSizer3.Add(self.textCtrl_vector_legend_u, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'ラベル：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_vector_legend_label = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
            wx.DefaultPosition, wx.Size(-1, -1), 0)
        self.textCtrl_vector_legend_label.SetToolTip(tooltip_tex)
        if sys.version_info.major <= 2:
            self.textCtrl_vector_legend_label.Disable()
        bSizer3.Add(self.textCtrl_vector_legend_label, 1, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

        bSizer2.Add(sbSizer1, 0, wx.EXPAND, 5)

        sbSizer1 = wx.StaticBoxSizer(wx.StaticBox(self.scrolledWindow_graph, wx.ID_ANY, _(u'等高線')), wx.VERTICAL)

        bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, u'Z-order：',
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.spinCtrl_zorder_contour = wx.SpinCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
            wx.DefaultPosition, wx.Size(spinCtrlWidth, -1), wx.SP_ARROW_KEYS, zorder_min, zorder_max, zorder_contour)
        self.spinCtrl_zorder_contour.SetToolTip(tooltip_zorder)
        bSizer3.Add(self.spinCtrl_zorder_contour, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'ファイル：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.filePicker_contour = wx.FilePickerCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.EmptyString,
            _(u'ファイルを開く'), u'*.*', wx.DefaultPosition, wx.DefaultSize,
            wx.FLP_USE_TEXTCTRL | wx.FLP_OPEN | wx.FLP_FILE_MUST_EXIST | wx.FLP_SMALL, name = 'filePicker_contour')
        self.filePicker_contour.SetForegroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOWTEXT))
        self.filePicker_contour.SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOW))
        self.filePicker_contour.SetToolTip(tooltip_drag_and_drop + '\n\n' +
            _(u"次の書式を用いた数式のプロットも可能です．\n" +
              u"分割数の前の'/'を'L/'にすれば，対数軸上で等分割にできます：\n" +
              u" 数式, xの変数 = [最小値, 最大値]/分割数, yの変数 = [最小値, 最大値]/分割数\n例）") +
              u"sqrt(x)*y**2, x = [-1.0, 1.0]/100, y = [-1.0, 1.0]/100")
        self.filePicker_vector.SetDropTarget(MyFileDropTarget(self.filePicker_vector))
        self.filePicker_contour.SetDropTarget(MyFileDropTarget(self.filePicker_contour))
        bSizer3.Add(self.filePicker_contour, 1, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        self.button_contour_open_file = wx.Button(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'開く'),
            wx.DefaultPosition, wx.DefaultSize, 0, name = 'bfilePicker_contour')
        self.button_contour_open_file.SetToolTip(tooltip_open_file)
        bSizer3.Add(self.button_contour_open_file, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

        bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

        self.checkBox_contour_grid_pattern = wx.CheckBox(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'碁盤の目状'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        self.checkBox_contour_grid_pattern.SetToolTip(_(u'データが碁盤の目状に並んでいる時にチェックを付けると，描画が高速になります．'))
        self.checkBox_contour_grid_pattern.SetValue(False)
        bSizer3.Add(self.checkBox_contour_grid_pattern, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        self.checkBox_contour_show_triangle = wx.CheckBox(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'三角形を表示'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        self.checkBox_contour_show_triangle.SetToolTip(_(u'「碁盤の目状」にチェックがついていない時に行う，デローニー分割した三角形を表示します．'))
        self.checkBox_contour_show_triangle.SetValue(False)
        bSizer3.Add(self.checkBox_contour_show_triangle, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)
        self.checkBox_contour_grid_patternOnCheck(None)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'xの列：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_contour_column_x = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'1',
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_contour_column_x.SetToolTip(tooltip_column)
        bSizer3.Add(self.textCtrl_contour_column_x, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'yの列：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_contour_column_y = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'2',
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_contour_column_y.SetToolTip(tooltip_column)
        bSizer3.Add(self.textCtrl_contour_column_y, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'zの列：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_contour_column_z = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'3',
            wx.DefaultPosition, wx.Size(60, -1), 0)
        self.textCtrl_contour_column_z.SetToolTip(tooltip_column)
        bSizer3.Add(self.textCtrl_contour_column_z, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'読み込み：'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.textCtrl_contour_every = wx.TextCtrl(sbSizer1.GetStaticBox(), wx.ID_ANY, u'1',
            wx.DefaultPosition, wx.Size(60, -1), 0)
        bSizer3.Add(self.textCtrl_contour_every, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM, 5)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'行ごと'), wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

        bSizer3 = wx.BoxSizer(wx.HORIZONTAL)

        staticText1 = wx.StaticText(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'塗り：'), wx.DefaultPosition, wx.DefaultSize, 0)
        staticText1.Wrap(-1)
        bSizer3.Add(staticText1, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.LEFT, 5)

        self.choice_contour_paint = wx.Choice(sbSizer1.GetStaticBox(), wx.ID_ANY, wx.DefaultPosition,
            wx.DefaultSize, paint_styles_wx, 0)
        self.choice_contour_paint.SetSelection(0)
        bSizer3.Add(self.choice_contour_paint, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        self.checkBox_contour_smooth_paint = wx.CheckBox(sbSizer1.GetStaticBox(), wx.ID_ANY, _(u'滑らかな面塗り'),
            wx.DefaultPosition, wx.DefaultSize, 0)
        self.checkBox_contour_smooth_paint.SetValue(True)
        bSizer3.Add(self.checkBox_contour_smooth_paint, 0, wx.ALIGN_CENTER_VERTICAL | wx.TOP | wx.BOTTOM | wx.RIGHT, 5)

        sbSizer1.Add(bSizer3, 0, wx.EXPAND, 5)

        bSizer2.Add(sbSizer1, 0, wx.EXPAND, 5)

        self.filePickers = (self.filePickers_scatter + [self.filePicker_png, self.filePicker_vector, self.filePicker_contour])

        self.scrolledWindow_graph.SetSizer(bSizer2)
        self.scrolledWindow_graph.Layout()
        bSizer2.Fit(self.scrolledWindow_graph)
        bSizer1.Add(self.scrolledWindow_graph, 1, wx.EXPAND |wx.ALL, 5)

        self.SetSizer(bSizer1)
        self.Layout()

        self.Centre(wx.BOTH)

        self.menubar = wx.MenuBar(0)

        self.menu_file = wx.Menu()
        self.menuItem_open = wx.MenuItem(self.menu_file, wx.ID_ANY, _(u'開く') + '\tCtrl+O', wx.EmptyString, wx.ITEM_NORMAL)
        self.menu_file.Append(self.menuItem_open)
        self.menuItem_save = wx.MenuItem(self.menu_file, wx.ID_ANY, _(u'保存') + '\tCtrl+S', wx.EmptyString, wx.ITEM_NORMAL)
        self.menu_file.Append(self.menuItem_save)
        self.menuItem_close = wx.MenuItem(self.menu_file, wx.ID_ANY, _(u'閉じる') + '\tCtrl+W', wx.EmptyString, wx.ITEM_NORMAL)
        self.menu_file.Append(self.menuItem_close)
        self.menu_file.AppendSeparator()
        self.menuItem_save_as = wx.MenuItem(self.menu_file, wx.ID_ANY, _(u'別名で保存') + '\tShift+Ctrl+S',
            wx.EmptyString, wx.ITEM_NORMAL)
        self.menu_file.Append(self.menuItem_save_as)
        self.menu_file.AppendSeparator()
        self.menuItem_save_as_script = wx.MenuItem(self.menu_file, wx.ID_ANY, _(u'pythonスクリプト形式で保存') + '\tCtrl+K',
            wx.EmptyString, wx.ITEM_NORMAL)
        self.menu_file.Append(self.menuItem_save_as_script)
        self.menu_file.AppendSeparator()
        self.menuItem_make_template = wx.MenuItem(self.menu_file, wx.ID_ANY, _(u'テンプレートを作る') + '\tCtrl+T',
            wx.EmptyString, wx.ITEM_NORMAL)
        self.menu_file.Append(self.menuItem_make_template)
        self.menubar.Append(self.menu_file, _(u'設定ファイル') + '(&F)')

        self.menu_edit = wx.Menu()
        self.menuItem_cut = wx.MenuItem(self.menu_edit, wx.ID_ANY, _(u'カット') + '\tCtrl+X',
            wx.EmptyString, wx.ITEM_NORMAL)
        self.menu_edit.Append(self.menuItem_cut)
        self.menuItem_copy = wx.MenuItem(self.menu_edit, wx.ID_ANY, _(u'コピー') + '\tCtrl+C',
            wx.EmptyString, wx.ITEM_NORMAL)
        self.menu_edit.Append(self.menuItem_copy)
        self.menuItem_paste = wx.MenuItem(self.menu_edit, wx.ID_ANY, _(u'ペースト') + '\tCtrl+V',
            wx.EmptyString, wx.ITEM_NORMAL)
        self.menu_edit.Append(self.menuItem_paste)
        self.menubar.Append(self.menu_edit, _(u'編集') + '(&E)')

        self.menu_plot = wx.Menu()
        self.menuItem_plot = wx.MenuItem(self.menu_plot, wx.ID_ANY, _(u'グラフをプロット') + '\tCtrl+P',
            wx.EmptyString, wx.ITEM_NORMAL)
        self.menu_plot.Append(self.menuItem_plot)
        self.menuItem_plot_direct = wx.MenuItem(self.menu_plot, wx.ID_ANY, _(u'設定ファイルから直接プロット') + '\tShift+Ctrl+P',
            wx.EmptyString, wx.ITEM_NORMAL)
        self.menu_plot.Append(self.menuItem_plot_direct)
        self.menubar.Append(self.menu_plot, _(u'プロット') + '(&P)')

        self.menu_help = wx.Menu()
        self.menuItem_update = wx.MenuItem(self.menu_help, wx.ID_ANY, _(u'アップデート'), wx.EmptyString, wx.ITEM_NORMAL)
        self.menu_help.Append(self.menuItem_update)
        self.menuItem_movie = wx.MenuItem(self.menu_help, wx.ID_ANY, _(u'使い方の動画'), wx.EmptyString, wx.ITEM_NORMAL)
        self.menu_help.Append(self.menuItem_movie)
        self.menubar.Append(self.menu_help, _(u'ヘルプ') + '(&H)')

        self.SetMenuBar(self.menubar)

        # Connect Events
        self.Bind(wx.EVT_CLOSE, self.FrameMainOnClose)
        self.filePicker_setting.Bind(wx.EVT_FILEPICKER_CHANGED, self.OnFileChanged)
        for i in self.filePickers:
            i.Bind(wx.EVT_FILEPICKER_CHANGED, self.OnFileChanged)
            i.Bind(wx.EVT_CHAR_HOOK, self.OnCharHook)
        self.grid_graph.Bind(wx.EVT_CHAR_HOOK, self.OnCharHook)
        self.textCtrl_title.Bind(wx.EVT_CHAR_HOOK, self.OnCharHook)
        self.grid_text.Bind(wx.EVT_CHAR_HOOK, self.OnCharHook)
        for i in self.textCtrls_scatter_label:
            i.Bind(wx.EVT_CHAR_HOOK, self.OnCharHook)
        self.button_setting_open_file.Bind(wx.EVT_BUTTON, self.button_openOnButtonClick)
        self.button_append_text.Bind(wx.EVT_BUTTON, self.button_append_gridOnButtonClick)
        self.button_delete_text.Bind(wx.EVT_BUTTON, self.button_delete_gridOnButtonClick)
        self.button_clear_text.Bind(wx.EVT_BUTTON, self.button_clear_gridOnButtonClick)
        self.button_append_param_dict.Bind(wx.EVT_BUTTON, self.button_append_gridOnButtonClick)
        self.button_delete_param_dict.Bind(wx.EVT_BUTTON, self.button_delete_gridOnButtonClick)
        self.button_clear_param_dict.Bind(wx.EVT_BUTTON, self.button_clear_gridOnButtonClick)
        self.button_append_arrow.Bind(wx.EVT_BUTTON, self.button_append_gridOnButtonClick)
        self.button_delete_arrow.Bind(wx.EVT_BUTTON, self.button_delete_gridOnButtonClick)
        self.button_clear_arrow.Bind(wx.EVT_BUTTON, self.button_clear_gridOnButtonClick)
        self.button_append_ellipse.Bind(wx.EVT_BUTTON, self.button_append_gridOnButtonClick)
        self.button_delete_ellipse.Bind(wx.EVT_BUTTON, self.button_delete_gridOnButtonClick)
        self.button_clear_ellipse.Bind(wx.EVT_BUTTON, self.button_clear_gridOnButtonClick)
        self.button_append_polygon.Bind(wx.EVT_BUTTON, self.button_append_gridOnButtonClick)
        self.button_delete_polygon.Bind(wx.EVT_BUTTON, self.button_delete_gridOnButtonClick)
        self.button_clear_polygon.Bind(wx.EVT_BUTTON, self.button_clear_gridOnButtonClick)
        self.button_append_naca4.Bind(wx.EVT_BUTTON, self.button_append_gridOnButtonClick)
        self.button_delete_naca4.Bind(wx.EVT_BUTTON, self.button_delete_gridOnButtonClick)
        self.button_clear_naca4.Bind(wx.EVT_BUTTON, self.button_clear_gridOnButtonClick)
        for i in self.buttons_scatter_open_file:
            i.Bind(wx.EVT_BUTTON, self.button_openOnButtonClick)
        self.button_vector_open_file.Bind(wx.EVT_BUTTON, self.button_openOnButtonClick)
        self.button_contour_open_file.Bind(wx.EVT_BUTTON, self.button_openOnButtonClick)
        self.checkBox_contour_grid_pattern.Bind(wx.EVT_CHECKBOX, self.checkBox_contour_grid_patternOnCheck)
        self.Bind(wx.EVT_MENU, self.menuItem_openOnMenuSelection, id = self.menuItem_open.GetId())
        self.Bind(wx.EVT_MENU, self.FrameMainOnClose, id = self.menuItem_close.GetId())
        self.Bind(wx.EVT_MENU, self.menuItem_saveOnMenuSelection, id = self.menuItem_save.GetId())
        self.Bind(wx.EVT_MENU, self.menuItem_save_asOnMenuSelection, id = self.menuItem_save_as.GetId())
        self.Bind(wx.EVT_MENU, self.menuItem_save_as_scriptOnMenuSelection, id = self.menuItem_save_as_script.GetId())
        self.Bind(wx.EVT_MENU, self.menuItem_make_templateOnMenuSelection, id = self.menuItem_make_template.GetId())
        self.Bind(wx.EVT_MENU, self.menuItem_cutOnMenuSelection, id = self.menuItem_cut.GetId())
        self.Bind(wx.EVT_MENU, self.menuItem_copyOnMenuSelection, id = self.menuItem_copy.GetId())
        self.Bind(wx.EVT_MENU, self.menuItem_pasteOnMenuSelection, id = self.menuItem_paste.GetId())
        self.Bind(wx.EVT_MENU, self.menuItem_plotOnMenuSelection, id = self.menuItem_plot.GetId())
        self.Bind(wx.EVT_MENU, self.menuItem_plot_directOnMenuSelection, id = self.menuItem_plot_direct.GetId())
        self.Bind(wx.EVT_MENU, self.menuItem_updateOnMenuSelection, id = self.menuItem_update.GetId())
        self.Bind(wx.EVT_MENU, self.menuItem_movieOnMenuSelection, id = self.menuItem_movie.GetId())

        backup_path_old = correct_file_name_in_unicode(os.path.join(os.path.dirname(
            os.path.realpath(decode_if_necessary(__file__))), u'backup_matplotlib.txt')) # unicode
        self.backup_path = correct_file_name_in_unicode(os.path.join(os.path.dirname(
            os.path.realpath(decode_if_necessary(__file__))), u'backup_matplotlibwx.txt')) # unicode
        if os.path.isfile(backup_path_old) and not os.path.isfile(self.backup_path):
            os.rename(backup_path_old, self.backup_path) # can overwrite
        self.cwd = correct_file_name_in_unicode(decode_if_necessary(os.getcwd())) # unicode

    def __del__(self):
        pass

    def FrameMainOnClose(self, event):
        with wx.MessageDialog(self,
            _(u'設定を保存しますか？') if self.filePicker_setting.GetPath() == u''
                else _(u'設定を ') + self.filePicker_setting.GetPath() + _(u' に保存しますか？'),
            _(u'保存'), style = wx.YES_NO | wx.CANCEL | wx.ICON_EXCLAMATION) as md:
            r = md.ShowModal()
        if r == wx.ID_YES:
            self.menuItem_saveOnMenuSelection(None)
        elif r == wx.ID_CANCEL:
            return
        quit()

    def OnCharHook(self, event):
        if event.GetKeyCode() == 165: # Yen mark
            if event.GetModifiers() & wx.MOD_CONTROL:
                event.GetEventObject().WriteText(u'¥')
            elif event.GetModifiers() & wx.MOD_SHIFT:
                self.GetEventObject().WriteText(u'|')
            else:
                event.GetEventObject().WriteText(u'\\')
        else:
            event.Skip()

    def OnFileChanged(self, event):
        p = correct_file_name_in_unicode(event.GetEventObject().GetPath()) # unicode
        if p == u'' or pat_eq_plot.match(p):
            event.GetEventObject().GetTextCtrl().SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOW))
        elif os.path.isfile(p):
            event.GetEventObject().SetPath(p)
            if event.GetEventObject().GetWindowStyle() & wx.FLP_FILE_MUST_EXIST:
                event.GetEventObject().GetTextCtrl().SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOW))
                if event.GetEventObject().GetName() == 'filePicker_setting':
                    self.load_settings(p)
            self.cwd = os.path.dirname(p) # unicode
            for i in self.filePickers:
                i.SetInitialDirectory(self.cwd)
        elif event.GetEventObject().GetWindowStyle() & wx.FLP_FILE_MUST_EXIST:
            event.GetEventObject().GetTextCtrl().SetBackgroundColour(wx.RED)

    def button_openOnButtonClick(self, event):
        p = self.FindWindowByName(event.GetEventObject().GetName()[1:], event.GetEventObject().GetParent()).GetPath() # unicode
        if p == u'':
            return
        if p.endswith(u'.zip'):
            with zipfile.ZipFile(p) as zf:
                zf.extractall(os.path.dirname(p))
                p = os.path.join(os.path.dirname(p), zf.namelist()[0])
        if sys.platform == 'win32':
            os.startfile(p)
        elif sys.platform == 'darwin':
            subprocess.call(u'open "{}"'.format(p), shell = True)
        else:
            subprocess.call(u'xdg-open "{}"'.format(p), shell = True)

    def button_insert_gridOnButtonClick(self, event):
        self.FindWindowByName(event.GetEventObject().GetName()[1:],
            event.GetEventObject().GetParent()).InsertRows(0, 1)

    def button_append_gridOnButtonClick(self, event):
        self.FindWindowByName(event.GetEventObject().GetName()[1:],
            event.GetEventObject().GetParent()).AppendRows(1)

    def button_delete_gridOnButtonClick(self, event):
        grid = self.FindWindowByName(event.GetEventObject().GetName()[1:], event.GetEventObject().GetParent())
        row, col = grid.GetGridCursorRow(), grid.GetGridCursorCol()
        grid.DeleteRows(row, 1)
        if grid.GetNumberRows() > 0:
            grid.SetGridCursor(min(row, grid.GetNumberRows() - 1), col)

    def button_clear_gridOnButtonClick(self, event):
        grid = self.FindWindowByName(event.GetEventObject().GetName()[1:], event.GetEventObject().GetParent())
        grid.ClearGrid()
        grid.ForceRefresh()

    def button_increase_gridOnButtonClick(self, event):
        grid = self.FindWindowByName(event.GetEventObject().GetName()[1:], event.GetEventObject().GetParent())
        row, col = grid.GetGridCursorRow(), grid.GetGridCursorCol()
        if row != len(grid.table.data) - 1:
            grid.table.data[row], grid.table.data[row + 1] = grid.table.data[row + 1], grid.table.data[row]
            grid.SetGridCursor(row + 1, col)

    def button_decrease_gridOnButtonClick(self, event):
        grid = self.FindWindowByName(event.GetEventObject().GetName()[1:], event.GetEventObject().GetParent())
        row, col = grid.GetGridCursorRow(), grid.GetGridCursorCol()
        if row != 0:
            grid.table.data[row], grid.table.data[row - 1] = grid.table.data[row - 1], grid.table.data[row]
            grid.SetGridCursor(row - 1, col)

    def checkBox_contour_grid_patternOnCheck(self, event):
        if self.checkBox_contour_grid_pattern.GetValue():
            self.checkBox_contour_show_triangle.Disable()
        else:
            self.checkBox_contour_show_triangle.Enable()

    def load_settings(self, path):
        s, path = load_plot_settings(path)
        dir_name = os.path.dirname(path) # unicode
        try:
            self.filePicker_png.SetPath(
                correct_file_name_in_unicode(os.path.join(dir_name, s[0]['png_file_name']))
                if 'png_file_name' in s[0] and s[0]['png_file_name'] is not None else u'')
        except:
            print(sys.exc_info())
        for i, j in ((self.textCtrl_fig_size_w, 0), (self.textCtrl_fig_size_h, 1)):
            try:
                i.SetValue(u'{:g}'.format(s[0]['fig_size'][j]) if 'fig_size' in s[0] and
                    s[0]['fig_size'] is not None and s[0]['fig_size'][j] is not None else u'')
            except:
                print(sys.exc_info())
        for i, j in ((self.textCtrl_skip, 'skip'), (self.textCtrl_delimiter, 'delimiter'),
                     (self.textCtrl_terminator, 'terminator'), (self.textCtrl_title, 'title')):
            try:
                i.SetValue(s[0][j] if j in s[0] and s[0][j] is not None else u'')
            except:
                print(sys.exc_info())
        for i, j, k, l in ((self.grid_graph.table.ROW_X, self.grid_graph.table.COL_FROM, 0, 0),
                           (self.grid_graph.table.ROW_X, self.grid_graph.table.COL_TO,   0, 1),
                           (self.grid_graph.table.ROW_Y, self.grid_graph.table.COL_FROM, 1, 0),
                           (self.grid_graph.table.ROW_Y, self.grid_graph.table.COL_TO,   1, 1),
                           (self.grid_graph.table.ROW_Z, self.grid_graph.table.COL_FROM, 2, 0),
                           (self.grid_graph.table.ROW_Z, self.grid_graph.table.COL_TO,   2, 1)):
            try:
                self.grid_graph.table.data[i][j] = (s[0]['ranges'][k][l] if 'ranges' in s[0] and
                    s[0]['ranges'] is not None and s[0]['ranges'][k] is not None else None)
            except:
                print(sys.exc_info())
        for i, j in ((self.grid_graph.table.ROW_X, 0), (self.grid_graph.table.ROW_Y, 1), (self.grid_graph.table.ROW_Z, 2)):
            try:
                self.grid_graph.table.data[i][self.grid_graph.table.COL_TICKS] = (s[0]['ticks'][j]
                    if 'ticks' in s[0] and s[0]['ticks'] is not None else None)
            except:
                print(sys.exc_info())
            try:
                self.grid_graph.table.data[i][self.grid_graph.table.COL_LOG_SCALE] = (s[0]['log_scale'][j]
                    if 'log_scale' in s[0] else False)
            except:
                print(sys.exc_info())
            try:
                self.grid_graph.table.data[i][self.grid_graph.table.COL_LABEL] = (s[0]['labels'][j]
                    if 'labels' in s[0] and s[0]['labels'] is not None else None)
            except:
                print(sys.exc_info())
        self.grid_graph.ForceRefresh()
        try:
            self.checkBox_x_grid.SetValue(s[0]['grids'][0] if 'grids' in s[0] else False)
        except:
            print(sys.exc_info())
        try:
            self.checkBox_y_grid.SetValue(s[0]['grids'][1] if 'grids' in s[0] else False)
        except:
            print(sys.exc_info())
        try:
            self.comboBox_aspect.SetValue((u'auto' if s[0]['aspect'] is None
                else (s[0]['aspect'] if s[0]['aspect'] in (u'auto', u'equal')
                else u'{:g}'.format(s[0]['aspect']))) if 'aspect' in s[0] else u'auto')
        except:
            print(sys.exc_info())
        try:
            if 'legend_location' in s[0] and s[0]['legend_location'] is not None:
                self.textCtrl_legend_left.SetValue(u'{:g}'.format(s[0]['legend_location'][0])
                    if s[0]['legend_location'][0] is not None else u'')
                self.textCtrl_legend_top.SetValue(u'{:g}'.format(s[0]['legend_location'][1])
                    if s[0]['legend_location'][1] is not None else u'')
            else:
                self.textCtrl_legend_left.SetValue(u'')
                self.textCtrl_legend_top.SetValue(u'')
        except:
            print(sys.exc_info())
        try:
            if 'graph_edges' in s[0] and s[0]['graph_edges'] is not None:
                l, b, r, t = s[0]['graph_edges']
                if l is not None and r is not None and l > r:
                    l, r = r, l
                if b is not None and t is not None and b > t:
                    b, t = t, b
                self.textCtrl_graph_left.SetValue(u'{:g}'.format(l) if l is not None else u'')
                self.textCtrl_graph_bottom.SetValue(u'{:g}'.format(b) if b is not None else u'')
                self.textCtrl_graph_right.SetValue(u'{:g}'.format(r) if r is not None else u'')
                self.textCtrl_graph_top.SetValue(u'{:g}'.format(t) if t is not None else u'')
            else:
                self.textCtrl_graph_left.SetValue(u'')
                self.textCtrl_graph_bottom.SetValue(u'')
                self.textCtrl_graph_right.SetValue(u'')
                self.textCtrl_graph_top.SetValue(u'')
        except:
            print(sys.exc_info())
        try:
            rows = len(self.grid_text.table.data)
            self.grid_text.table.data = []
            if 'texts' in s[0] and s[0]['texts'] is not None:
                for i in s[0]['texts']:
                    self.grid_text.table.data.append([i[0], i[1], i[2], i[3][0], i[4], i[5], i[6]])
            for i in range(len(self.grid_text.table.data), 3):
                self.grid_text.table.data.append(self.grid_text.table.new_data[:])
            self.grid_text.UpdateView(len(self.grid_text.table.data) - rows)
        except:
            print(sys.exc_info())
        for i, j in (('arrows', self.grid_arrow), ('ellipses', self.grid_ellipse),
            ('polygons', self.grid_polygon), ('naca4', self.grid_naca4)):
            try:
                rows = len(j.table.data)
                j.table.data = []
                if i in s[0] and s[0][i] is not None:
                    for k in s[0][i]:
                        j.table.data.append(list(k))
                for k in range(len(j.table.data), 3):
                    j.table.data.append(j.table.new_data[:])
                j.UpdateView(len(j.table.data) - rows)
            except:
                print(sys.exc_info())
        try:
            rows = len(self.grid_param_dict.table.data)
            self.grid_param_dict.table.data = []
            if 'param_dict' in s[0] and s[0]['param_dict'] is not None:
                for k, v in s[0]['param_dict'].items():
                    self.grid_param_dict.table.data.append([k, v])
            for i in range(len(self.grid_param_dict.table.data), 3):
                self.grid_param_dict.table.data.append(self.grid_param_dict.table.new_data[:])
            self.grid_param_dict.UpdateView(len(self.grid_param_dict.table.data) - rows)
        except:
            print(sys.exc_info())
        n = 0
        for x in s[1:]:
            if x['type'] == u'scatter' and ('file_name' in x and x['file_name'] is not None or
                    'equation' in x and x['equation'] is not None):
                if n == n_scatter:
                    continue
                try:
                    self.spinCtrls_zorder_scatter[n].SetValue(x['zorder'])
                except:
                    print(sys.exc_info())
                try:
                    self.filePickers_scatter[n].SetPath(
                        correct_file_name_in_unicode(os.path.join(dir_name, x['file_name'])))
                except:
                    try:
                        self.filePickers_scatter[n].SetPath(x['equation'])
                    except:
                        print(sys.exc_info())
                for i, j in ((self.textCtrls_scatter_column_x[n], 0), (self.textCtrls_scatter_column_y[n], 1),
                             (self.textCtrls_scatter_column_z[n], 2)):
                    try:
                        if type(x['columns'][j]) is str:
                            i.SetValue(u"'" + (x['columns'][j].decode('UTF-8')
                            if sys.version_info.major <= 2 else x['columns'][j]) + u"'")
                        else:
                            i.SetValue(u'{:d}'.format(x['columns'][j]))
                    except:
                        i.SetValue(u'')
                self.textCtrls_scatter_err_column_xn[n].SetValue(u'')
                self.textCtrls_scatter_err_column_xp[n].SetValue(u'')
                self.textCtrls_scatter_err_column_yn[n].SetValue(u'')
                self.textCtrls_scatter_err_column_yp[n].SetValue(u'')
                if 'err_columns' in x and x['err_columns'] is not None:
                    for i, j in enumerate(x['err_columns']):
                        try:
                            if type(j) is tuple:
                                j = [k for k in sorted(set(j), key = j.index) if k is not None]
                                if len(j) == 0:
                                    pass
                                elif len(j) == 1:
                                    if type(j[0]) is str:
                                        (self.textCtrls_scatter_err_column_xn[n] if i == 0
                                            else self.textCtrls_scatter_err_column_yn[n]).SetValue(
                                            u"'" + (j[0].decode('UTF-8') if sys.version_info.major <= 2 else j[0]) + u"'")
                                    else:
                                        (self.textCtrls_scatter_err_column_xn[n] if i == 0
                                            else self.textCtrls_scatter_err_column_yn[n]).SetValue(u'{:d}'.format(j[0]))
                                else:
                                    if type(j[0]) is str:
                                        (self.textCtrls_scatter_err_column_xn[n] if i == 0
                                            else self.textCtrls_scatter_err_column_yn[n]).SetValue(
                                            u"'" + (j[0].decode('UTF-8') if sys.version_info.major <= 2 else j[0]) + u"'")
                                    else:
                                        (self.textCtrls_scatter_err_column_xn[n] if i == 0
                                            else self.textCtrls_scatter_err_column_yn[n]).SetValue(u'{:d}'.format(j[0]))
                                    if type(j[1]) is str:
                                        (self.textCtrls_scatter_err_column_xp[n] if i == 0
                                            else self.textCtrls_scatter_err_column_yp[n]).SetValue(
                                            u"'" + (j[1].decode('UTF-8') if sys.version_info.major <= 2 else j[1]) + u"'")
                                    else:
                                        (self.textCtrls_scatter_err_column_xp[n] if i == 0
                                            else self.textCtrls_scatter_err_column_yp[n]).SetValue(u'{:d}'.format(j[1]))
                            elif j is not None:
                                if type(j) is str:
                                    (self.textCtrls_scatter_err_column_xn[n] if i == 0
                                        else self.textCtrls_scatter_err_column_yn[n]).SetValue(
                                        u"'" + (j.decode('UTF-8') if sys.version_info.major <= 2 else j) + u"'")
                                else:
                                    (self.textCtrls_scatter_err_column_xn[n] if i == 0
                                        else self.textCtrls_scatter_err_column_yn[n]).SetValue(u'{:d}'.format(j))
                        except:
                            print(sys.exc_info())
                try:
                    self.textCtrls_scatter_every[n].SetValue(u'{:d}'.format(x['every']))
                except:
                    self.textCtrls_scatter_every[n].SetValue(u'1')
                try:
                    self.textCtrls_scatter_label[n].SetValue(x['label_in_legend']
                        if 'label_in_legend' in x and x['label_in_legend'] is not None else u'')
                except:
                    print(sys.exc_info())
                try:
                    self.choices_scatter_marker[n].SetSelection(markers.index(x['marker']))
                except:
                    print(sys.exc_info())
                try:
                    self.textCtrls_scatter_marker_size[n].SetValue(u'{:g}'.format(x['marker_size'])
                        if 'marker_size' in x and x['marker_size'] is not None else u'')
                except:
                    print(sys.exc_info())
                self.checkBoxes_scatter_marker_edge[n].SetValue(True)
                for i, j in ((self.colourPickers_scatter_marker[n], 'marker_color'),
                             (self.colourPickers_scatter_marker_edge[n], 'marker_edge_color'),
                             (self.colourPickers_scatter_line[n], 'line_color')):
                    try:
                        if j in x:
                            if x[j] is None:
                                i.SetColour(wx.WHITE if j == 'marker_color' else wx.BLACK)
                            elif sys.version_info.major <= 2 and type(x[j]) is unicode or type(x[j]) is str:
                                if j == 'marker_edge_color' and x[j] == 'none':
                                    self.checkBoxes_scatter_marker_edge[n].SetValue(False)
                                else:
                                    i.SetColour(x[j])
                            else: # tuple, each element is in a range between 0 and 1
                                i.SetColour(wx.Colour(min(255, int(255*x[j][0])),
                                    min(255, int(255*x[j][1])), min(255, int(255*x[j][2]))))
                    except:
                        print(sys.exc_info())
                try:
                    self.choices_scatter_line_style[n].SetSelection(line_styles.index(x['line_style']))
                except:
                    print(sys.exc_info())
                try:
                    if 'show_z_by' not in x or x['show_z_by'] is None:
                        self.textCtrls_scatter_column_z[n].SetValue(u'')
                    else:
                        self.choices_show_z_by[n].SetSelection(by_what_show_z.index(x['show_z_by']))
                except:
                    print(sys.exc_info())
                try:
                    self.textCtrls_scatter_marker_size_ratio[n].SetValue(u'{:g}'.format(x['marker_size_ratio'])
                        if 'marker_size_ratio' in x and x['marker_size_ratio'] is not None else u'')
                except:
                    print(sys.exc_info())
                n += 1
            elif x['type'] == u'vector' and 'file_name' in x and x['file_name'] is not None:
                try:
                    self.spinCtrl_zorder_vector.SetValue(x['zorder'])
                except:
                    print(sys.exc_info())
                try:
                    self.filePicker_vector.SetPath(
                        correct_file_name_in_unicode(os.path.join(dir_name, x['file_name'])))
                except:
                    print(sys.exc_info())
                for i, j in ((self.textCtrl_vector_column_x, 0), (self.textCtrl_vector_column_y, 1),
                             (self.textCtrl_vector_column_u, 2), (self.textCtrl_vector_column_v, 3)):
                    try:
                        if type(x['columns'][j]) is str:
                            i.SetValue(u"'" + (x['columns'][j].decode('UTF-8')
                            if sys.version_info.major <= 2 else x['columns'][j]) + u"'")
                        else:
                            i.SetValue(u'{:d}'.format(x['columns'][j]))
                    except:
                        i.SetValue(u'')
                try:
                    self.textCtrl_vector_every.SetValue(u'{:d}'.format(x['every']))
                except:
                    self.textCtrl_vector_every.SetValue(u'1')
                try:
                    if x['color'] == u'len':
                        self.choice_vector_color.SetSelection(1)
                    elif x['color'] == u'len1':
                        self.choice_vector_color.SetSelection(2)
                    else:
                        self.choice_vector_color.SetSelection(0)
                        if x['color'] is None:
                            self.colourPicker_vector.SetColour(wx.BLACK)
                        elif sys.version_info.major <= 2 and type(x['color']) is unicode or type(x['color']) is str:
                            self.colourPicker_vector.SetColour(x['color'])
                        else: # tuple, each element is in a range between 0 and 1
                            self.colourPicker_vector.SetColour(wx.Colour(
                                min(255, int(255*x['color'][0])), min(255, int(255*x['color'][1])),
                                min(255, int(255*x['color'][2]))))
                except:
                    print(sys.exc_info())
                try:
                    self.textCtrl_vector_scale.SetValue(u'{:g}'.format(x['scale'])
                        if 'scale' in x and x['scale'] is not None else u'')
                except:
                    print(sys.exc_info())
                for i, j in ((self.textCtrl_vector_legend_x, 0), (self.textCtrl_vector_legend_y, 1),
                             (self.textCtrl_vector_legend_u, 2)):
                    try:
                        i.SetValue(u'{:g}'.format(x['legend'][j]) if 'legend' in x and x['legend'] is not None else u'')
                    except:
                        print(sys.exc_info())
                try:
                    self.textCtrl_vector_legend_label.SetValue(x['legend'][3]
                        if 'legend' in x and x['legend'] is not None else u'')
                except:
                    print(sys.exc_info())
            elif x['type'] == u'contour' and 'file_name' in x and x['file_name'] is not None:
                try:
                    self.spinCtrl_zorder_contour.SetValue(x['zorder'])
                except:
                    print(sys.exc_info())
                try:
                    self.filePicker_contour.SetPath(correct_file_name_in_unicode(os.path.join(dir_name, x['file_name'])))
                except:
                    print(sys.exc_info())
                self.checkBox_contour_grid_pattern.SetValue((True if x['structured_grid'] is not None else False)
                    if 'structured_grid' in x else False)
                self.checkBox_contour_show_triangle.SetValue(x['show_triangle'] if 'show_triangle' in x else False)
                for i, j in ((self.textCtrl_contour_column_x, 0), (self.textCtrl_contour_column_y, 1),
                             (self.textCtrl_contour_column_z, 2)):
                    try:
                        if type(x['columns'][j]) is str:
                            i.SetValue(u"'" + (x['columns'][j].decode('UTF-8')
                            if sys.version_info.major <= 2 else x['columns'][j]) + u"'")
                        else:
                            i.SetValue(u'{:d}'.format(x['columns'][j]))
                    except:
                        self.i.SetValue(u'')
                try:
                    self.textCtrl_contour_every.SetValue(u'{:d}'.format(x['every']))
                except:
                    self.textCtrl_contour_every.SetValue(u'1')
                try:
                    self.choice_contour_paint.SetSelection(paint_styles.index(x['paint']))
                except:
                    print(sys.exc_info())
                self.checkBox_contour_smooth_paint.SetValue(x['smooth_paint'] if 'smooth_paint' in x else True)
        if path != self.backup_path:
            self.filePicker_setting.SetPath(path)
        self.cwd = dir_name
        for i in self.filePickers:
            i.SetInitialDirectory(dir_name)

    def menuItem_openOnMenuSelection(self, event):
        with wx.FileDialog(self, _(u'設定ファイルを開く'), style = wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fd:
            fd.SetDirectory(self.cwd)
            if fd.ShowModal() == wx.ID_CANCEL:
                return
            self.load_settings(fd.GetPath())

    def MakeSettingStrings(self, path):
        path = correct_file_name_in_unicode(path) # unicode
        try:
            s = u"{\n    'type': 'global',\n"
            v = correct_file_name_in_unicode(self.filePicker_png.GetPath()) # unicode
            if v == u'':
                s += u"    'png_file_name': None,\n"
            else:
                if not v.endswith(u'.png'):
                    v += u'.png'
                self.filePicker_png.SetPath(v)
                v = correct_file_name_in_unicode(os.path.relpath(v, start = os.path.dirname(path)))
                s += u"    'png_file_name': '{}',\n".format(v.replace("\\", r"\\").replace("'", r"\'"))
            try:
                s += u"    'fig_size': ({:g}, {:g}),\n".format(
                    float(self.textCtrl_fig_size_w.GetValue()), float(self.textCtrl_fig_size_h.GetValue()))
            except:
                s += u"    'fig_size': None,\n"
            for i, j in ((self.textCtrl_skip, u'skip'), (self.textCtrl_delimiter, u'delimiter'),
                         (self.textCtrl_terminator, u'terminator')):
                v = i.GetValue() # unicode
                if v == u'':
                    s += u"    '{}': None,\n".format(j)
                else:
                    s += u"    '{}': '{}',\n".format(j, v.replace("\\", r"\\").replace("'", r"\'"))
            s += u"    'ranges': ("
            for i, j, k in (
                (self.grid_graph.table.ROW_X, self.grid_graph.table.COL_FROM, self.grid_graph.table.COL_TO),
                (self.grid_graph.table.ROW_Y, self.grid_graph.table.COL_FROM, self.grid_graph.table.COL_TO),
                (self.grid_graph.table.ROW_Z, self.grid_graph.table.COL_FROM, self.grid_graph.table.COL_TO)
            ):
                try:
                    s += u"({:g}, ".format(self.grid_graph.table.data[i][j])
                    try:
                        s += u"{:g}), ".format(self.grid_graph.table.data[i][k])
                    except:
                        s += u"None), "
                except:
                    try:
                        s += u"(None, {:g}), ".format(self.grid_graph.table.data[i][k])
                    except:
                        s += u"None, "
            s = s[:-2] + u"),\n"
            s += u"    'ticks': ("
            for i in (self.grid_graph.table.ROW_X, self.grid_graph.table.ROW_Y, self.grid_graph.table.ROW_Z):
                try:
                    s += u"{:g}, ".format(self.grid_graph.table.data[i][self.grid_graph.table.COL_TICKS])
                except:
                    s += u"None, "
            s = s[:-2] + u"),\n"
            s += u"    'log_scale': " + str((
                self.grid_graph.table.data[self.grid_graph.table.ROW_X][self.grid_graph.table.COL_LOG_SCALE],
                self.grid_graph.table.data[self.grid_graph.table.ROW_Y][self.grid_graph.table.COL_LOG_SCALE],
                self.grid_graph.table.data[self.grid_graph.table.ROW_Z][self.grid_graph.table.COL_LOG_SCALE]
            )) + u",\n"
            s += u"    'labels': ("
            for i in (self.grid_graph.table.ROW_X, self.grid_graph.table.ROW_Y, self.grid_graph.table.ROW_Z):
                if self.grid_graph.table.data[i][self.grid_graph.table.COL_LABEL] is None:
                    s += u"None, "
                else:
                    s += u"'{}', ".format(self.grid_graph.table.data[i][
                        self.grid_graph.table.COL_LABEL].replace("\\", r"\\").replace("'", r"\'"))
            s = s[:-2] + u"),\n"
            s += (u"    'grids': ({}, {}),\n").format(str(self.checkBox_x_grid.GetValue()), str(self.checkBox_y_grid.GetValue()))
            v = self.comboBox_aspect.GetValue() # unicode
            if v == u'':
                s += u"    'aspect': 'auto',\n"
            elif v in (u'auto', u'equal'):
                s += u"    'aspect': '{}',\n".format(v)
            else:
                try:
                    s += u"    'aspect': {:g},\n".format(float(v))
                except:
                    s += u"    'aspect': 'auto',\n"
            v = self.textCtrl_title.GetValue() # unicode
            if v == u'':
                s += u"    'title': None,\n"
            else:
                s += u"    'title': '{}',\n".format(v.replace("\\", r"\\").replace("'", r"\'"))
            x = self.textCtrl_legend_left.GetValue() # unicode
            y = self.textCtrl_legend_top.GetValue() # unicode
            if x == u''and y == u'':
                s += u"    'legend_location': None,\n"
            else:
                s += u"    'legend_location': ({}, {}),\n".format(
                    u'None' if x == u'' else u'{:g}'.format(float(x)), u'None' if y == u'' else u'{:g}'.format(float(y)))
            l = self.textCtrl_graph_left.GetValue() # unicode
            b = self.textCtrl_graph_bottom.GetValue() # unicode
            r = self.textCtrl_graph_right.GetValue() # unicode
            t = self.textCtrl_graph_top.GetValue() # unicode
            if l == u'' and b == u''and r == u'' and t == u'':
                s += u"    'graph_edges': None,\n"
            else:
                if l != u'' and r != u'' and float(l) > float(r):
                    l, r = r, l
                if b != u'' and t != u'' and float(b) > float(t):
                    b, t = t, b
                l = u'None' if l == u'' else u'{:g}'.format(float(l))
                b = u'None' if b == u'' else u'{:g}'.format(float(b))
                r = u'None' if r == u'' else u'{:g}'.format(float(r))
                t = u'None' if t == u'' else u'{:g}'.format(float(t))
                s += u"    'graph_edges': ({}, {}, {}, {}),\n".format(l, b, r, t)
            v = []
            for i in self.grid_text.table.data:
                if None not in i:
                    v.append(i)
            if len(v) == 0:
                s += u"    'texts': None,\n"
            else:
                s += u"    'texts': ("
                for i in v:
                    s += u"('{}', {:g}, {:g}, '{}', '{}', {:g}, {}), ".format(
                        i[0], i[1], i[2], i[3], i[4].replace("\\", r"\\").replace("'", r"\'"), i[5], i[6])
                s = s[:-1] + u"),\n"
            v = []
            for i in self.grid_arrow.table.data:
                if None not in i:
                    v.append(i)
            if len(v) == 0:
                s += u"    'arrows': None,\n"
            else:
                s += u"    'arrows': ("
                for i in v:
                    s += u"('{}', {:g}, {:g}, {:g}, {:g}, {:g}, '{}', {:g}, {}), ".format(*i)
                s = s[:-1] + u"),\n"
            v = []
            for i in self.grid_ellipse.table.data:
                if None not in i:
                    v.append(i)
            if len(v) == 0:
                s += u"    'ellipses': None,\n"
            else:
                s += u"    'ellipses': ("
                for i in v:
                    s += u"('{}', {:g}, {:g}, {:g}, {:g}, {:g}, {:g}, ".format(*i[:self.grid_ellipse.table.COL_FILL])
                    s += str(i[self.grid_ellipse.table.COL_FILL])
                    s += ", {}), ".format(i[self.grid_ellipse.table.COL_ZORDER])
                s = s[:-1] + u"),\n"
            v = []
            for i in self.grid_polygon.table.data:
                if None not in i:
                    v.append(i)
            if len(v) == 0:
                s += u"    'polygons': None,\n"
            else:
                s += u"    'polygons': ("
                for i in v:
                    s += u"('{}', [".format(i[self.grid_polygon.table.COL_COORDINATE])
                    for j in i[self.grid_polygon.table.COL_XY]:
                        s += u"{:g}, {:g}, ".format(j[0], j[1])
                    s = s[:-2]
                    s += "], {:g}, ".format(i[self.grid_polygon.table.COL_LINE_WIDTH])
                    s += str(i[self.grid_polygon.table.COL_FILL])
                    s += u", {}), ".format(i[self.grid_polygon.table.COL_ZORDER])
                s = s[:-1] + u"),\n"
            v = []
            for i in self.grid_naca4.table.data:
                if None not in i:
                    v.append(i)
            if len(v) == 0:
                s += u"    'naca4': None,\n"
            else:
                s += u"    'naca4': ("
                for i in v:
                    s += u"('{}', '{}', {:g}, {:g}, {:g}, {:g}, {:g}, ".format(*i[:self.grid_naca4.table.COL_FILL])
                    s += str(i[self.grid_naca4.table.COL_FILL])
                    s += u", {}), ".format(i[self.grid_naca4.table.COL_ZORDER])
                s = s[:-1] + u"),\n"
            v = []
            for i in self.grid_param_dict.table.data:
                if None not in i:
                    v.append(i)
            if len(v) == 0:
                s += u"    'param_dict': None,\n"
            else:
                s += u"    'param_dict': {"
                for i in v:
                    s += u"'{}': {:g}, ".format(i[0], i[1])
                s = s[:-2] + u"},\n"
            s += "},\n"
            for n in range(n_scatter):
                fn = self.filePickers_scatter[n].GetPath()
                if fn == u'':
                    continue
                s += u"{\n    'type': 'scatter',\n"
                s += u"    'zorder': {},\n".format(self.spinCtrls_zorder_scatter[n].GetValue())
                eq = self.filePickers_scatter[n].GetTextCtrl().GetValue()
                file_plot = True if pat_eq_plot.match(eq) is None else False
                if file_plot:
                    fn = correct_file_name_in_unicode(os.path.relpath(fn, start = os.path.dirname(path)))
                    s += u"    'file_name': '{}',\n".format(fn.replace("\\", r"\\").replace("'", r"\'"))
                    s += u"    'columns': ("
                    for i in (self.textCtrls_scatter_column_x[n].GetValue(),
                            self.textCtrls_scatter_column_y[n].GetValue(),
                            self.textCtrls_scatter_column_z[n].GetValue()): # unicode
                        if i == u'':
                            s += u"None, "
                        elif fn.endswith(u'.csv') or fn.endswith(u'.xls') or fn.endswith(u'.xlsx'):
                            s += u"'{}', ".format(i.strip(u"'"))
                        else:
                            try:
                                s += u"{}, ".format(int(i))
                            except:
                                s += u"'{}', ".format(i.strip(u"'"))
                    s = s[:-8 if s.endswith(u', None, ') else -2] + u"),\n"
                    s += u"    'err_columns': ("
                    for v in ((self.textCtrls_scatter_err_column_xn[n].GetValue(),
                            self.textCtrls_scatter_err_column_xp[n].GetValue()),
                            (self.textCtrls_scatter_err_column_yn[n].GetValue(),
                            self.textCtrls_scatter_err_column_yp[n].GetValue())): # unicode
                        v = [i for i in sorted(set(v), key = v.index) if i != u'']
                        if len(v) == 0:
                            s += u"None, "
                        elif len(v) == 1:
                            if fn.endswith(u'.csv') or fn.endswith(u'.xls') or fn.endswith(u'.xlsx'):
                                s += u"'{}', ".format(v[0].strip(u"'"))
                            else:
                                try:
                                    s += u"{}, ".format(int(v[0]))
                                except:
                                    s += u"'{}', ".format(v[0].strip(u"'"))
                        else:
                            if fn.endswith(u'.csv') or fn.endswith(u'.xls') or fn.endswith(u'.xlsx'):
                                s += u"('{}', '{}'), ".format(v[0].strip(u"'"), v[1].strip(u"'"))
                            else:
                                try:
                                    s += u"({}, ".format(int(v[0]))
                                except:
                                    s += u"('{}', ".format(v[0].strip(u"'"))
                                try:
                                    s += u"{}), ".format(int(v[1]))
                                except:
                                    s += u"'{}'), ".format(v[1].strip(u"'"))
                    s = s[:-2] + u"),\n"
                    try:
                        s += u"    'every': {},\n".format(int(self.textCtrls_scatter_every[n].GetValue()))
                    except:
                        s += u"    'every': 1,\n"
                else: # not file_plot
                    s += u"    'equation': '{}',\n".format(eq)
                v = self.choices_scatter_marker[n].GetSelection()
                if v == len(markers) - 1:
                    s += u"    'marker': None,\n"
                else:
                    s += u"    'marker': '{}',\n".format(markers[v])
                try:
                    s += u"    'marker_size': {:g},\n".format(float(self.textCtrls_scatter_marker_size[n].GetValue()))
                except:
                    s += u"    'marker_size': None,\n"
                v = self.colourPickers_scatter_marker[n].GetColour()
                s += u"    'marker_color': ({:g}, {:g}, {:g}),\n".format(v.Red()/255.0, v.Green()/255.0, v.Blue()/255.0)
                if self.checkBoxes_scatter_marker_edge[n].GetValue():
                    v = self.colourPickers_scatter_marker_edge[n].GetColour()
                    s += u"    'marker_edge_color': ({:g}, {:g}, {:g}),\n".format(v.Red()/255.0, v.Green()/255.0, v.Blue()/255.0)
                else:
                    s += u"    'marker_edge_color': 'none',\n" # not 'None', but 'none'
                v = self.choices_scatter_line_style[n].GetSelection()
                if v == len(line_styles) - 1:
                    s += u"    'line_style': None,\n"
                else:
                    s += u"    'line_style': '{}',\n".format(line_styles[v])
                v = self.colourPickers_scatter_line[n].GetColour()
                s += u"    'line_color': ({:g}, {:g}, {:g}),\n".format(v.Red()/255.0, v.Green()/255.0, v.Blue()/255.0)
                if self.textCtrls_scatter_column_z[n].GetValue() == u'':
                    s += u"    'show_z_by': None,\n"
                else:
                    s += u"    'show_z_by': '{}',\n".format(by_what_show_z[self.choices_show_z_by[n].GetSelection()])
                try:
                    s += u"    'marker_size_ratio': {:g},\n".format(float(self.textCtrls_scatter_marker_size_ratio[n].GetValue()))
                except:
                    s += u"    'marker_size_ratio': None,\n"
                v = self.textCtrls_scatter_label[n].GetValue() # unicode
                if v == u'':
                    s += u"    'label_in_legend': None,\n"
                else:
                    s += u"    'label_in_legend': '{}',\n".format(v.replace("\\", r"\\").replace("'", r"\'"))
                s += u"},\n"
            fn = self.filePicker_vector.GetPath()
            if fn != u'':
                s += u"{\n    'type': 'vector',\n"
                s += u"    'zorder': {},\n".format(self.spinCtrl_zorder_vector.GetValue())
                eq = self.filePicker_vector.GetTextCtrl().GetValue()
                file_plot = True if pat_eq_plot.match(eq) is None else False
                if file_plot:
                    fn = correct_file_name_in_unicode(
                        os.path.relpath(fn, start = os.path.dirname(path)))
                    s += u"    'file_name': '{}',\n".format(fn.replace("\\", r"\\").replace("'", r"\'"))
                    s += u"    'columns': ("
                    for i in (self.textCtrl_vector_column_x.GetValue(),
                            self.textCtrl_vector_column_y.GetValue(),
                            self.textCtrl_vector_column_u.GetValue(),
                            self.textCtrl_vector_column_v.GetValue()): # unicode
                        if i == u'':
                            s += u"None, "
                        elif fn.endswith(u'.csv') or fn.endswith(u'.xls') or fn.endswith(u'.xlsx'):
                            s += u"'{}', ".format(i.strip(u"'"))
                        else:
                            try:
                                s += u"{}, ".format(int(i))
                            except:
                                s += u"'{}', ".format(i.strip(u"'"))
                    s = s[:-2] + u"),\n"
                    try:
                        s += u"    'every': {},\n".format(int(self.textCtrl_vector_every.GetValue()))
                    except:
                        s += u"    'every': 1,\n"
                else: # not file_plot
                    s += u"    'equation': '{}',\n".format(eq)
                v = self.choice_vector_color.GetSelection()
                if v == 0:
                    v = self.colourPicker_vector.GetColour()
                    s += u"    'color': ({:g}, {:g}, {:g}),\n".format(v.Red()/255.0, v.Green()/255.0, v.Blue()/255.0)
                elif v == 1:
                    s += u"    'color': 'len',\n"
                else:
                    s += u"    'color': 'len1',\n"
                try:
                    s += u"    'scale': {:g},\n".format(float(self.textCtrl_vector_scale.GetValue()))
                except:
                    s += u"    'scale': 1.0,\n"
                try:
                    v = self.textCtrl_vector_legend_label.GetValue()
                    v = 'None' if v == u'' else "'{}'".format(v.replace("\\", r"\\").replace("'", r"\'"))
                    s += u"    'legend': ({:g}, {:g}, {:g}, {}),\n".format(
                        float(self.textCtrl_vector_legend_x.GetValue()),
                        float(self.textCtrl_vector_legend_y.GetValue()),
                        float(self.textCtrl_vector_legend_u.GetValue()), v)
                except:
                    s += u"    'legend': None,\n"
                s += u"},\n"
            fn = self.filePicker_contour.GetPath()
            if fn != u'':
                s += u"{\n    'type': 'contour',\n"
                s += u"    'zorder': {},\n".format(self.spinCtrl_zorder_contour.GetValue())
                eq = self.filePicker_contour.GetTextCtrl().GetValue()
                file_plot = True if pat_eq_plot.match(eq) is None else False
                if file_plot:
                    fn = correct_file_name_in_unicode(
                        os.path.relpath(fn, start = os.path.dirname(path)))
                    s += u"    'file_name': '{}',\n".format(fn.replace("\\", r"\\").replace("'", r"\'"))
                else: # not file_plot
                    s += u"    'equation': '{}',\n".format(eq)
                s += u"    'grid_pattern': " + str(self.checkBox_contour_grid_pattern.GetValue()) + u",\n"
                s += u"    'show_triangle': " + str(self.checkBox_contour_show_triangle.GetValue()) + u",\n"
                if file_plot:
                    s += u"    'columns': ("
                    for i in (self.textCtrl_contour_column_x.GetValue(),
                            self.textCtrl_contour_column_y.GetValue(),
                            self.textCtrl_contour_column_z.GetValue()): # unicode
                        if i == u'':
                            s += u"None, "
                        elif fn.endswith(u'.csv') or fn.endswith(u'.xls') or fn.endswith(u'.xlsx'):
                            s += u"'{}', ".format(i.strip(u"'"))
                        else:
                            try:
                                s += u"{}, ".format(int(i))
                            except:
                                s += u"'{}', ".format(i.strip(u"'"))
                    s = s[:-2] + u"),\n"
                    try:
                        s += u"    'every': {},\n".format(int(self.textCtrl_contour_every.GetValue()))
                    except:
                        s += u"    'every': 1,\n"
                v = self.choice_contour_paint.GetSelection()
                if v == len(paint_styles) - 1:
                    s += u"    'paint': None,\n"
                else:
                    s += u"    'paint': '{}',\n".format(paint_styles[v])
                s += u"    'smooth_paint': " + str(self.checkBox_contour_smooth_paint.GetValue()) + u",\n"
                s += u"},\n"
        except:
            print(sys.exc_info())
        return s, path

    def SaveSettings(self, path):
        s, path = self.MakeSettingStrings(path)
        with codecs.open(path, 'w', encoding = 'UTF-8') as f:
            f.write(s)
        if path != self.backup_path:
            self.filePicker_setting.SetPath(path)

    def menuItem_saveOnMenuSelection(self, event):
        if self.filePicker_setting.GetPath() == u'':
            self.menuItem_save_asOnMenuSelection(event)
        else:
            self.SaveSettings(self.filePicker_setting.GetPath())

    def menuItem_save_asOnMenuSelection(self, event):
        with wx.FileDialog(self, _(u'設定ファイルを保存'), wildcard = u'Text files (*.txt)|*.txt',
            style = wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fd:
            fd.SetDirectory(self.cwd)
            fd.SetFilename(u'settings_matplotlib.txt')
            if fd.ShowModal() == wx.ID_CANCEL:
                return
            self.SaveSettings(fd.GetPath())

    def menuItem_save_as_scriptOnMenuSelection(self, event):
        with wx.FileDialog(self, _(u'pythonスクリプト形式で保存'), wildcard = u'python script files (*.py)|*.py',
            style = wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fd:
            fd.SetDirectory(self.cwd)
            fd.SetFilename(u'script_matplotlib.py')
            if fd.ShowModal() == wx.ID_CANCEL:
                return
            p = fd.GetPath() # unicode
        s, p = self.MakeSettingStrings(p)
        make_script(s, p)

    def menuItem_make_templateOnMenuSelection(self, event):
        with wx.FileDialog(self, _(u'テンプレートを作る'), wildcard = u'Text files (*.txt)|*.txt',
            style = wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fd:
            fd.SetDirectory(self.cwd)
            fd.SetFilename(u'template.txt')
            if fd.ShowModal() == wx.ID_CANCEL:
                return
            make_template(fd.GetPath())

    def menuItem_cutOnMenuSelection(self, event):
        try:
            self.FindFocus().Cut()
        except:
            pass

    def menuItem_copyOnMenuSelection(self, event):
        try:
            self.FindFocus().Copy()
        except:
            pass

    def menuItem_pasteOnMenuSelection(self, event):
        try:
            self.FindFocus().Paste()
        except:
            pass

    def menuItem_plotOnMenuSelection(self, event):
        self.SaveSettings(self.backup_path)
        try:
            plot(load_plot_settings(self.backup_path)[0], show = True)
        except:
            print(sys.exc_info())

    def menuItem_plot_directOnMenuSelection(self, event):
        with wx.FileDialog(self, _(u'設定ファイルから直接プロット'), style = wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fd:
            fd.SetDirectory(self.cwd)
            if fd.ShowModal() == wx.ID_CANCEL:
                return
            try:
                plot(load_plot_settings(fd.GetPath())[0], show = True)
            except:
                print(sys.exc_info())

    def menuItem_updateOnMenuSelection(self, event):
#       s = get_file_from_google_drive('18jsRVeShgjhvKi_X6m9z-JQKlHkedDnT')
        s = get_file_from_github_public(user = 'gitwamoto', repository = 'matplotlibwx',
            branch = 'main', file_path = 'matplotlibwx.py')
        if s is None:
            with wx.MessageDialog(self,
                _(u'GitHubに接続できませんでした．後でやり直して下さい．'),
                _(u'接続エラー'), style = wx.ICON_ERROR) as md:
                md.ShowModal()
            return
        r = re.search(b"version\\s*=\\s*'([0-9/ :APM]+)'\n", s[0])
        if r is not None and time_str_a_is_newer_than_b(a = r[1].decode(s[1]), b = version):
            p = correct_file_name_in_unicode(os.path.realpath(decode_if_necessary(__file__)))
            with open(p, 'wb') as f:
                f.write(s[0])
            pd = os.path.dirname(p)
            d = os.path.join(pd, u'locale', u'en', u'LC_MESSAGES')
            if not os.path.isdir(d):
                os.makedirs(d)
            s = get_file_from_github_public(user = 'gitwamoto', repository = 'matplotlibwx',
                branch = 'main', file_path = 'locale/en/LC_MESSAGES/messages.mo')
            if s is not None:
                with open(os.path.join(d, u'messages.mo'), 'wb') as f:
                    f.write(s[0])
            s = get_file_from_github_public(user = 'gitwamoto', repository = 'matplotlibwx',
                branch = 'main', file_path = 'locale/en/LC_MESSAGES/messages.po')
            if s is not None:
                with open(os.path.join(d, u'messages.po'), 'wb') as f:
                    f.write(s[0])
            s = get_file_from_github_public(user = 'gitwamoto', repository = 'matplotlibwx',
                branch = 'main', file_path = 'locale/messages.pot')
            if s is not None:
                with open(os.path.join(pd, u'locale', u'messages.pot'), 'wb') as f:
                    f.write(s[0])
            s = get_file_from_github_public(user = 'gitwamoto', repository = 'matplotlibwx',
                branch = 'main', file_path = 'README.md')
            if s is not None:
                with open(os.path.join(pd, u'README.md'), 'wb') as f:
                    f.write(s[0])
            if os.path.isfile(os.path.join(pd, u'modules_needed.txt')):
                os.remove(os.path.join(pd, u'modules_needed.txt'))
            if sys.platform != 'darwin':
                for curDir, dirs, files in os.walk(os.path.dirname(p)):
                    for name in files:
                        if re.match('\\._.+|[._]+DS_Store', name):
                            os.remove(os.path.join(curDir, name))
            with wx.MessageDialog(self, _(u'アップデートされました．再起動します．'),
                _(u'アップデート完了'), style = wx.ICON_INFORMATION) as md:
                md.ShowModal()
                os.execv(sys.executable, ['python', __file__])
        else:
            with wx.MessageDialog(self, _(u'アップデートの必要はありません．'),
                _(u'プログラムは最新です．'), style = wx.ICON_INFORMATION) as md:
                md.ShowModal()

    def menuItem_movieOnMenuSelection(self, event):
        webbrowser.open(url = 'https://youtu.be/WKnJaJfyNww')

if __name__ == '__main__':
    if len(sys.argv) == 1:
        app = wx.App()
        frame_main = FrameMain(None)
        frame_main.Show()
        app.MainLoop()
    else:
        if sys.argv[1] == '-h':
            print(u'Usage: %s {<setting_file_name> | -t [<template_file_name>]}' %
                os.path.basename(decode_if_necessary(sys.argv[0])))
        elif sys.argv[1] == '-t':
            t = sys.argv[2] if len(sys.argv) > 2 else 'template.txt'
            t = make_template(decode_if_necessary(t))
            print(u'Template file \'%s\' has be made.' % t)
            with open('bbb_b.txt', 'wb') as f:
                f.write(b) # mac + python 2 -> success, mac + python 3 -> ?
        else:
            plot(load_plot_settings(os.path.abspath(decode_if_necessary(sys.argv[1])))[0], show = True)
